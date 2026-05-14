"""Parse a 38DN Internal Pricing Bible xlsx into a `Bible` record.

Usage
-----
    from lib.bible_loader import load_bible_from_excel
    bible = load_bible_from_excel(
        Path("Desktop/38DN Internal Pricing Bible_H1-2026_vf_*.xlsx"),
        label="H1-2026 Perch May Update",
    )

Strategy (MVP scope)
--------------------
The loader **overlays values** from the uploaded Excel onto the bundled
Q1'26 vintage. Tolerances, labels, units, and the model-row mapping all
stay code-curated — they don't change across bible vintages. The Excel
is the source of *values*, not the source of *audit structure*.

What's overlaid (Phase 4 MVP):
  - CS tab column H (Average scenario) for ~13 line items that map to
    model rows in `cs_average`

What's NOT overlaid (deferred to a follow-up):
  - Size-dependent EPC override (CS!H11 vs CS!I11) — the <5 MWdc value
    stays hardcoded in `lib/audit/context.py` for now
  - Market Specific Assumptions tab — state/utility/program structures
    are heterogeneous and need per-state parsing logic
  - cs_tax_equity, cs_perm_debt_*, cs_construction_loan informational
    sections — not currently consumed by the audit pipeline

If a mapped Bible row is unreadable (missing cell, non-numeric value,
sentinel like "See Market Specific Assumptions"), the loader keeps the
bundled value and records a warning in the returned `Bible.source` note.
"""

from __future__ import annotations

import logging
from datetime import UTC, datetime
from pathlib import Path
from typing import BinaryIO

import openpyxl

from lib.audit.bible import Bible, vintage_id_from_upload
from lib.utils import safe_float

logger = logging.getLogger(__name__)


# Bible CS-tab row → model row mapping.
# Hand-curated to match the existing `lib.bible_reference.CS_AVERAGE`
# entries. Adding a new row here = adding it to CS_AVERAGE in code first.
#
# Reference: `Claude-Work/About-Me/pricing-bible-source-mapping.md` §
# "CS tab → Pricing Model Project Inputs row map (col H values)".
CS_TAB_ROW_MAP: dict[int, int] = {
    9: 123,  # Closing & Legal Costs → Closing and Legal
    11: 118,  # EPC ($/W) → PV EPC Cost (>5 MWdc — col H)
    12: 119,  # LNTP ($/W) → PV LNTP Cost
    35: 225,  # PV O&M Preventative → PV O&M Preventative
    36: 226,  # PV O&M Corrective → PV O&M Corrective
    37: 227,  # PV O&M Esc → PV O&M Esc
    38: 230,  # AM for Financing → AM Fee (financing)
    39: 231,  # AM for Financing Esc → AM Esc
    40: 302,  # AM - Internally Incurred → Internal AM Costs
    42: 296,  # Insurance → P&C Insurance
    43: 297,  # Insurance Esc → P&C Insurance Esc
    45: 241,  # Customer Management Esc → Customer Mgmt Esc
    48: 286,  # Decom Bond Annual Premium → Decom Annual Premium
}

# The Average scenario column in the CS tab.
# See `pricing-bible-source-mapping.md` § "CS tab column layout".
AVERAGE_COL_IDX = 8  # column "H"


class BibleParseError(Exception):
    """Raised when the Excel doesn't look like a valid Pricing Bible.

    Examples
    --------
      - "CS" worksheet missing
      - All mapped rows return None (probably wrong file)
      - File can't be opened by openpyxl
    """


def load_bible_from_excel(
    source: Path | BinaryIO,
    label: str = "",
    *,
    filename: str | None = None,
) -> Bible:
    """Parse a Pricing Bible xlsx into a `Bible` vintage record.

    Args
    ----
    source
        Path to the xlsx, or an open binary file-like object (used by
        the API upload endpoint).
    label
        Human-readable vintage label (e.g., "H1-2026 Perch May Update").
        Falls back to the filename if empty.
    filename
        Optional original filename when `source` is a BytesIO. Used in
        the `Bible.source` field for traceability.

    Returns
    -------
    A `Bible` instance with `vintage_id` time-stamped, `source` set to
    the filename, and `uploaded_at` set to now (UTC).

    Raises
    ------
    BibleParseError if the workbook doesn't have a "CS" tab or all
    mapped values are unreadable.
    """
    src_name = _source_name(source, filename)

    try:
        wb = openpyxl.load_workbook(
            source,
            data_only=True,
            read_only=True,
        )
    except Exception as e:  # openpyxl raises a grab-bag of types
        raise BibleParseError(f"Could not open {src_name!r} as an xlsx workbook: {e}") from e

    if "CS" not in wb.sheetnames:
        raise BibleParseError(
            f"{src_name!r} has no 'CS' worksheet — expected the 38DN Internal "
            f"Pricing Bible layout (CS tab with Average column at H). Found "
            f"sheets: {wb.sheetnames}"
        )

    ws = wb["CS"]

    # Start from the bundled record so tolerances + labels + state
    # overrides + market entries + benchmarks are all preserved.
    bible = Bible.bundled_q1_2026()

    # Mutate vintage metadata to identify as uploaded.
    bible.vintage_id = vintage_id_from_upload(src_name)
    bible.label = label or src_name or "Uploaded Bible"
    bible.source = src_name
    bible.uploaded_at = datetime.now(UTC).isoformat()

    overlaid: list[int] = []
    skipped: list[tuple[int, str]] = []

    for bible_row, model_row in CS_TAB_ROW_MAP.items():
        raw = ws.cell(row=bible_row, column=AVERAGE_COL_IDX).value
        if raw is None:
            skipped.append((bible_row, "cell blank"))
            continue
        # The Average column carries either a number or a sentinel string
        # like "See Market Specific Assumptions". Only the numbers are
        # bible benchmarks; sentinel strings mean "look elsewhere".
        if isinstance(raw, str):
            skipped.append((bible_row, f"sentinel: {raw!r}"))
            continue
        value = safe_float(raw)
        if value is None:
            skipped.append((bible_row, f"non-numeric: {raw!r}"))
            continue
        if model_row not in bible.cs_average:
            skipped.append((bible_row, f"model row {model_row} not in cs_average"))
            continue
        # Preserve int representation when the value is integer-valued.
        # Matches the bundled Python literals (e.g., `4750` for O&M, not
        # `4750.0` from safe_float) so audit-output snapshots stay stable
        # across bundled vs upload-loaded vintages.
        typed_value: float | int = int(value) if value == int(value) else value
        bible.cs_average[model_row]["value"] = typed_value
        overlaid.append(bible_row)

    wb.close()

    if not overlaid:
        raise BibleParseError(
            f"{src_name!r} matched the CS-tab layout but no numeric values "
            f"could be overlaid. Skipped: {skipped[:5]}"
        )

    logger.info(
        "Loaded bible vintage %s from %s — overlaid %d row(s), skipped %d",
        bible.vintage_id,
        src_name,
        len(overlaid),
        len(skipped),
    )
    if skipped:
        logger.debug("Skipped Bible rows: %s", skipped)

    return bible


def _source_name(source: Path | BinaryIO, filename: str | None) -> str:
    """Best-effort display name for the source argument."""
    if isinstance(source, Path):
        return source.name
    if filename:
        return filename
    name_attr = getattr(source, "name", None)
    return name_attr if isinstance(name_attr, str) else "<stream>"
