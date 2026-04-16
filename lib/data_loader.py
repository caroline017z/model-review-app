"""
38DN Pricing Model Review — Data Loading
Reads pricing model workbooks and extracts project data.
Pure Python — no Streamlit dependency.
"""
import copy
import logging
import openpyxl
from pathlib import Path
import re

logger = logging.getLogger(__name__)

from lib.config import (
    INPUT_ROW_LABELS, OUTPUT_ROWS, TEXT_ROWS,
    RATE_COMP_STARTS, EQUITY_RATE_TOGGLE_START,
    DEBT_RATE_TOGGLE_START, APPRAISAL_RATE_TOGGLE_START,
)
from lib.utils import safe_float


# ---------------------------------------------------------------------------
# Dynamic label-based row mapping
# ---------------------------------------------------------------------------

def _normalize_label(s):
    """Normalize a label for fuzzy matching: lowercase, strip, collapse whitespace, remove punctuation noise."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r'\s+', ' ', s)
    return s


_UNIT_SUFFIXES_RE = re.compile(r"\$/(kWh|MWh|MW[\s-]*dc|MW|W|kW[\s-]*yr)", re.IGNORECASE)


def _labels_match(canonical, actual):
    """Check if two normalized labels are equivalent.

    Tightened: we only accept substring matches when the shorter label is
    AT LEAST half the length of the longer one. This prevents false
    positives like "Customer" matching "Customer Mgmt Esc".

    Unit-aware: if both labels contain a unit suffix ($/W, $/kWh, etc.)
    and they differ, the match is rejected immediately — prevents mapping
    a $/kWh row to a $/W canonical or vice versa.
    """
    if canonical == actual:
        return True

    # Unit-suffix check: reject if both carry different unit tokens.
    c_units = _UNIT_SUFFIXES_RE.findall(canonical)
    a_units = _UNIT_SUFFIXES_RE.findall(actual)
    if c_units and a_units:
        # Normalize: lowercase, strip whitespace/hyphens for comparison
        c_norm = {u.lower().replace(" ", "").replace("-", "") for u in c_units}
        a_norm = {u.lower().replace(" ", "").replace("-", "") for u in a_units}
        if c_norm.isdisjoint(a_norm):
            return False

    # Common substitutions — but DO NOT collapse "esc" to "escalator"
    # inside word boundaries, because that makes "customer" vs
    # "customer mgmt esc" look more similar than it should.
    def _canon(s):
        s = s.replace("&", "and").replace("-", " ").replace("_", " ")
        # Only expand "esc" when it appears as a standalone word.
        s = re.sub(r"\besc\b", "escalator", s)
        # Normalize common spelling variants
        s = re.sub(r"\bpreventative\b", "preventive", s)
        s = re.sub(r"\bprogramme\b", "program", s)
        # Replace parens/slash/percent with spaces rather than dropping their
        # contents — "Size (MWDC)" should still match "Size MWDC".
        s = re.sub(r"[()/%]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s
    c = _canon(canonical)
    a = _canon(actual)
    if c == a:
        return True
    # Conservative substring match: both must be ≥8 chars AND the shorter
    # side must be ≥65% of the longer, to avoid over-generous matches like
    # "Annual Premium" ⇆ "Decom Bond Annual Premium".
    if len(c) >= 8 and len(a) >= 8:
        shorter, longer = (c, a) if len(c) <= len(a) else (a, c)
        if shorter in longer and len(shorter) >= 0.65 * len(longer):
            return True
    return False


# Canonical row → list of alternative labels accepted as matches.
# Keeps matching tolerant to template drift without relaxing _labels_match.
ROW_LABEL_ALIASES: dict[int, list[str]] = {
    2: ["Project #", "Project Number", "Project No", "#", "Project ID", "Proj #", "Proj No", "Project Code", "Proj", "No.", "No"],
    4: ["Project Name", "Projects", "Project"],
    7: ["Toggle (On/Off)", "Project Toggle (on/off)", "Project Toggle", "Toggle"],
    10: ["Developer", "Developer Name"],
    11: ["Size MWDC", "Size (MWDC)", "DC Size", "MWdc"],
    12: ["Size MWAC", "Size (MWAC)", "AC Size", "MWac"],
    # Returns block — use template's actual labels first, fall back to legacy
    31: ["Live Appraisal IRR", "Appraisal IRR", "Live Appraisal Model IRR"],
    33: ["FMV Calculated", "FMV Calculated ($/W)", "FMV $/W"],
    36: ["Target IRR"],
    37: ["Live Levered Pre-Tax IRR", "Levered Pre-Tax IRR", "Live Levered PT IRR", "Levered PT IRR"],
    38: ["NPP ($/W) - SOLVE", "NPP ($/W)", "NPP $/W", "NPP per W"],
    39: ["NPP ($)", "NPP $"],
    118: ["PV EPC Cost", "EPC Cost", "PV EPC"],
    119: ["PV LNTP Cost", "LNTP Cost", "PV LNTP"],
    122: ["IX Cost", "Interconnection Cost"],
    123: ["Closing and Legal", "Closing & Legal", "Closing and Legal Cost"],
    216: ["Upfront Incentive", "Upfront Incentives"],
    217: ["Upfront Incentive Lag", "Upfront Lag"],
    157: ["Energy Rate (at COD)", "Energy Rate", "PPA Rate", "Rate at COD"],
    158: ["Energy Rate Escalator", "Rate Escalator", "Escalator"],
    161: ["Rate Discount", "Discount Rate", "GH Discount", "Guidehouse Discount"],
    162: ["Rate UCB Fee", "UCB Fee", "Utility Credit Buyer Fee"],
    225: ["PV O&M Preventative", "O&M Preventative", "PV O&M Preventive", "O&M Preventive"],
    226: ["PV O&M Corrective", "O&M Corrective"],
    227: ["PV O&M Esc", "PV O&M Escalator", "O&M Escalator"],
    230: ["AM Fee", "Asset Management Fee", "Asset Mgmt Fee"],
    231: ["AM Esc", "AM Escalator", "Asset Management Escalator", "Asset Mgmt Escalator"],
    240: ["Customer Mgmt Cost", "Customer Management Cost", "Cust Mgmt", "Customer Mgmt"],
    241: ["Customer Mgmt Esc", "Customer Management Escalator", "Cust Mgmt Escalator"],
    286: ["Decom Annual Premium", "Decom Bond Annual Premium"],
    296: ["P&C Insurance", "P&C Insurance Annual Premium", "Insurance"],
    297: ["P&C Insurance Esc", "P&C Insurance Escalator", "Insurance Escalator"],
    291: ["Custom PropTax Toggle", "Custom Property Tax Toggle", "Custom Prop Tax"],
    292: ["Property Taxes Yr 1", "Property Tax Year 1", "Property Taxes"],
    293: ["PropTax Escalator", "Property Tax Escalator", "Prop Tax Esc"],
    302: ["Internal AM Costs", "Internal Asset Management"],
    587: ["COD Quarter"],
    681: ["Active FMV", "Active Fair Market Value", "FMV", "Active Market Value"],
    591: ["Tax Treatment", "Tax Equity Treatment"],
    596: ["TE Structure", "Tax Equity Structure"],
    597: ["ITC Rate", "ITC %", "Investment Tax Credit Rate", "Investment Tax Credit %", "Tax Credit Rate"],
    602: ["Eligible Costs %", "ITC Eligible Costs", "Eligible Cost %", "ITC Eligible Cost %", "Eligible Costs", "Eligible Cost"],
}


def _detect_label_column(ws, max_row=1000):
    """Detect which column (B-E) contains the input row labels."""
    known = {_normalize_label(v) for v in INPUT_ROW_LABELS.values() if v}
    best_col, best_count = 2, 0
    for col in range(2, 6):
        count = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=col).value
            if val is not None and _normalize_label(val) in known:
                count += 1
        if count > best_count:
            best_col, best_count = col, count
    return best_col


# ---------------------------------------------------------------------------
# Wrapped-EPC component detection
# ---------------------------------------------------------------------------
# The bible's $1.65/W EPC benchmark is the WRAPPED EPC = EPC + LNTP +
# Safe Harbor module amount + EPC Contingency. Models vary in how they place
# these rows, so we scan the label column for any row whose label matches
# one of the patterns below and sum the per-W values for each project.
WRAPPED_EPC_LABEL_PATTERNS = [
    # (regex, component name) — order is informational only
    (r"^pv\s*epc\s*cost$",                           "EPC"),
    (r"^pv\s*lntp(\s*cost)?$",                       "LNTP"),
    (r"safe\s*harbor.*module|module.*safe\s*harbor", "Safe Harbor Modules"),
    (r"epc\s*contingency|contingency.*epc",          "EPC Contingency"),
]


def _scan_wrapped_epc_rows(ws, label_col, max_row=1000):
    """Return list of (row, component_name) for wrapped-EPC build rows."""
    found = []
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=label_col).value
        if val is None:
            continue
        norm = _normalize_label(val)
        for pat, comp_name in WRAPPED_EPC_LABEL_PATTERNS:
            if re.search(pat, norm):
                found.append((r, comp_name))
                break
    return found


# ---------------------------------------------------------------------------
# Rate-component label scanner — Guidehouse discount + ABP REC detection
# ---------------------------------------------------------------------------
# The bible expects a Guidehouse-derived discount applied to community-solar
# rate components. The actual discount % lives at offset+7 of each rate
# component start row. Separately, if any rate component is named "ABP REC"
# (or similar) AND that component's equity toggle is on, the project should
# be priced using ABP market assumptions for its (state, utility) regardless
# of how the program field is labeled.
GUIDEHOUSE_NAME_PATTERNS = [
    r"guidehouse",
    r"\bgh\b",                  # "GH discount" abbreviation
]
ABP_REC_NAME_PATTERNS = [
    r"\babp\s*rec\b",
    r"adjustable\s*block",
    r"\babp\b.*\brec\b",
]


def _matches_any(name, patterns):
    if not name:
        return False
    n = _normalize_label(name)
    return any(re.search(p, n) for p in patterns)


def _scan_rate_components(ws, col):
    """Inspect each rate component's name + toggles for this project column.

    Returns:
        {
          "guidehouse": [{idx, name, discount, equity_on, debt_on, appraisal_on}, ...],
          "abp_rec":    [{idx, name, energy_rate, discount, equity_on, ...}, ...],
          "abp_rec_live": bool,    # any ABP REC component with equity_on truthy
        }
    """
    out = {"guidehouse": [], "abp_rec": [], "abp_rec_live": False}
    for i, start in enumerate(RATE_COMP_STARTS, 1):
        name = ws.cell(row=start + 1, column=col).value
        if not name:
            continue
        eq_on  = safe_float(ws.cell(row=EQUITY_RATE_TOGGLE_START + (i - 1), column=col).value)
        dt_on  = safe_float(ws.cell(row=DEBT_RATE_TOGGLE_START   + (i - 1), column=col).value)
        ap_on  = safe_float(ws.cell(row=APPRAISAL_RATE_TOGGLE_START + (i - 1), column=col).value)
        rec = {
            "idx": i, "name": str(name).strip(),
            "energy_rate": safe_float(ws.cell(row=start + 3, column=col).value),
            "discount":    safe_float(ws.cell(row=start + 7, column=col).value),
            "equity_on": bool(eq_on), "debt_on": bool(dt_on), "appraisal_on": bool(ap_on),
        }
        if _matches_any(name, GUIDEHOUSE_NAME_PATTERNS):
            out["guidehouse"].append(rec)
        if _matches_any(name, ABP_REC_NAME_PATTERNS):
            out["abp_rec"].append(rec)
            if rec["equity_on"]:
                out["abp_rec_live"] = True
    return out


#: Critical canonical rows — we warn loudly when these are unresolved, since
#: downstream consumers show stale / garbage data without them.
_CRITICAL_CANONICAL_ROWS = (4, 7, 10, 11, 18, 22, 31, 33, 37, 38, 39, 118, 681)


def _build_row_mapping(ws, label_col, max_row=1000):
    """Build canonical_row -> actual_row mapping by scanning labels in the model.

    Returns a dict where keys are canonical row numbers (from INPUT_ROW_LABELS)
    and values are the actual row numbers found in this specific model, or
    `None` when no confident match was found.

    Critically, unmatched rows map to None (NOT to the canonical row number)
    so downstream code reads None instead of whatever cell happens to sit at
    the canonical position (which is usually garbage when labels have drifted).
    """
    # Read all labels from the label column. Track ALL occurrences so we can
    # detect duplicated labels (common source of subtle mapping bugs).
    actual_by_norm_first: dict[str, int] = {}
    actual_by_norm_all: dict[str, list[int]] = {}
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=label_col).value
        if val is not None:
            norm = _normalize_label(val)
            if norm:
                actual_by_norm_all.setdefault(norm, []).append(r)
                if norm not in actual_by_norm_first:
                    actual_by_norm_first[norm] = r  # first occurrence wins

    for norm, rows in actual_by_norm_all.items():
        if len(rows) > 1:
            logger.warning(
                "Duplicate label %r appears at rows %s — first wins (row %d). "
                "If the later rows matter, add more-specific aliases.",
                norm, rows, rows[0],
            )

    all_canonical = dict(INPUT_ROW_LABELS)
    all_canonical.update(OUTPUT_ROWS)

    mapping: dict[int, int | None] = {}
    unresolved_critical: list[int] = []
    for canonical_row, label in all_canonical.items():
        candidates = [label] + ROW_LABEL_ALIASES.get(canonical_row, [])
        resolved = None
        # Exact-normalized pass first (strongest signal).
        for cand in candidates:
            norm = _normalize_label(cand)
            if norm in actual_by_norm_first:
                resolved = actual_by_norm_first[norm]
                break
        # Fuzzy pass, using the tightened _labels_match.
        if resolved is None:
            for cand in candidates:
                norm = _normalize_label(cand)
                for actual_norm, actual_row in actual_by_norm_first.items():
                    if _labels_match(norm, actual_norm):
                        resolved = actual_row
                        break
                if resolved is not None:
                    break
        mapping[canonical_row] = resolved  # may be None
        if resolved is None and canonical_row in _CRITICAL_CANONICAL_ROWS:
            unresolved_critical.append(canonical_row)

    # Collision detection: if two canonical rows resolved to the same actual
    # row, the data for one of them is silently wrong. Keep the canonical row
    # whose label had an exact (not fuzzy) match; set the other(s) to None.
    actual_to_canonicals: dict[int, list[int]] = {}
    for canon_r, actual_r in mapping.items():
        if actual_r is not None:
            actual_to_canonicals.setdefault(actual_r, []).append(canon_r)
    for actual_r, canon_list in actual_to_canonicals.items():
        if len(canon_list) <= 1:
            continue
        # Prefer the canonical whose primary label is an exact normalized match.
        actual_norm = actual_by_norm_first.get(
            next((n for n, r in actual_by_norm_first.items() if r == actual_r), ""),
            "",
        )
        exact_winner = None
        for cr in canon_list:
            primary = _normalize_label(all_canonical.get(cr, ""))
            if primary and primary == actual_norm:
                exact_winner = cr
                break
        if exact_winner is None:
            exact_winner = canon_list[0]  # fallback: first canonical wins
        losers = [cr for cr in canon_list if cr != exact_winner]
        for cr in losers:
            mapping[cr] = None
        logger.warning(
            "Collision: actual row %d claimed by canonical rows %s — "
            "keeping %d (exact match), setting %s to None.",
            actual_r, canon_list, exact_winner, losers,
        )

    if unresolved_critical:
        logger.warning(
            "Critical row(s) unresolved: %s — downstream UI may show blanks "
            "or default to a fallback. Consider adding labels in column %s.",
            unresolved_critical, chr(64 + label_col) if label_col else "?",
        )
    return mapping


# ---------------------------------------------------------------------------
# Model loading
# ---------------------------------------------------------------------------


def load_pricing_model(file):
    """Load a pricing model workbook and extract all project data.

    Uses read_only=False for fast random cell access (read_only streams rows
    sequentially which is catastrophically slow for sparse random reads:
    456s with read_only vs ~3s without on an 8.5MB xlsm).

    Row positions are detected dynamically per model by scanning column B
    labels, so models with slightly different row layouts are handled correctly.
    """
    wb = openpyxl.load_workbook(file, data_only=True, read_only=False)
    # Locate the Project Inputs sheet tolerantly — names drift across templates
    # ("Project Inputs", "Project Input", "Project_Inputs", "ProjectInputs"…).
    ws = None
    for sn in wb.sheetnames:
        norm = re.sub(r"[^a-z]", "", sn.lower())
        if norm == "projectinputs":
            ws = wb[sn]
            break
    if ws is None:
        # Second pass: any sheet whose name starts with "project input"
        for sn in wb.sheetnames:
            if sn.strip().lower().startswith("project input"):
                ws = wb[sn]
                break
    if ws is None:
        logger.error(
            "Project Inputs sheet not found. Sheets present: %s",
            wb.sheetnames,
        )
        raise KeyError(
            "Could not find a 'Project Inputs' sheet in this workbook. "
            f"Sheets present: {', '.join(wb.sheetnames)}"
        )
    logger.info("Loaded workbook; Project Inputs sheet = %r", ws.title)

    # Detect label column and build per-model row mapping
    label_col = _detect_label_column(ws)
    row_map = _build_row_mapping(ws, label_col)

    all_needed_canonical = set(INPUT_ROW_LABELS.keys()) | set(OUTPUT_ROWS.keys())
    rate_rows = set()
    for start in RATE_COMP_STARTS:
        for offset in range(1, 9):
            rate_rows.add(start + offset)
    for i in range(6):
        rate_rows.add(EQUITY_RATE_TOGGLE_START + i)
        rate_rows.add(DEBT_RATE_TOGGLE_START + i)
        rate_rows.add(APPRAISAL_RATE_TOGGLE_START + i)
    rate_rows.add(400)   # debt match toggle
    rate_rows.add(512)   # appraisal match toggle

    # Resolve actual row numbers for key fixed-position rows
    # (name+toggle rarely drift, but fall back to canonical so we can still
    # enumerate project columns even on templates with unusual label layouts.)
    name_row = row_map.get(4) or 4
    toggle_row = row_map.get(7) or 7

    # One-shot scan for wrapped-EPC component rows (label-based, model-agnostic)
    wrapped_epc_rows = _scan_wrapped_epc_rows(ws, label_col)

    projects = {}
    for col in range(6, 88):
        name_cell = ws.cell(row=name_row, column=col).value
        if name_cell is None or not str(name_cell).strip():
            continue
        toggle_cell = ws.cell(row=toggle_row, column=col).value
        # openpyxl may return 1 as int or 1.0 as float depending on cell format.
        # Accept numeric truthiness first, then fall back to a text check.
        if toggle_cell is None:
            is_on = False
        else:
            num = safe_float(toggle_cell)
            if num is not None:
                is_on = num != 0
            else:
                is_on = str(toggle_cell).strip().lower() in ("1", "on", "true", "yes", "y")

        data = {}
        for canonical_r in all_needed_canonical:
            actual_r = row_map.get(canonical_r)
            # Unresolved labels → None (don't read whatever lives at the
            # canonical row, which would yield misleading values).
            data[canonical_r] = ws.cell(row=actual_r, column=col).value if actual_r else None

        # Read ALL non-empty cells for this project column, keyed by the
        # column B label. This enables the walk builder to compare ANY
        # hardcoded input between two models by matching on label text,
        # not just the rows we have canonical definitions for.
        all_inputs: dict[str, Any] = {}
        for r in range(1, max_row + 1):
            label_val = ws.cell(row=r, column=label_col).value
            cell_val = ws.cell(row=r, column=col).value
            if label_val is not None and cell_val is not None:
                label_str = str(label_val).strip()
                if label_str:
                    all_inputs[label_str] = cell_val
        data["_all_inputs"] = all_inputs

        rate_comps = {}
        for i, start in enumerate(RATE_COMP_STARTS, 1):
            rate_comps[i] = {
                "name": ws.cell(row=start+1, column=col).value,
                "custom_generic": ws.cell(row=start+2, column=col).value,
                "energy_rate": ws.cell(row=start+3, column=col).value,
                "escalator": ws.cell(row=start+4, column=col).value,
                "start_date": ws.cell(row=start+5, column=col).value,
                "term": ws.cell(row=start+6, column=col).value,
                "discount": ws.cell(row=start+7, column=col).value,
                "ucb_fee": ws.cell(row=start+8, column=col).value,
                "equity_on": safe_float(ws.cell(row=EQUITY_RATE_TOGGLE_START+(i-1), column=col).value),
                "debt_on": safe_float(ws.cell(row=DEBT_RATE_TOGGLE_START+(i-1), column=col).value),
                "appraisal_on": safe_float(ws.cell(row=APPRAISAL_RATE_TOGGLE_START+(i-1), column=col).value),
            }

        # For Custom rate components, the escalator input on Project Inputs
        # is not used — the rate curve defines year-by-year rates directly.
        # Suppress the escalator to avoid showing misleading values.
        for i, start in enumerate(RATE_COMP_STARTS, 1):
            cg = ws.cell(row=start+2, column=col).value  # Custom/Generic
            if cg and str(cg).strip().lower() == "custom":
                esc_canonical = start + 4  # escalator row for this rate comp
                if esc_canonical in data:
                    data[esc_canonical] = "N/A (Custom)"

        data["_debt_match_equity"] = ws.cell(row=400, column=col).value
        data["_appraisal_match_equity"] = ws.cell(row=512, column=col).value

        dscr_label = ws.cell(row=341, column=col).value
        dscr_schedule = {}
        for yr in range(1, 31):
            v = safe_float(ws.cell(row=341 + yr, column=col).value)
            if v is not None:
                dscr_schedule[yr] = v

        data["_front_back_toggle"] = ws.cell(row=320, column=col).value
        data["_debt_sizing_method"] = ws.cell(row=321, column=col).value

        # ---- Wrapped EPC build: sum EPC + LNTP + Safe Harbor + Contingency ----
        # The bible's $1.65/W EPC benchmark is the WRAPPED total. We sum every
        # component row found by the label scan; missing components contribute 0.
        epc_components = []
        wrapped_total = 0.0
        any_value = False
        for r, comp_name in wrapped_epc_rows:
            v = safe_float(ws.cell(row=r, column=col).value)
            epc_components.append({"row": r, "component": comp_name, "value": v})
            if v is not None:
                wrapped_total += v
                any_value = True
        data["_wrapped_epc_components"] = epc_components
        # Plausibility check: each wrapped EPC component should be in $/W range (0-10).
        # Values outside this range suggest a unit mismatch (e.g. total $ not per-watt).
        for comp in epc_components:
            v = comp.get("value")
            if v is not None and (v < 0 or v > 10):
                logger.warning(
                    "Wrapped EPC component %r = %.2f at row %d for col %d — "
                    "outside $/W plausibility range (0-10). Possible unit mismatch.",
                    comp.get("component"), v, comp.get("row"), col,
                )
        data["_wrapped_epc_total"] = wrapped_total if any_value else None
        # Override row 118 used by the bible audit so the comparison runs
        # against the WRAPPED build, not raw PV EPC alone. Original raw EPC
        # is preserved at _raw_epc_118 for diagnostics.
        if any_value:
            data["_raw_epc_118"] = data.get(118)
            data[118] = wrapped_total

        # ---- Rate-component scan: Guidehouse discount + ABP REC live state ----
        rate_scan = _scan_rate_components(ws, col)
        data["_guidehouse_components"] = rate_scan["guidehouse"]
        data["_abp_rec_components"]    = rate_scan["abp_rec"]
        data["_abp_rec_live"]          = rate_scan["abp_rec_live"]

        clean_name = " | ".join(line.strip() for line in str(name_cell).strip().splitlines() if line.strip())
        projects[col] = {
            "name": clean_name, "toggle": is_on,
            "col_letter": openpyxl.utils.get_column_letter(col),
            "data": data, "rate_comps": rate_comps,
            "dscr_label": dscr_label, "dscr_schedule": dscr_schedule,
        }

    # Ops Sandbox
    ops_sandbox = {"revenue_adders": [], "opex_overrides": []}
    if "Ops Sandbox" in wb.sheetnames:
        ws_ops = wb["Ops Sandbox"]
        ops_sandbox["live_project"] = ws_ops.cell(row=15, column=4).value
        for r in range(19, 29):
            label = ws_ops.cell(row=r, column=2).value
            if label and "Placeholder" not in str(label):
                ops_sandbox["revenue_adders"].append({
                    "label": str(label).strip(),
                    "annual": safe_float(ws_ops.cell(row=r, column=3).value),
                    "equity": safe_float(ws_ops.cell(row=r, column=4).value) or 0,
                    "debt": safe_float(ws_ops.cell(row=r, column=5).value) or 0,
                    "appraisal": safe_float(ws_ops.cell(row=r, column=6).value) or 0,
                    "npv_total": safe_float(ws_ops.cell(row=r, column=8).value),
                })
        for r in range(36, 46):
            label = ws_ops.cell(row=r, column=2).value
            if label and "Placeholder" not in str(label):
                ops_sandbox["opex_overrides"].append({
                    "label": str(label).strip(),
                    "annual": safe_float(ws_ops.cell(row=r, column=3).value),
                    "equity": safe_float(ws_ops.cell(row=r, column=4).value) or 0,
                    "debt": safe_float(ws_ops.cell(row=r, column=5).value) or 0,
                    "appraisal": safe_float(ws_ops.cell(row=r, column=6).value) or 0,
                    "npv_total": safe_float(ws_ops.cell(row=r, column=8).value),
                })
        for attr, row in [("equity_adder_npv", 31), ("debt_adder_npv", 32), ("appraisal_adder_npv", 33),
                          ("equity_opex_npv", 48), ("debt_opex_npv", 49), ("appraisal_opex_npv", 50)]:
            ops_sandbox[attr] = safe_float(ws_ops.cell(row=row, column=8).value)

    # --- Rate Curves sheet ---
    rate_curves = {}
    if "Rate Curves" in wb.sheetnames:
        ws_rc = wb["Rate Curves"]

        # Read date row (row 5, cols 10+)
        rc_dates = {}
        for c in range(10, min(510, ws_rc.max_column + 1)):
            d = ws_rc.cell(row=5, column=c).value
            if d is not None and hasattr(d, "year"):
                rc_dates[c] = d
            elif d is None and c > 20:
                break

        # Build project name -> RC custom-rate row mapping
        # RC1 projects start at row 35, RC2 at row 117, etc.
        # Each RC block is 82 rows apart: 30, 112, 194, 276, 358, 440
        rc_block_starts = [30, 112, 194, 276, 358, 440]
        proj_row_offset = 5  # projects start 5 rows after block start

        # Map project names to their rate curve row within each block
        rc_proj_rows = {}  # {rc_index: {project_name: row_number}}
        for rc_idx, blk_start in enumerate(rc_block_starts, 1):
            rc_proj_rows[rc_idx] = {}
            rc_label_row = blk_start  # e.g. "Rate Component 1"
            rc_name = ws_rc.cell(row=rc_label_row, column=4).value  # curve name (e.g. "GH25 -22.5")

            # Custom rate row (aggregate) is blk_start + 2
            custom_rate_row = blk_start + 2

            # Per-project custom rates start at blk_start + 5
            for pr in range(blk_start + proj_row_offset, blk_start + 80):
                pname = ws_rc.cell(row=pr, column=2).value
                if pname and "Project " not in str(pname) and "Anchor" not in str(pname):
                    rc_proj_rows[rc_idx][str(pname).strip()] = pr
                elif pname and "Anchor" in str(pname):
                    break

            rate_curves[f"rc{rc_idx}_name"] = str(rc_name or "")
            rate_curves[f"rc{rc_idx}_custom_row"] = custom_rate_row
            rate_curves[f"rc{rc_idx}_proj_rows"] = rc_proj_rows[rc_idx]

        # Extract monthly rate data per project per RC
        rate_curves["dates"] = rc_dates
        rate_curves["projects"] = {}

        for col_idx, proj in projects.items():
            pname = proj["name"]
            # Find matching project row in Rate Curves by name matching
            proj_rc_data = {}
            for rc_idx in range(1, 7):
                proj_rows_map = rc_proj_rows.get(rc_idx, {})
                # Try exact match first, then partial
                matched_row = None
                for rc_pname, rc_row in proj_rows_map.items():
                    if rc_pname in pname or pname in rc_pname:
                        matched_row = rc_row
                        break
                    # Also try first part of multi-line name
                    first_part = pname.split(" | ")[0].strip()
                    if rc_pname == first_part or first_part in rc_pname:
                        matched_row = rc_row
                        break

                if matched_row:
                    monthly = {}
                    for c, dt in rc_dates.items():
                        v = safe_float(ws_rc.cell(row=matched_row, column=c).value)
                        if v is not None and v != 0:
                            monthly[dt] = v
                    proj_rc_data[rc_idx] = monthly

            rate_curves["projects"][pname] = proj_rc_data

    wb.close()
    # Deep-copy before returning so the Streamlit cache stores an immutable
    # snapshot. Downstream code (mockup_view, render_html, filter_projects)
    # sometimes iterates and mutates in-place; without the copy a second
    # render in the same session would see a corrupted dict.
    return copy.deepcopy(
        {"projects": projects, "ops_sandbox": ops_sandbox, "rate_curves": rate_curves}
    )


def get_projects(model_result):
    """Extract projects dict. Cache already returns a deepcopy (line 637),
    so no additional copy needed here."""
    if isinstance(model_result, dict) and "projects" in model_result:
        return model_result["projects"]
    return {}


def get_ops_sandbox(model_result):
    if isinstance(model_result, dict) and "ops_sandbox" in model_result:
        return model_result["ops_sandbox"]
    return {}


def get_rate_curves(model_result):
    if isinstance(model_result, dict) and "rate_curves" in model_result:
        return model_result["rate_curves"]
    return {}





def load_mapper_output(file):
    """Load a portfolio-model-mapper output workbook.

    The mapper produces a 'Model Paste Format' (or 'Full Row Mapping 2') sheet
    with this layout:
        Col A: Model Row#   (matches our canonical INPUT_ROW_LABELS keys)
        Col B: Field        (label, ignored — we trust the row #)
        Col C: Units        (ignored)
        Col D: Source       (CIM / Default / Manual — kept for tooltip)
        Col E..N: one column per project, header is project name in row 1

    Returns a dict shaped exactly like load_pricing_model()'s 'projects' so the
    UI can drop it in alongside Model 1 / Model 2:
        {col_idx: {"name": ..., "toggle": True, "col_letter": ...,
                   "data": {row#: value}, "rate_comps": {}, "_source_map": {row#: src}}}
    """
    wb = openpyxl.load_workbook(file, data_only=True, read_only=False)

    # Prefer the paste-format sheet (canonical row #s); fall back to Full Row Mapping
    target_sheet = None
    for candidate in ("Model Paste Format", "Full Row Mapping 2", "Full Row Mapping"):
        if candidate in wb.sheetnames:
            target_sheet = candidate
            break
    if target_sheet is None:
        wb.close()
        return {}

    ws = wb[target_sheet]

    # Header row 1: cols 5+ contain project names. Skip empty headers.
    project_cols = []
    for c in range(5, ws.max_column + 1):
        name = ws.cell(row=1, column=c).value
        if name and str(name).strip():
            # Cleanup: strip leading state code/site code in parentheses if any
            project_cols.append((c, str(name).strip()))

    if not project_cols:
        wb.close()
        return {}

    projects = {}
    for c, pname in project_cols:
        data = {}
        sources = {}
        for r in range(2, ws.max_row + 1):
            row_num_cell = ws.cell(row=r, column=1).value
            row_num = safe_float(row_num_cell)
            if row_num is None:
                continue
            row_int = int(row_num)
            val = ws.cell(row=r, column=c).value
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            data[row_int] = val
            sources[row_int] = ws.cell(row=r, column=4).value

        if not data:
            continue

        # Clean name (mirror load_pricing_model logic)
        clean_name = " | ".join(line.strip() for line in pname.splitlines() if line.strip())
        projects[c] = {
            "name": clean_name,
            "toggle": True,
            "col_letter": openpyxl.utils.get_column_letter(c),
            "data": data,
            "rate_comps": {},
            "dscr_label": None,
            "dscr_schedule": {},
            "_source_map": sources,
            "_origin": "mapper",
        }

    wb.close()
    return projects



def load_data_room(files):
    data_room = {}
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
        fname = Path(f.name).stem if hasattr(f, 'name') else str(f)
        sheets_info = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=False):
                rows.append([c.value for c in row])
            sheets_info[sn] = rows
        data_room[fname] = {"sheets": wb.sheetnames, "data": sheets_info}
        wb.close()
    return data_room
