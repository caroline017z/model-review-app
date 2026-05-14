"""
38DN Build Walk — Excel Walk Summary Generator

Compares two pricing models side-by-side, generating a formatted .xlsx file
that shows per-project NPP/IRR deltas and auto-detected variance drivers
grouped by category.

Output format matches existing 38DN Walk Summary files exactly:
  - Navy project name fills, white bold font
  - NPP in 0.000 accounting format with red negatives
  - IRR in 0.00% format
  - ∆ Base with Excel formulas (zero shows as dash)
  - Yellow highlights on changed values
  - MW-weighted SUMPRODUCT averages
  - Category-grouped variance drivers below
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from lib.bible_audit import audit_project, verdict_from_summary
from lib.config import (
    BIBLE_BENCHMARKS,
    DATE_ROWS,
    DPW_ROWS,
    INPUT_ROW_LABELS,
    INPUT_ROW_UNITS,
    INT_ROWS,
    PCT_ROWS,
    TEXT_ROWS,
)
from lib.data_loader import get_projects
from lib.impact import portfolio_impact
from lib.rows import ROW_DC_MW, ROW_LEVERED_PT_IRR, ROW_NPP, ROW_PROJECT_NUMBER
from lib.utils import canonicalize_name, canonicalize_pnum, safe_float

logger = logging.getLogger(__name__)


def _equality_numeric(v: Any) -> float | None:
    """Numeric form for equality comparison.

    Differs from safe_float in that bools coerce to 0.0/1.0. Without this, a
    toggle row that stores True in one model and 1 in the other would route
    through the text branch and falsely flag as differing.
    """
    if isinstance(v, bool):
        return float(v)
    return safe_float(v)


def _values_differ(v1: Any, v2: Any, *, is_text: bool = False, tol: float = 1e-6) -> bool:
    """Equality check used across diff_inputs. Returns True if v1 != v2.

    Handles None (matches None), bool/int parity (True == 1), text tolerance
    (case/whitespace), and date equality.
    """
    if v1 is None and v2 is None:
        return False
    if is_text:
        return str(v1 or "").strip() != str(v2 or "").strip()
    f1 = _equality_numeric(v1)
    f2 = _equality_numeric(v2)
    if f1 is not None and f2 is not None:
        return abs(f1 - f2) > tol
    if f1 is None and f2 is None:
        return str(v1 or "").strip() != str(v2 or "").strip()
    return True


# Rows where models sometimes disagree on unit storage. Conversion factor
# from "kWh basis" to the "displayed" basis uses project yield (row 14,
# kWh/Wdc). Magnitude heuristic: if one side is < ~1 and other side is
# >> 100, it's almost certainly a unit mismatch — canonicalize both to the
# displayed basis before diff-comparing to avoid false positives.
_UNIT_MISMATCH_ROWS: dict[int, dict] = {
    121: {
        "label": "Cust Acquisition",
        "to_displayed": lambda v, y: v * y,  # $/kWh → $/W
        "displayed_unit": "$/W",
    },
    240: {
        "label": "Cust Mgmt",
        "to_displayed": lambda v, y: v * y * 1_000_000,  # $/kWh → $/MW/yr
        "displayed_unit": "$/MW/yr",
    },
}


def _reconcile_unit_sensitive(
    row: int,
    v1: Any,
    v2: Any,
    m1_yield: float | None,
    m2_yield: float | None,
) -> tuple[Any, Any, str | None]:
    """For rows where $/kWh vs $/W|$/MW/yr storage drift is known to happen,
    detect magnitude-based mismatch and canonicalize both sides to the
    displayed unit. Returns (v1_new, v2_new, note). Note is populated when
    a conversion was applied — walk renders it as a footnote on the row.
    """
    spec = _UNIT_MISMATCH_ROWS.get(row)
    if spec is None:
        return v1, v2, None
    f1 = safe_float(v1)
    f2 = safe_float(v2)
    if f1 is None or f2 is None:
        return v1, v2, None
    # Mismatch signature: one side in kWh basis (< ~1), the other in
    # displayed basis (> ~100). Symmetric check — convert the small side up.
    small, large = (f1, f2) if f1 < f2 else (f2, f1)
    if not (small < 1.0 and large > 100.0):
        return v1, v2, None
    if f1 == small:
        y = m1_yield
        if y is None or y <= 0:
            return v1, v2, None
        return (
            spec["to_displayed"](f1, y),
            f2,
            (f"M1 converted from $/kWh to {spec['displayed_unit']} (yield={y:.3f})"),
        )
    else:
        y = m2_yield
        if y is None or y <= 0:
            return v1, v2, None
        return (
            f1,
            spec["to_displayed"](f2, y),
            (f"M2 converted from $/kWh to {spec['displayed_unit']} (yield={y:.3f})"),
        )


# ---------------------------------------------------------------------------
# Placeholder detection (template projects we don't treat as real portfolio)
# ---------------------------------------------------------------------------

_PLACEHOLDER_RE = re.compile(r"^\s*project\s+\d+\s*$", re.IGNORECASE)


def _is_placeholder(name: str) -> bool:
    return bool(_PLACEHOLDER_RE.match(name)) or name.strip().lower() in ("anchor", "sample", "")


# ---------------------------------------------------------------------------
# Project pairing & orphan helpers (used by build_walk_xlsx to populate the
# Unmatched sheet with reason codes)
# ---------------------------------------------------------------------------


def _build_pnum_set(projects: dict) -> set[int]:
    """Integer Project # index for fast 'does this # exist in that model' checks."""
    out: set[int] = set()
    for _c, p in projects.items():
        if not isinstance(p, dict):
            continue
        pn = safe_float((p.get("data") or {}).get(ROW_PROJECT_NUMBER))
        if pn is not None:
            out.add(int(pn))
    return out


def _build_name_set(projects: dict) -> set[str]:
    """Canonicalized project-name index for orphan reason resolution."""
    out: set[str] = set()
    for _c, p in projects.items():
        if not isinstance(p, dict):
            continue
        n = canonicalize_name(p.get("name"))
        if n:
            out.add(n)
    return out


def _orphan_reason(proj: dict, other_pnums: set[int], other_names: set[str]) -> str:
    """Classify why a project failed to pair. Reason codes:
    missing_proj_num       — own proj_number is None/blank
    proj_num_not_in_other  — has proj#, but no project in the other model has it
    name_not_in_other      — proj# matches but name doesn't (unusual)
    unknown                — pairing logic missed it for an unclassified reason
    """
    data = proj.get("data") or {}
    pn = safe_float(data.get(ROW_PROJECT_NUMBER))
    name_canon = canonicalize_name(proj.get("name"))
    if pn is None:
        return "missing_proj_num"
    if int(pn) not in other_pnums:
        return "proj_num_not_in_other"
    if name_canon and name_canon not in other_names:
        return "name_not_in_other"
    return "unknown"


# ---------------------------------------------------------------------------
# Rate Curves COD lookup (used by Pass 1c of diff_inputs)
# ---------------------------------------------------------------------------

from lib.rate_curve import (
    rate_at_cod as _rate_at_cod,  # noqa: E402 — re-export for inline call sites in this module
)

# ---------------------------------------------------------------------------
# Formatting constants (match reference Walk Summary files exactly)
# ---------------------------------------------------------------------------

NAVY_FILL = PatternFill("solid", fgColor="002060")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC")

# Verdict fills — match Excel's conditional-formatting palette so they read
# the same way reviewers expect (green=good, yellow=caution, red=bad).
_VERDICT_FILLS = {
    "CLEAN": PatternFill("solid", fgColor="C6EFCE"),  # light green
    "REVIEW": PatternFill("solid", fgColor="FFEB9C"),  # light yellow
    "REWORK REQUIRED": PatternFill("solid", fgColor="FFC7CE"),  # light red
}

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
BLUE_FONT = Font(color="0000FF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)

_THIN = Side(style="thin")
_DOUBLE = Side(style="double")
THIN_BOTTOM = Border(bottom=_THIN)
DOUBLE_BOTTOM = Border(bottom=_DOUBLE)
# Column separators for the NPP/IRR/delta grid
THIN_LEFT = Border(left=_THIN)
THIN_LEFT_RIGHT = Border(left=_THIN, right=_THIN)
THIN_BOTTOM_LEFT = Border(bottom=_THIN, left=_THIN)
THIN_BOTTOM_LEFT_RIGHT = Border(bottom=_THIN, left=_THIN, right=_THIN)
# Boxed cell for variance section values
THIN_BOX = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
# Header row: bottom + left separator
HDR_NPP = Border(bottom=_THIN, left=_THIN, top=_THIN)
HDR_DELTA = Border(bottom=_THIN, left=_THIN, right=_THIN, top=_THIN)

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER_CONT = Alignment(horizontal="centerContinuous", vertical="center")

FMT_MW = "0.00"
FMT_NPP = "0.000_);[Red]\\(0.000\\)"
FMT_IRR = "0.00%"
FMT_DELTA = '0.000_);[Red]\\(0.000\\);"-"'
FMT_DPW = "#,##0.000"
FMT_INT = "#,##0"
FMT_PCT = "0.00%"

# Rows to skip in variance detection (identity / metadata rows)
_SKIP_ROWS = {2, 4, 7, 8, 10, 18, 19}

# Labels to EXCLUDE from the walk variance section (covered elsewhere or outputs)
_WALK_EXCLUDE_LABELS = {
    "live appraisal model irr",
    "live appraisal irr",
    "max fmv w/ constraint",
    "max fmv with constraint",
    "minimum equity dscr multiple",
    "npp ($)",
    "npp ($/w) - solve",
    "npp ($/w)",
    "other upfront costs",
    "project toggle (on/off)",
    "toggle (on/off)",
    "step up dev fee - solve",
    "step up dev fee",
    "te pre-commitment amount for cl sizing",
    "tax equity insurance costs",
    "total capex excl. financing costs",
    "total capex excl. financing",
    "total capex incl. financing costs",
    "total capex incl. financing",
    "total itc",
    "unconstrained (calculated) fmv",
    "unconstrained fmv",
    "live levered pre-tax irr",
    "levered pre-tax irr",
    "active fmv",
    "active fair market value",
    "fmv step up",
    "fmv step-up",
    "step up cap",
    "fmv wacc",
    "fmv wacc (target)",
}

# Category display order
_CATEGORY_ORDER = [
    "CapEx",
    "System Sizing",
    "Revenue",
    "Incentives & Tax",
    "System Details",
    "OpEx",
    "Other",
]


# ---------------------------------------------------------------------------
# Row categorization
# ---------------------------------------------------------------------------


def _build_category_map() -> dict[int, str]:
    """Invert BIBLE_BENCHMARKS to get {row_number: category_name}."""
    cat_map: dict[int, str] = {}
    for category, checks in BIBLE_BENCHMARKS.items():
        for _label, spec in checks.items():
            if "row" in spec:
                cat_map[int(spec["row"])] = category
    return cat_map


_CATEGORY_MAP = _build_category_map()

# Fallback ranges for rows not in BIBLE_BENCHMARKS
_FALLBACK_RANGES = [
    (range(117, 130), "CapEx"),
    (range(143, 215), "Revenue"),
    (range(215, 222), "Incentives & Tax"),
    (range(225, 311), "OpEx"),
    (range(587, 611), "Incentives & Tax"),
]


def _categorize_row(row: int) -> str:
    if row in _CATEGORY_MAP:
        return _CATEGORY_MAP[row]
    for rng, cat in _FALLBACK_RANGES:
        if row in rng:
            return cat
    return "Other"


# ---------------------------------------------------------------------------
# Project matching
# ---------------------------------------------------------------------------


def _build_pnum_index(projects: dict) -> dict[str, tuple[int, Any]]:
    """Project # → (col, raw_pnum) index keyed on canonicalize_pnum output.

    Handles int/float/string variance ("1" vs 1 vs 1.0 all key the same)
    and string project numbers like "P-001". The raw pnum is preserved so
    the returned match dict can report the original display value.
    """
    idx: dict[str, tuple[int, Any]] = {}
    for col, proj in projects.items():
        if not isinstance(proj, dict) or "data" not in proj:
            continue
        raw = proj["data"].get(ROW_PROJECT_NUMBER)
        canon = canonicalize_pnum(raw)
        if canon is not None:
            idx[canon] = (col, raw)
    return idx


def _build_positional_pnum_index(projects: dict) -> dict[str, tuple[int, Any]]:
    """Same as _build_pnum_index but falls back to column-position-inferred
    project numbers when the row-2 value is blank. Typical models start
    projects at column F (6), so col - 5 = inferred project #.
    """
    idx: dict[str, tuple[int, Any]] = {}
    for col, proj in projects.items():
        if not isinstance(proj, dict) or "data" not in proj:
            continue
        raw = proj["data"].get(ROW_PROJECT_NUMBER)
        canon = canonicalize_pnum(raw)
        if canon is not None:
            idx[canon] = (col, raw)
        elif proj.get("name"):
            inferred = col - 5
            if inferred >= 1:
                idx[str(inferred)] = (col, inferred)
    return idx


def _build_name_match_index(projects: dict) -> dict[str, int]:
    """Canonical-name → col index for name-based fallback matching."""
    idx: dict[str, int] = {}
    for col, proj in projects.items():
        if not isinstance(proj, dict):
            continue
        canon = canonicalize_name(proj.get("name"))
        if canon:
            idx[canon] = col
    return idx


def _pnum_display(raw: Any) -> Any:
    """Convert a raw Project # cell value into the form downstream uses:
    int when the value is integer-valued numeric (the common case); the
    original string otherwise (e.g. "P-001"). Preserves int-filter
    compatibility with the frontend's include_proj_numbers request field.
    """
    fv = safe_float(raw)
    if fv is not None and fv == int(fv):
        return int(fv)
    return raw


def _make_matched(
    proj_number: Any,
    name: str,
    m1_col: int,
    m2_col: int,
    match_source: str,
) -> dict:
    return {
        "proj_number": _pnum_display(proj_number),
        "name": name,
        "m1_col": m1_col,
        "m2_col": m2_col,
        "match_source": match_source,
    }


def match_projects(
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Match projects between two models.

    Strategy (first match wins):
      1. Canonical Project # (row 2) — handles int/float/string variance.
      2. Positional fallback — infer proj# from column position (col - 5).
         Used only when primary yields zero matches.
      3. Name fallback — canonical-name equality. Used only when both
         primary and positional yield zero matches.

    Returns list of matched dicts with keys: proj_number, name, m1_col,
    m2_col, match_source (one of "proj_num" | "positional" | "name").
    """
    m1_idx = _build_pnum_index(m1_projects)
    m2_idx = _build_pnum_index(m2_projects)
    common = sorted(set(m1_idx.keys()) & set(m2_idx.keys()))
    matched = [
        _make_matched(
            proj_number=m1_idx[k][1],
            name=str(m1_projects[m1_idx[k][0]].get("name") or "Unnamed").strip(),
            m1_col=m1_idx[k][0],
            m2_col=m2_idx[k][0],
            match_source="proj_num",
        )
        for k in common
    ]
    if matched:
        return matched

    # Fallback 1: positional project # (col - 5)
    m1_pos = _build_positional_pnum_index(m1_projects)
    m2_pos = _build_positional_pnum_index(m2_projects)
    common_pos = sorted(set(m1_pos.keys()) & set(m2_pos.keys()))
    for k in common_pos:
        m1_col, raw_pnum = m1_pos[k]
        m2_col, _ = m2_pos[k]
        matched.append(
            _make_matched(
                proj_number=raw_pnum,
                name=str(m1_projects[m1_col].get("name") or "Unnamed").strip(),
                m1_col=m1_col,
                m2_col=m2_col,
                match_source="positional",
            )
        )
    if matched:
        logger.info(
            "Project # matching yielded 0 results; positional fallback matched %d projects.",
            len(matched),
        )
        return matched

    # Fallback 2: canonical name equality
    m1_names = _build_name_match_index(m1_projects)
    m2_names = _build_name_match_index(m2_projects)
    common_names = sorted(set(m1_names.keys()) & set(m2_names.keys()))
    for i, canon in enumerate(common_names, start=1):
        m1_col = m1_names[canon]
        name = str(m1_projects[m1_col].get("name") or "").strip()
        matched.append(
            _make_matched(
                proj_number=i,
                name=name,
                m1_col=m1_col,
                m2_col=m2_names[canon],
                match_source="name",
            )
        )
    if matched:
        logger.info(
            "Project # matching yielded 0 results; name-based fallback matched %d projects.",
            len(matched),
        )
    return matched


# ---------------------------------------------------------------------------
# Metrics extraction
# ---------------------------------------------------------------------------


def extract_metrics(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Extract NPP, IRR, MWdc per matched project per model.

    Also runs M1-side audit and records the verdict (CLEAN / REVIEW /
    REWORK REQUIRED) so the Anchor section can surface review status
    alongside NPP/IRR. M2 is the comparison base; its verdict would be
    symmetric and adds no new signal.
    """
    results = []
    for m in matched:
        m1_data = m1_projects[m["m1_col"]]["data"]
        m2_data = m2_projects[m["m2_col"]]["data"]
        try:
            audit = audit_project(m1_data)
            m1_verdict = verdict_from_summary(audit.get("summary", {}))
        except Exception as exc:
            logger.warning(
                "audit_project failed for %r: %s — defaulting to REVIEW", m.get("name"), exc
            )
            m1_verdict = "REVIEW"
        results.append(
            {
                "proj_number": m["proj_number"],
                "name": m["name"],
                "mwdc": safe_float(m1_data.get(ROW_DC_MW)) or 0,
                "m1_npp": safe_float(m1_data.get(ROW_NPP)),
                "m1_irr": safe_float(m1_data.get(ROW_LEVERED_PT_IRR)),
                "m2_npp": safe_float(m2_data.get(ROW_NPP)),
                "m2_irr": safe_float(m2_data.get(ROW_LEVERED_PT_IRR)),
                "match_source": m.get("match_source", "proj_num"),
                "m1_verdict": m1_verdict,
            }
        )
    return results


# ---------------------------------------------------------------------------
# Input diff detection
# ---------------------------------------------------------------------------

# Rate Component field groupings for Pass 1a.
# SHAPE: always compared — describes what the component IS.
# VALUE: compared only when the RC is on (equity_on != 0) in BOTH models.
# TOGGLE: always compared — scope of the revenue stream.
_RC_SHAPE_FIELDS = [
    ("name", "Name", "text"),
    ("custom_generic", "Custom/Generic", "text"),
]
_RC_VALUE_FIELDS = [
    ("energy_rate", "Energy Rate", "$/kWh"),
    ("escalator", "Escalator", "%"),
    ("start_date", "Start Date", "date"),
    ("term", "Term", "yrs"),
    ("discount", "Customer Discount", "%"),
    ("ucb_fee", "UCB Fee", "%"),
]
_RC_TOGGLE_FIELDS = [
    ("equity_on", "Equity", "toggle"),
    ("debt_on", "Debt", "toggle"),
    ("appraisal_on", "Appraisal", "toggle"),
]

# Special (non-row-number) data keys that also represent walk-relevant inputs.
_SPECIAL_KEYS: list[tuple[str, str, str, str]] = [
    # (data_key, label, unit, category)
    ("_debt_match_equity", "Debt Rate: match equity", "toggle", "Financing"),
    ("_appraisal_match_equity", "Appraisal Rate: match equity", "toggle", "Financing"),
    ("_front_back_toggle", "Front/Back Debt toggle", "toggle", "Financing"),
    ("_debt_sizing_method", "Debt Sizing Method", "text", "Financing"),
]


def _rc_on_both_sides(rc1: dict, rc2: dict) -> bool:
    """RC value-field gate: both models must have equity_on truthy."""

    def _on(v):
        return bool(safe_float(v)) if not isinstance(v, bool) else v

    return _on(rc1.get("equity_on")) and _on(rc2.get("equity_on"))


def _diff_canonical_rows(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
    seen_labels: set[str],
    units_by_row: dict[int, str],
) -> list[dict]:
    """Pass 1: Canonical INPUT_ROW_LABELS rows with unit reconciliation.

    Mutates seen_labels so Pass 2 can skip labels already covered here.
    """
    diffs: list[dict] = []
    for row_num, label in INPUT_ROW_LABELS.items():
        if row_num in _SKIP_ROWS:
            continue
        if label.strip().lower() in _WALK_EXCLUDE_LABELS:
            continue

        is_text = row_num in TEXT_ROWS or row_num in DATE_ROWS
        per_project: dict[int, tuple[Any, Any]] = {}
        n_diff = 0
        row_unit_note: str | None = None

        for m in matched:
            m1_val = m1_projects[m["m1_col"]]["data"].get(row_num)
            m2_val = m2_projects[m["m2_col"]]["data"].get(row_num)

            if not is_text and row_num in _UNIT_MISMATCH_ROWS:
                m1_y = safe_float(m1_projects[m["m1_col"]]["data"].get(14))
                m2_y = safe_float(m2_projects[m["m2_col"]]["data"].get(14))
                m1_val, m2_val, note = _reconcile_unit_sensitive(
                    row_num,
                    m1_val,
                    m2_val,
                    m1_y,
                    m2_y,
                )
                if note and row_unit_note is None:
                    row_unit_note = note

            if (
                not is_text
                and _equality_numeric(m1_val) is None
                and _equality_numeric(m2_val) is None
            ):
                continue

            if _values_differ(m1_val, m2_val, is_text=is_text):
                n_diff += 1
            per_project[m["proj_number"]] = (m1_val, m2_val)

        if n_diff > 0 and per_project:
            unit = units_by_row.get(row_num) or INPUT_ROW_UNITS.get(row_num, "")
            if row_unit_note and row_num in _UNIT_MISMATCH_ROWS:
                unit = _UNIT_MISMATCH_ROWS[row_num]["displayed_unit"]
            diffs.append(
                {
                    "row": row_num,
                    "label": label,
                    "unit": unit,
                    "category": _categorize_row(row_num),
                    "values": per_project,
                    "n_diff": n_diff,
                    "n_total": len(per_project),
                    "source": "canonical",
                    **({"unit_note": row_unit_note} if row_unit_note else {}),
                }
            )
        seen_labels.add(label.strip().lower())
    return diffs


def _diff_rate_comps(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Pass 1a: per-component / per-field rate-component diffs for RC1-6.

    INPUT_ROW_LABELS covers only RC1 + partial RC2. Walks rate_comps directly
    so RC3-6 (community solar adders, REC streams) don't fall through.
    """
    diffs: list[dict] = []
    for rc_idx in range(1, 7):
        for field, label_suffix, unit in _RC_SHAPE_FIELDS + _RC_VALUE_FIELDS + _RC_TOGGLE_FIELDS:
            is_value = (field, label_suffix, unit) in _RC_VALUE_FIELDS
            is_text = unit in ("text", "date")
            per_project: dict[int, tuple[Any, Any]] = {}
            n_diff = 0
            for m in matched:
                rc1 = (m1_projects[m["m1_col"]].get("rate_comps") or {}).get(rc_idx) or {}
                rc2 = (m2_projects[m["m2_col"]].get("rate_comps") or {}).get(rc_idx) or {}
                if is_value and not _rc_on_both_sides(rc1, rc2):
                    continue
                m1_val = rc1.get(field)
                m2_val = rc2.get(field)
                if m1_val is None and m2_val is None:
                    continue
                if _values_differ(m1_val, m2_val, is_text=is_text):
                    n_diff += 1
                per_project[m["proj_number"]] = (m1_val, m2_val)
            if n_diff > 0 and per_project:
                diffs.append(
                    {
                        "row": 0,
                        "label": f"RC{rc_idx} {label_suffix}",
                        "unit": unit,
                        "category": "Revenue",
                        "values": per_project,
                        "n_diff": n_diff,
                        "n_total": len(per_project),
                        "source": "rate_comps",
                    }
                )
    return diffs


def _diff_rate_curves_cod(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Pass 1c: Rate Curves COD-period rate for Custom-Custom RC pairs.

    When both sides have Custom RC, Project Inputs energy_rate is blank —
    the revenue driver lives on the Rate Curves tab. Compare the $/kWh at
    each project's COD period to surface per-curve drift.
    """
    diffs: list[dict] = []
    for rc_idx in range(1, 7):
        per_project: dict[int, tuple[Any, Any]] = {}
        n_diff = 0
        n_extrapolated = 0
        for m in matched:
            m1p = m1_projects[m["m1_col"]]
            m2p = m2_projects[m["m2_col"]]
            rc1 = (m1p.get("rate_comps") or {}).get(rc_idx) or {}
            rc2 = (m2p.get("rate_comps") or {}).get(rc_idx) or {}
            m1_custom = str(rc1.get("custom_generic") or "").strip().lower() == "custom"
            m2_custom = str(rc2.get("custom_generic") or "").strip().lower() == "custom"
            if not (m1_custom and m2_custom):
                continue
            m1_curve = m1p.get(f"_rate_curves_rc{rc_idx}") or {}
            m2_curve = m2p.get(f"_rate_curves_rc{rc_idx}") or {}
            m1_rate, m1_conf = _rate_at_cod(m1_curve, m1p.get("data") or {})
            m2_rate, m2_conf = _rate_at_cod(m2_curve, m2p.get("data") or {})
            if m1_rate is None and m2_rate is None:
                continue
            if _values_differ(m1_rate, m2_rate):
                n_diff += 1
            # Count this project as non-exact if EITHER side had to extrapolate;
            # reviewers can then read the Notes column to know the rate may
            # not reflect the true COD-period value.
            if (m1_conf not in ("exact", None)) or (m2_conf not in ("exact", None)):
                n_extrapolated += 1
            per_project[m["proj_number"]] = (m1_rate, m2_rate)
        if n_diff > 0 and per_project:
            diffs.append(
                {
                    "row": 0,
                    "label": f"RC{rc_idx} Rate Curve (COD)",
                    "unit": "$/kWh",
                    "category": "Revenue",
                    "values": per_project,
                    "n_diff": n_diff,
                    "n_total": len(per_project),
                    "source": "rate_curve",
                    **({"extrapolated_count": n_extrapolated} if n_extrapolated else {}),
                }
            )
    return diffs


def _diff_special_keys(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Pass 1b: match toggles, debt sizing method, and DSCR schedule Y1-10.

    These live under non-row-number data keys and on proj["dscr_schedule"]
    — invisible to INPUT_ROW_LABELS iteration.
    """
    diffs: list[dict] = []
    for key, label, unit, category in _SPECIAL_KEYS:
        per_project: dict[int, tuple[Any, Any]] = {}
        n_diff = 0
        for m in matched:
            m1_val = m1_projects[m["m1_col"]]["data"].get(key)
            m2_val = m2_projects[m["m2_col"]]["data"].get(key)
            if m1_val is None and m2_val is None:
                continue
            if _values_differ(m1_val, m2_val, is_text=(unit == "text")):
                n_diff += 1
            per_project[m["proj_number"]] = (m1_val, m2_val)
        if n_diff > 0 and per_project:
            diffs.append(
                {
                    "row": 0,
                    "label": label,
                    "unit": unit,
                    "category": category,
                    "values": per_project,
                    "n_diff": n_diff,
                    "n_total": len(per_project),
                    "source": "special",
                }
            )

    # DSCR years 1-10 — practical debt-sizing horizon.
    for year in range(1, 11):
        per_project = {}
        n_diff = 0
        for m in matched:
            m1_sched = m1_projects[m["m1_col"]].get("dscr_schedule") or {}
            m2_sched = m2_projects[m["m2_col"]].get("dscr_schedule") or {}
            m1_val = m1_sched.get(year)
            m2_val = m2_sched.get(year)
            if m1_val is None and m2_val is None:
                continue
            if _values_differ(m1_val, m2_val):
                n_diff += 1
            per_project[m["proj_number"]] = (m1_val, m2_val)
        if n_diff > 0 and per_project:
            diffs.append(
                {
                    "row": 0,
                    "label": f"DSCR Y{year}",
                    "unit": "x",
                    "category": "Financing",
                    "values": per_project,
                    "n_diff": n_diff,
                    "n_total": len(per_project),
                    "source": "special",
                }
            )
    return diffs


def _proptax_label_excluded(label_lower: str, m1_all: dict, m2_all: dict) -> bool:
    """Property Tax label gate for Pass 2: drop Y2-Y5 entirely; include
    Y1/escalator only when the Custom PropTax toggle is on in either model.
    """
    if "property tax" not in label_lower:
        return False
    # Skip Y2+ entirely.
    if any(
        x in label_lower
        for x in [
            "year 2",
            "year 3",
            "year 4",
            "year 5",
            "yr 2",
            "yr 3",
            "yr 4",
            "yr 5",
            "y2",
            "y3",
            "y4",
            "y5",
        ]
    ):
        return True
    if "toggle" in label_lower:
        return False
    m1_toggle = m1_all.get("Custom Property Tax Schedule Toggle (On/Off)") or m1_all.get(
        "Custom PropTax Toggle"
    )
    m2_toggle = m2_all.get("Custom Property Tax Schedule Toggle (On/Off)") or m2_all.get(
        "Custom PropTax Toggle"
    )
    toggle_on = any(
        str(t).strip().lower() in ("1", "on", "true", "yes") or (safe_float(t) or 0) != 0
        for t in [m1_toggle, m2_toggle]
        if t is not None
    )
    return not toggle_on


def _diff_all_inputs_labels(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
    seen_labels: set[str],
    units_by_label: dict[str, str],
) -> list[dict]:
    """Pass 2: label-based scan of _all_inputs — catches non-canonical rows
    that Pass 1 missed (template additions, developer-added lines, etc.).
    """
    if not matched:
        return []
    first_m = matched[0]
    m1_all = m1_projects[first_m["m1_col"]]["data"].get("_all_inputs", {})
    m2_all = m2_projects[first_m["m2_col"]]["data"].get("_all_inputs", {})
    all_labels = set(m1_all.keys()) | set(m2_all.keys())

    diffs: list[dict] = []
    for label in sorted(all_labels):
        if not label or not label.strip():
            continue
        label_lower = label.strip().lower()
        if label_lower in seen_labels:
            continue
        if label_lower in _WALK_EXCLUDE_LABELS:
            continue
        if _proptax_label_excluded(label_lower, m1_all, m2_all):
            continue

        per_project: dict[int, tuple[Any, Any]] = {}
        n_diff = 0
        for m in matched:
            m1_inputs = m1_projects[m["m1_col"]]["data"].get("_all_inputs", {})
            m2_inputs = m2_projects[m["m2_col"]]["data"].get("_all_inputs", {})
            m1_val = m1_inputs.get(label)
            m2_val = m2_inputs.get(label)
            if m1_val is None and m2_val is None:
                continue
            if _values_differ(m1_val, m2_val):
                n_diff += 1
            per_project[m["proj_number"]] = (m1_val, m2_val)

        if n_diff > 0 and per_project:
            diffs.append(
                {
                    "row": 0,
                    "label": label,
                    "unit": units_by_label.get(label, ""),
                    "category": "Other",
                    "values": per_project,
                    "n_diff": n_diff,
                    "n_total": len(per_project),
                    "source": "label",
                }
            )
    return diffs


def diff_inputs(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Find all per-project differences across the two models.

    Five passes:
      1  canonical rows (INPUT_ROW_LABELS) with unit reconciliation
      1a rate components 1-6 (shape / value / toggle fields)
      1c rate curves COD-period rate (Custom-Custom pairs only)
      1b match toggles, debt sizing method, DSCR schedule years 1-10
      2  _all_inputs label scan for non-canonical rows

    Returns list of {row, label, unit, category, values, n_diff, n_total}.
    """
    units_by_row: dict[int, str] = {}
    units_by_label: dict[str, str] = {}
    if matched:
        first_data = m1_projects[matched[0]["m1_col"]]["data"]
        units_by_row = first_data.get("_units_by_row") or {}
        units_by_label = first_data.get("_all_units") or {}

    seen_labels: set[str] = set()
    return [
        *_diff_canonical_rows(matched, m1_projects, m2_projects, seen_labels, units_by_row),
        *_diff_rate_comps(matched, m1_projects, m2_projects),
        *_diff_rate_curves_cod(matched, m1_projects, m2_projects),
        *_diff_special_keys(matched, m1_projects, m2_projects),
        *_diff_all_inputs_labels(matched, m1_projects, m2_projects, seen_labels, units_by_label),
    ]


# ---------------------------------------------------------------------------
# Number format selection
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Walk sheet layout constants (shared by anchor + variance section writers)
# ---------------------------------------------------------------------------

# Column layout: B=proj name, C=MWdc, D=spacer.
# Cases start at col E (5). Each case = 3 cols: NPP, IRR, ∆ Base.
# E(5)=NPP1, F(6)=IRR1, G(7)=delta, H(8)=NPP2, I(9)=IRR2, J(10)=Notes (variance only).
_DELTA_COL = 7
_WALK_LAST_COL = 9


def _case_cols(case_idx: int) -> tuple[int, int, int]:
    """Return (npp_col, irr_col, delta_col) for a case (0-indexed)."""
    return (5, 6, _DELTA_COL) if case_idx == 0 else (8, 9, _DELTA_COL)


# ---------------------------------------------------------------------------
# Sheet writers — each populates one region. build_walk_xlsx orchestrates.
# ---------------------------------------------------------------------------


def _write_anchor_section(
    ws,
    metrics: list[dict],
    case_labels: list[str],
    n_cases: int = 2,
) -> int:
    """Write the top NPP/IRR table (header rows 3-6, data rows 7+, summary
    row). Returns the summary row number (caller uses it to anchor the
    variance section that follows).
    """
    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    # Column D used to be an empty spacer; it now holds the M1-side review
    # verdict (CLEAN / REVIEW / REWORK REQUIRED) with color-coded fill.
    # Widen enough to fit the longest label.
    ws.column_dimensions["D"].width = 18
    for ci in range(5, _WALK_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions["J"].width = 26  # Notes column (variance-only)

    # Row 3: case numbers with double-bottom borders across the band.
    ws.merge_cells("E3:F3")
    ws.cell(row=3, column=5, value=1).font = NORMAL_FONT
    ws.cell(row=3, column=5).alignment = CENTER
    for c in range(5, 7):
        ws.cell(row=3, column=c).border = DOUBLE_BOTTOM
    ws.cell(row=3, column=7).border = DOUBLE_BOTTOM
    ws.merge_cells("H3:I3")
    ws.cell(row=3, column=8, value=2).font = NORMAL_FONT
    ws.cell(row=3, column=8).alignment = CENTER
    for c in range(7, 10):
        ws.cell(row=3, column=c).border = DOUBLE_BOTTOM

    # Row 5: case labels
    ws.merge_cells("E5:F5")
    c = ws.cell(row=5, column=5, value=case_labels[0])
    c.font = BOLD_FONT
    c.alignment = CENTER_WRAP
    c.border = THIN_BOTTOM
    ws.cell(row=5, column=6).border = THIN_BOTTOM
    ws.cell(row=5, column=7).border = THIN_BOTTOM
    ws.merge_cells("H5:I5")
    c = ws.cell(row=5, column=8, value=case_labels[1])
    c.font = BOLD_FONT
    c.alignment = CENTER_WRAP
    c.border = THIN_BOTTOM
    ws.cell(row=5, column=9).border = THIN_BOTTOM

    # Row 6: fixed headers + per-case NPP/IRR headers + delta
    for col, text in [(2, "Project Name"), (3, "MWdc"), (4, "M1 Verdict")]:
        cell = ws.cell(row=6, column=col, value=text)
        cell.font = BOLD_FONT if col == 2 else NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = THIN_BOTTOM
        cell.alignment = LEFT if col == 2 else CENTER
    for npp_c, irr_c, lbl_npp, lbl_irr in [
        (5, 6, "NPP ($/W)", "IRR (%)"),
        (8, 9, "NPP ($/W)", "IRR (%)"),
    ]:
        cell = ws.cell(row=6, column=npp_c, value=lbl_npp)
        cell.font = NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = HDR_NPP
        cell.alignment = CENTER
        cell = ws.cell(row=6, column=irr_c, value=lbl_irr)
        cell.font = NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = THIN_BOTTOM
        cell.alignment = CENTER
    cell = ws.cell(row=6, column=_DELTA_COL, value="\u2206 Base")
    cell.font = BLUE_FONT
    cell.fill = GREY_FILL
    cell.border = HDR_DELTA
    cell.alignment = CENTER

    # Data rows 7+
    data_start = 7
    base_npp_letter = get_column_letter(_case_cols(0)[0])
    for pi, pm in enumerate(metrics):
        r = data_start + pi
        is_last = pi == len(metrics) - 1

        cell = ws.cell(row=r, column=2, value=pm["name"])
        cell.font = WHITE_BOLD
        cell.fill = NAVY_FILL
        if is_last:
            cell.border = THIN_BOTTOM
        # Non-standard pairing? attach a hover-comment so reviewers can
        # audit how the pair was formed. Doesn't clutter the print view.
        match_source = pm.get("match_source", "proj_num")
        if match_source != "proj_num":
            cell.comment = Comment(
                f"Matched via {match_source} fallback — verify the pair "
                f"is correct. Standard matching uses Project # (row 2).",
                "walk_builder",
            )

        cell = ws.cell(row=r, column=3, value=pm["mwdc"])
        cell.number_format = FMT_MW
        cell.alignment = CENTER
        if is_last:
            cell.border = THIN_BOTTOM

        # Verdict cell — color-coded fill matches review-panel semantics.
        verdict = pm.get("m1_verdict", "REVIEW")
        v_cell = ws.cell(row=r, column=4, value=verdict)
        v_cell.alignment = CENTER
        v_cell.font = BOLD_FONT
        fill = _VERDICT_FILLS.get(verdict)
        if fill is not None:
            v_cell.fill = fill
        if is_last:
            v_cell.border = THIN_BOTTOM

        case_vals = [(pm["m1_npp"], pm["m1_irr"]), (pm["m2_npp"], pm["m2_irr"])]
        for ci, (npp_val, irr_val) in enumerate(case_vals):
            npp_c, irr_c, delta_c = _case_cols(ci)

            cell = ws.cell(row=r, column=npp_c, value=npp_val)
            cell.number_format = FMT_NPP
            cell.alignment = CENTER
            cell.border = THIN_BOTTOM_LEFT if is_last else THIN_LEFT

            cell = ws.cell(row=r, column=irr_c, value=irr_val)
            cell.number_format = FMT_IRR
            cell.alignment = CENTER
            if is_last:
                cell.border = THIN_BOTTOM

            # Delta (only for non-base cases). Δ direction = M1 - M2.
            if ci > 0:
                npp_letter = get_column_letter(npp_c)
                cell = ws.cell(
                    row=r, column=delta_c, value=f"={base_npp_letter}{r}-{npp_letter}{r}"
                )
                cell.number_format = FMT_DELTA
                cell.alignment = CENTER
                cell.border = THIN_BOTTOM_LEFT_RIGHT if is_last else THIN_LEFT_RIGHT

    # Summary row: MW-weighted averages via SUMPRODUCT.
    summary_r = data_start + len(metrics)
    last_data_r = data_start + len(metrics) - 1
    if metrics:
        cell = ws.cell(row=summary_r, column=3, value=f"=SUM(C{data_start}:C{last_data_r})")
        cell.number_format = FMT_MW
        cell.alignment = CENTER
        for ci in range(n_cases):
            npp_c, irr_c, _ = _case_cols(ci)
            npp_l = get_column_letter(npp_c)
            irr_l = get_column_letter(irr_c)
            mw_wgt_fmt = (
                f"=SUMPRODUCT({{col}}{data_start}:{{col}}{last_data_r},"
                f"$C${data_start}:$C${last_data_r})"
                f"/SUM($C${data_start}:$C${last_data_r})"
            )
            cell = ws.cell(row=summary_r, column=npp_c, value=mw_wgt_fmt.format(col=npp_l))
            cell.number_format = FMT_NPP
            cell.alignment = CENTER
            cell = ws.cell(row=summary_r, column=irr_c, value=mw_wgt_fmt.format(col=irr_l))
            cell.number_format = FMT_IRR
            cell.alignment = CENTER
    return summary_r


def _aggregate_variance_values(
    v: dict,
    metrics: list[dict],
    total_mw: float,
) -> tuple[Any, Any, bool]:
    """For a single variance row, compute the display values across all
    matched projects: MW-weighted average for numerics, first-project for
    text. Returns (m1_display, m2_display, is_text_val).
    """
    m1_sum = m2_sum = 0.0
    m1_count = m2_count = 0
    is_text_val = False
    first_m1 = first_m2 = None
    for pnum, (m1v, m2v) in v["values"].items():
        mw = next((pm["mwdc"] for pm in metrics if pm["proj_number"] == pnum), 1.0)
        f1 = safe_float(m1v)
        f2 = safe_float(m2v)
        if f1 is not None:
            m1_sum += f1 * mw
            m1_count += 1
        elif m1v is not None:
            is_text_val = True
            if first_m1 is None:
                first_m1 = m1v
        if f2 is not None:
            m2_sum += f2 * mw
            m2_count += 1
        elif m2v is not None:
            is_text_val = True
            if first_m2 is None:
                first_m2 = m2v

    if is_text_val:
        return first_m1, first_m2, True
    m1_display = m1_sum / total_mw if m1_count else None
    m2_display = m2_sum / total_mw if m2_count else None
    return m1_display, m2_display, False


def _write_variance_section(
    ws,
    grouped: dict[str, list[dict]],
    metrics: list[dict],
    var_start: int,
    m1_data_by_pnum: dict | None = None,
) -> int:
    """Write the 'Project Inputs' variance drivers block starting at
    var_start. Returns the row after the last written diff.

    m1_data_by_pnum: optional {proj_number: m1_data_dict} for $-impact
        computation. Omit to suppress the Impact column (e.g., legacy
        callers that don't pass project data through).
    """
    cell = ws.cell(row=var_start, column=2, value="Project Inputs")
    cell.font = NORMAL_FONT
    cell.border = DOUBLE_BOTTOM
    for vc in range(3, _WALK_LAST_COL + 1):
        ws.cell(row=var_start, column=vc).border = DOUBLE_BOTTOM

    cur_row = var_start + 1
    ws.cell(row=cur_row, column=3, value="Unit").font = Font(size=10, color="7d8694")
    ws.cell(row=cur_row, column=3).alignment = CENTER
    notes_hdr = ws.cell(row=cur_row, column=10, value="Notes")
    notes_hdr.font = Font(size=10, color="7d8694")
    notes_hdr.alignment = LEFT
    src_hdr = ws.cell(row=cur_row, column=11, value="Source")
    src_hdr.font = Font(size=10, color="7d8694")
    src_hdr.alignment = CENTER
    imp_hdr = ws.cell(row=cur_row, column=12, value="$ Impact")
    imp_hdr.font = Font(size=10, color="7d8694")
    imp_hdr.alignment = CENTER
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 14
    cur_row += 1

    total_mw = sum(pm["mwdc"] for pm in metrics) or 1.0

    for cat_name in _CATEGORY_ORDER:
        cat_vars = grouped.get(cat_name)
        if not cat_vars:
            continue
        cell = ws.cell(row=cur_row, column=2, value=cat_name)
        cell.font = BOLD_FONT
        cell.border = DOUBLE_BOTTOM
        cur_row += 1

        for v in sorted(cat_vars, key=lambda x: x["label"].lower()):
            _write_variance_row(ws, cur_row, v, metrics, total_mw, m1_data_by_pnum)
            cur_row += 1
    return cur_row


def _write_variance_row(
    ws,
    r: int,
    v: dict,
    metrics: list[dict],
    total_mw: float,
    m1_data_by_pnum: dict | None,
) -> None:
    """Render a single variance row: label, unit, M1/Δ/M2 values, Notes,
    Source, $ Impact. Extracted to keep _write_variance_section under the
    complexity threshold as columns get added tranche-over-tranche."""
    nfmt = _num_format(v["row"])
    cell = ws.cell(row=r, column=2, value=v["label"])
    cell.font = NORMAL_FONT
    cell.alignment = LEFT

    if v["unit"]:
        cell = ws.cell(row=r, column=3, value=v["unit"])
        cell.font = Font(color="7d8694", size=10, italic=True)
        cell.alignment = CENTER

    m1_display, m2_display, is_text_val = _aggregate_variance_values(v, metrics, total_mw)

    c_e = ws.cell(row=r, column=5, value=m1_display)
    if not is_text_val and m1_display is not None:
        c_e.number_format = nfmt
    c_e.alignment = CENTER_CONT
    c_e.border = THIN_BOX

    if not is_text_val:
        delta_cell = ws.cell(row=r, column=7, value=f"=E{r}-H{r}")
        delta_cell.number_format = FMT_DELTA
        delta_cell.alignment = CENTER
        delta_cell.border = THIN_BOX

    c_h = ws.cell(row=r, column=8, value=m2_display)
    if not is_text_val and m2_display is not None:
        c_h.number_format = nfmt
    c_h.alignment = CENTER_CONT
    c_h.border = THIN_BOX

    if m1_display != m2_display:
        f1 = safe_float(m1_display)
        f2 = safe_float(m2_display)
        if (
            f1 is not None
            and f2 is not None
            and abs(f1 - f2) > 1e-6
            or is_text_val
            and str(m1_display or "") != str(m2_display or "")
        ):
            c_h.fill = YELLOW_FILL

    n_diff = v.get("n_diff", 0)
    n_total = v.get("n_total", len(v.get("values", {})))
    if n_diff:
        note_text = f"differs for {n_diff} of {n_total} projects"
        extrapolated = v.get("extrapolated_count", 0)
        if extrapolated:
            note_text += f" · rate extrapolated: {extrapolated}"
        notes_cell = ws.cell(row=r, column=10, value=note_text)
        notes_cell.font = Font(size=10, color="7d8694", italic=True)
        notes_cell.alignment = LEFT

    src = v.get("source")
    if src:
        src_cell = ws.cell(row=r, column=11, value=src)
        src_cell.font = Font(size=9, color="7d8694", italic=True)
        src_cell.alignment = CENTER

    if m1_data_by_pnum is not None:
        impact_val = portfolio_impact(v.get("row", 0), v.get("values", {}), m1_data_by_pnum)
        if impact_val is not None:
            imp_cell = ws.cell(row=r, column=12, value=impact_val)
            imp_cell.number_format = '$#,##0_);[Red]($#,##0);"-"'
            imp_cell.alignment = CENTER


_REASON_DETAIL = {
    "missing_proj_num": "This model has no Project # assigned (row 2 blank).",
    "proj_num_not_in_other": "Project # exists here but the counterpart model has no matching Project #.",
    "name_not_in_other": "Project # exists but the project name doesn't match the counterpart.",
    "unknown": "Unable to classify — inspect both rows manually.",
}


def _write_unmatched_sheet(
    wb,
    unmatched_m1: list,
    unmatched_m2: list,
    m1_label: str,
    m2_label: str,
) -> None:
    """Create the Unmatched sheet listing orphan projects from each side
    with reason codes. Called only when at least one orphan exists."""
    un = wb.create_sheet("Unmatched")
    un.sheet_view.showGridLines = False
    for col_letter, width in zip("ABCDEF", (8, 32, 12, 10, 24, 44), strict=True):
        un.column_dimensions[col_letter].width = width
    headers = ["Side", "Project Name", "Project #", "MWdc", "Reason Code", "Detail"]
    for c, text in enumerate(headers, start=1):
        cell = un.cell(row=1, column=c, value=text)
        cell.font = WHITE_BOLD
        cell.fill = NAVY_FILL
        cell.alignment = CENTER

    r = 2
    for side, projects in [(m1_label or "M1", unmatched_m1), (m2_label or "M2", unmatched_m2)]:
        for _col, proj, reason in projects:
            data = proj.get("data", {}) or {}
            un.cell(row=r, column=1, value=side).alignment = CENTER
            un.cell(row=r, column=2, value=str(proj.get("name") or "").strip()).alignment = LEFT
            un.cell(
                row=r, column=3, value=safe_float(data.get(ROW_PROJECT_NUMBER))
            ).alignment = CENTER
            mwdc_cell = un.cell(row=r, column=4, value=safe_float(data.get(ROW_DC_MW)))
            mwdc_cell.number_format = FMT_MW
            mwdc_cell.alignment = CENTER
            un.cell(row=r, column=5, value=reason).alignment = CENTER
            un.cell(row=r, column=6, value=_REASON_DETAIL.get(reason, reason)).alignment = LEFT
            r += 1


def _num_format(row: int) -> str:
    if row in PCT_ROWS:
        return FMT_PCT
    if row in DPW_ROWS:
        return FMT_DPW
    if row in INT_ROWS:
        return FMT_INT
    return FMT_DPW  # default to 3-decimal


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------


def build_walk_xlsx(
    m1_result: dict,
    m2_result: dict,
    m1_label: str,
    m2_label: str,
    include_proj_numbers: set[int] | None = None,
    include_proj_names: set[str] | None = None,
) -> tuple[io.BytesIO, dict]:
    """Build a formatted Walk Summary .xlsx comparing two models.

    Args:
        include_proj_numbers: If provided, include projects whose proj_number
            is in this set. Union'd with include_proj_names.
        include_proj_names: If provided, include projects whose name (case /
            whitespace tolerant) is in this set. Acts as a fallback when the
            project review panel returns projects with null projNumber.

    A project is kept if it matches EITHER the number set OR the name set.
    When both filters are None, all matched projects are kept (no filter).

    Returns (BytesIO with xlsx data, summary dict for UI).
    """
    m1_projects = get_projects(m1_result) or {}
    m2_projects = get_projects(m2_result) or {}

    all_matched = match_projects(m1_projects, m2_projects)

    # Apply the include filters. Union semantics: a project passes if its
    # proj_number OR its name is in the allowed set. This is what the review
    # panel sends when some rows have a valid project# and others don't.
    canon_names = (
        {canonicalize_name(n) for n in include_proj_names} if include_proj_names else set()
    )
    if include_proj_numbers is not None or include_proj_names is not None:
        matched = [
            m
            for m in all_matched
            if (include_proj_numbers and m["proj_number"] in include_proj_numbers)
            or (canon_names and canonicalize_name(m["name"]) in canon_names)
        ]
    else:
        matched = list(all_matched)

    # Filter out template placeholders ("Project 15", "Anchor", etc.) via
    # the module-level _is_placeholder helper.
    matched = [m for m in matched if not _is_placeholder(m["name"])]

    # Compute true orphans with reason codes. Use the unfiltered match set
    # so a deselected-but-matched project isn't mistaken for an orphan.
    matched_m1_cols = {m["m1_col"] for m in all_matched}
    matched_m2_cols = {m["m2_col"] for m in all_matched}

    m1_pnums = _build_pnum_set(m1_projects)
    m2_pnums = _build_pnum_set(m2_projects)
    m1_names = _build_name_set(m1_projects)
    m2_names = _build_name_set(m2_projects)

    unmatched_m1 = [
        (col, p, _orphan_reason(p, m2_pnums, m2_names))
        for col, p in m1_projects.items()
        if isinstance(p, dict)
        and col not in matched_m1_cols
        and p.get("name")
        and not _is_placeholder(str(p.get("name") or ""))
    ]
    unmatched_m2 = [
        (col, p, _orphan_reason(p, m1_pnums, m1_names))
        for col, p in m2_projects.items()
        if isinstance(p, dict)
        and col not in matched_m2_cols
        and p.get("name")
        and not _is_placeholder(str(p.get("name") or ""))
    ]

    if not matched:
        logger.warning("No projects matched between the two models by Project #.")
        # Flag so we can write an explanatory message into the xlsx below
        _no_matches = True
    else:
        _no_matches = False

    metrics = extract_metrics(matched, m1_projects, m2_projects)
    variances = diff_inputs(matched, m1_projects, m2_projects)

    # Group variances by category
    grouped: dict[str, list[dict]] = {}
    for v in variances:
        grouped.setdefault(v["category"], []).append(v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Build Walk"
    ws.sheet_view.showGridLines = False  # Clean look matching reference files

    case_labels = [m1_label, m2_label]

    # ===================================================================
    # TOP SECTION: NPP / IRR comparison table
    # ===================================================================
    summary_r = _write_anchor_section(ws, metrics, case_labels)

    # Build proj_number → M1 data dict for $impact computation.
    m1_data_by_pnum = {m["proj_number"]: m1_projects[m["m1_col"]].get("data", {}) for m in matched}

    # Variance drivers block starts 3 rows below the anchor summary row.
    _write_variance_section(
        ws,
        grouped,
        metrics,
        var_start=summary_r + 3,
        m1_data_by_pnum=m1_data_by_pnum,
    )

    if _no_matches:
        cell = ws.cell(
            row=3,
            column=2,
            value="No projects matched between models. Check that Project # "
            "(row 2) is populated in both models.",
        )
        cell.font = Font(color="FF0000", bold=True, size=11)

    if unmatched_m1 or unmatched_m2:
        _write_unmatched_sheet(wb, unmatched_m1, unmatched_m2, m1_label, m2_label)

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Build summary dict for the UI panel
    cats_found = [c for c in _CATEGORY_ORDER if c in grouped]
    matches_by_source: dict[str, int] = {"proj_num": 0, "positional": 0, "name": 0}
    for m in matched:
        key = m.get("match_source", "proj_num")
        matches_by_source[key] = matches_by_source.get(key, 0) + 1
    verdict_counts: dict[str, int] = {"CLEAN": 0, "REVIEW": 0, "REWORK REQUIRED": 0}
    for pm in metrics:
        vk = pm.get("m1_verdict", "REVIEW")
        verdict_counts[vk] = verdict_counts.get(vk, 0) + 1
    summary = {
        "n_matched": len(matched),
        "n_diffs": len(variances),
        "n_unmatched_m1": len(unmatched_m1),
        "n_unmatched_m2": len(unmatched_m2),
        "matches_by_source": matches_by_source,
        "verdict_counts": verdict_counts,
        "categories": cats_found,
        "m1_label": m1_label,
        "m2_label": m2_label,
    }

    return buf, summary
