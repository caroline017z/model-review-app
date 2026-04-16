"""
38DN Build Walk — Excel Walk Summary Generator

Compares two pricing models side-by-side, generating a formatted .xlsx file
that shows per-project NPP/IRR deltas and auto-detected variance drivers
grouped by category.

Output format matches existing 38DN Walk Summary files exactly:
  - Navy project name fills, white bold font
  - NPP in 0.000 accounting format with red negatives
  - IRR in 0.00% format
  - ∆ Base with Excel formulas (zero shows as dash)
  - Yellow highlights on changed values
  - MW-weighted SUMPRODUCT averages
  - Category-grouped variance drivers below
"""
from __future__ import annotations

import io
import logging
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from config import (
    BIBLE_BENCHMARKS, INPUT_ROW_LABELS, INPUT_ROW_UNITS,
    PCT_ROWS, DPW_ROWS, INT_ROWS, TEXT_ROWS, DATE_ROWS,
)
from data_loader import get_projects
from rows import ROW_PROJECT_NUMBER, ROW_DC_MW, ROW_NPP, ROW_FMV_IRR
from utils import safe_float

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Formatting constants (match reference Walk Summary files exactly)
# ---------------------------------------------------------------------------

NAVY_FILL = PatternFill("solid", fgColor="002060")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC")

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
BLUE_FONT = Font(color="0000FF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)

_THIN = Side(style="thin")
_DOUBLE = Side(style="double")
THIN_BOTTOM = Border(bottom=_THIN)
DOUBLE_BOTTOM = Border(bottom=_DOUBLE)
# Column separators for the NPP/IRR/delta grid
THIN_LEFT = Border(left=_THIN)
THIN_LEFT_RIGHT = Border(left=_THIN, right=_THIN)
THIN_BOTTOM_LEFT = Border(bottom=_THIN, left=_THIN)
THIN_BOTTOM_LEFT_RIGHT = Border(bottom=_THIN, left=_THIN, right=_THIN)
# Boxed cell for variance section values
THIN_BOX = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
# Header row: bottom + left separator
HDR_NPP = Border(bottom=_THIN, left=_THIN, top=_THIN)
HDR_DELTA = Border(bottom=_THIN, left=_THIN, right=_THIN, top=_THIN)

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER_CONT = Alignment(horizontal="centerContinuous", vertical="center")

FMT_MW = "0.00"
FMT_NPP = '0.000_);[Red]\\(0.000\\)'
FMT_IRR = "0.00%"
FMT_DELTA = '0.000_);[Red]\\(0.000\\);"-"'
FMT_DPW = "#,##0.000"
FMT_INT = "#,##0"
FMT_PCT = "0.00%"

# Rows to skip in variance detection (identity / metadata rows)
_SKIP_ROWS = {2, 4, 7, 8, 10, 18, 19}

# Labels to EXCLUDE from the walk variance section (covered elsewhere or outputs)
_WALK_EXCLUDE_LABELS = {
    "live appraisal model irr", "live appraisal irr",
    "max fmv w/ constraint", "max fmv with constraint",
    "minimum equity dscr multiple",
    "npp ($)", "npp ($/w) - solve", "npp ($/w)",
    "other upfront costs",
    "project toggle (on/off)", "toggle (on/off)",
    "step up dev fee - solve", "step up dev fee",
    "te pre-commitment amount for cl sizing",
    "tax equity insurance costs",
    "total capex excl. financing costs", "total capex excl. financing",
    "total capex incl. financing costs", "total capex incl. financing",
    "total itc",
    "unconstrained (calculated) fmv", "unconstrained fmv",
    "live levered pre-tax irr", "levered pre-tax irr",
    "active fmv", "active fair market value",
}

# Category display order
_CATEGORY_ORDER = [
    "CapEx", "System Sizing", "Revenue",
    "Incentives & Tax", "System Details", "OpEx", "Other",
]


# ---------------------------------------------------------------------------
# Row categorization
# ---------------------------------------------------------------------------

def _build_category_map() -> dict[int, str]:
    """Invert BIBLE_BENCHMARKS to get {row_number: category_name}."""
    cat_map: dict[int, str] = {}
    for category, checks in BIBLE_BENCHMARKS.items():
        for _label, spec in checks.items():
            if "row" in spec:
                cat_map[spec["row"]] = category
    return cat_map


_CATEGORY_MAP = _build_category_map()

# Fallback ranges for rows not in BIBLE_BENCHMARKS
_FALLBACK_RANGES = [
    (range(117, 130), "CapEx"),
    (range(143, 215), "Revenue"),
    (range(215, 222), "Incentives & Tax"),
    (range(225, 311), "OpEx"),
    (range(587, 611), "Incentives & Tax"),
]


def _categorize_row(row: int) -> str:
    if row in _CATEGORY_MAP:
        return _CATEGORY_MAP[row]
    for rng, cat in _FALLBACK_RANGES:
        if row in rng:
            return cat
    return "Other"


# ---------------------------------------------------------------------------
# Project matching
# ---------------------------------------------------------------------------

def match_projects(
    m1_projects: dict, m2_projects: dict,
) -> list[dict]:
    """Match projects between two models by Project # (row 2).

    Falls back to name-based matching or column-position-inferred project
    numbers when row 2 data is missing/unresolved.

    Returns list of {proj_number, name, m1_col, m2_col} sorted by proj_number.
    """
    def _build_index(projects: dict) -> dict[int, int]:
        idx: dict[int, int] = {}
        for col, proj in projects.items():
            if not isinstance(proj, dict) or "data" not in proj:
                continue
            pnum = safe_float(proj["data"].get(ROW_PROJECT_NUMBER))
            if pnum is not None:
                idx[int(pnum)] = col
        return idx

    m1_idx = _build_index(m1_projects)
    m2_idx = _build_index(m2_projects)

    # Inner join on proj_number
    common = sorted(set(m1_idx.keys()) & set(m2_idx.keys()))
    matched = []
    for pnum in common:
        m1_col = m1_idx[pnum]
        m1_proj = m1_projects[m1_col]
        matched.append({
            "proj_number": pnum,
            "name": str(m1_proj.get("name") or "Unnamed").strip(),
            "m1_col": m1_col,
            "m2_col": m2_idx[pnum],
        })

    if matched:
        return matched

    # ----- Fallback 1: infer project numbers from column position -----
    # Typical models have projects starting at column F (6), so col - 5 = project #.
    # Try this for projects whose data[ROW_PROJECT_NUMBER] is None.
    def _build_index_positional(projects: dict) -> dict[int, int]:
        idx: dict[int, int] = {}
        for col, proj in projects.items():
            if not isinstance(proj, dict) or "data" not in proj:
                continue
            pnum = safe_float(proj["data"].get(ROW_PROJECT_NUMBER))
            if pnum is not None:
                idx[int(pnum)] = col
            elif proj.get("name"):
                # Infer project number from column position (col F=6 -> #1)
                inferred = col - 5
                if inferred >= 1:
                    idx[inferred] = col
        return idx

    m1_pos = _build_index_positional(m1_projects)
    m2_pos = _build_index_positional(m2_projects)
    already_matched = {m["proj_number"] for m in matched}
    common_pos = sorted(set(m1_pos.keys()) & set(m2_pos.keys()))
    for pnum in common_pos:
        if pnum in already_matched:
            continue
        m1_col = m1_pos[pnum]
        m1_proj = m1_projects[m1_col]
        matched.append({
            "proj_number": pnum,
            "name": str(m1_proj.get("name") or "Unnamed").strip(),
            "m1_col": m1_col,
            "m2_col": m2_pos[pnum],
        })

    if matched:
        logger.info(
            "Project # matching yielded 0 results; positional fallback matched %d projects.",
            len(matched),
        )
        return matched

    # ----- Fallback 2: match by project name (exact string match) -----
    def _build_name_index(projects: dict) -> dict[str, int]:
        idx: dict[str, int] = {}
        for col, proj in projects.items():
            if not isinstance(proj, dict):
                continue
            name = str(proj.get("name") or "").strip()
            if name:
                idx[name] = col
        return idx

    m1_names = _build_name_index(m1_projects)
    m2_names = _build_name_index(m2_projects)
    common_names = sorted(set(m1_names.keys()) & set(m2_names.keys()))
    for i, name in enumerate(common_names, start=1):
        m1_col = m1_names[name]
        matched.append({
            "proj_number": i,
            "name": name,
            "m1_col": m1_col,
            "m2_col": m2_names[name],
        })

    if matched:
        logger.info(
            "Project # matching yielded 0 results; name-based fallback matched %d projects.",
            len(matched),
        )
    return matched


# ---------------------------------------------------------------------------
# Metrics extraction
# ---------------------------------------------------------------------------

def extract_metrics(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Extract NPP, IRR, MWdc per matched project per model."""
    results = []
    for m in matched:
        m1_data = m1_projects[m["m1_col"]]["data"]
        m2_data = m2_projects[m["m2_col"]]["data"]
        results.append({
            "proj_number": m["proj_number"],
            "name": m["name"],
            "mwdc": safe_float(m1_data.get(ROW_DC_MW)) or 0,
            "m1_npp": safe_float(m1_data.get(ROW_NPP)),
            "m1_irr": safe_float(m1_data.get(ROW_FMV_IRR)),
            "m2_npp": safe_float(m2_data.get(ROW_NPP)),
            "m2_irr": safe_float(m2_data.get(ROW_FMV_IRR)),
        })
    return results


# ---------------------------------------------------------------------------
# Input diff detection
# ---------------------------------------------------------------------------

def diff_inputs(
    matched: list[dict],
    m1_projects: dict,
    m2_projects: dict,
) -> list[dict]:
    """Find ALL Project Inputs rows that differ between the two models.

    Uses two passes:
    1. Canonical rows (INPUT_ROW_LABELS) — compared by row number via row mapping
    2. ALL column B labels (_all_inputs) — compared by label text, catching any
       hardcoded input not in our canonical list

    Returns list of {row, label, unit, category, values: {proj_number: (m1, m2)}}.
    """
    diffs: list[dict] = []
    seen_labels: set[str] = set()

    # Model-sourced units (by canonical row and by label); prefer model over hardcoded.
    _m1_units_by_row: dict[int, str] = {}
    _m1_units_by_label: dict[str, str] = {}
    if matched:
        first_data = m1_projects[matched[0]["m1_col"]]["data"]
        _m1_units_by_row = first_data.get("_units_by_row") or {}
        _m1_units_by_label = first_data.get("_all_units") or {}

    # --- Pass 1: Canonical rows (by mapped row number) ---
    for row_num, label in INPUT_ROW_LABELS.items():
        if row_num in _SKIP_ROWS:
            continue
        if label.strip().lower() in _WALK_EXCLUDE_LABELS:
            continue

        is_text = row_num in TEXT_ROWS or row_num in DATE_ROWS
        per_project: dict[int, tuple[Any, Any]] = {}
        any_diff = False

        for m in matched:
            m1_val = m1_projects[m["m1_col"]]["data"].get(row_num)
            m2_val = m2_projects[m["m2_col"]]["data"].get(row_num)

            if is_text:
                s1 = str(m1_val or "").strip()
                s2 = str(m2_val or "").strip()
                if s1 != s2:
                    any_diff = True
            else:
                f1 = safe_float(m1_val)
                f2 = safe_float(m2_val)
                if f1 is None and f2 is None:
                    continue
                if f1 is None or f2 is None or abs(f1 - f2) > 1e-6:
                    any_diff = True

            per_project[m["proj_number"]] = (m1_val, m2_val)

        if any_diff and per_project:
            unit = _m1_units_by_row.get(row_num) or INPUT_ROW_UNITS.get(row_num, "")
            diffs.append({
                "row": row_num,
                "label": label,
                "unit": unit,
                "category": _categorize_row(row_num),
                "values": per_project,
            })
        seen_labels.add(label.strip().lower())

    # --- Pass 2: ALL inputs by column B label (catches non-canonical rows) ---
    if matched:
        # Collect all unique labels across both models
        first_m = matched[0]
        m1_all = m1_projects[first_m["m1_col"]]["data"].get("_all_inputs", {})
        m2_all = m2_projects[first_m["m2_col"]]["data"].get("_all_inputs", {})
        all_labels = set(m1_all.keys()) | set(m2_all.keys())

        for label in sorted(all_labels):
            if not label or not label.strip():
                continue  # skip unlabeled rows
            if label.strip().lower() in seen_labels:
                continue  # already compared in Pass 1
            if label.strip().lower() in _WALK_EXCLUDE_LABELS:
                continue  # excluded from walk
            # Skip individual year property tax rows (only Y1 matters)
            label_lower = label.strip().lower()
            if "property tax" in label_lower and any(
                x in label_lower for x in ["year 2", "year 3", "year 4", "year 5",
                    "yr 2", "yr 3", "yr 4", "yr 5", "y2", "y3", "y4", "y5"]
            ):
                continue

            per_project = {}
            any_diff = False
            for m in matched:
                m1_inputs = m1_projects[m["m1_col"]]["data"].get("_all_inputs", {})
                m2_inputs = m2_projects[m["m2_col"]]["data"].get("_all_inputs", {})
                m1_val = m1_inputs.get(label)
                m2_val = m2_inputs.get(label)

                if m1_val is None and m2_val is None:
                    continue

                f1 = safe_float(m1_val)
                f2 = safe_float(m2_val)
                if f1 is not None and f2 is not None:
                    if abs(f1 - f2) > 1e-6:
                        any_diff = True
                elif str(m1_val or "").strip() != str(m2_val or "").strip():
                    any_diff = True

                per_project[m["proj_number"]] = (m1_val, m2_val)

            if any_diff and per_project:
                unit = _m1_units_by_label.get(label, "")
                diffs.append({
                    "row": 0,  # unknown canonical row
                    "label": label,
                    "unit": unit,
                    "category": "Other",
                    "values": per_project,
                })

    return diffs


# ---------------------------------------------------------------------------
# Number format selection
# ---------------------------------------------------------------------------

def _num_format(row: int) -> str:
    if row in PCT_ROWS:
        return FMT_PCT
    if row in DPW_ROWS:
        return FMT_DPW
    if row in INT_ROWS:
        return FMT_INT
    return FMT_DPW  # default to 3-decimal


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def build_walk_xlsx(
    m1_result: dict,
    m2_result: dict,
    m1_label: str,
    m2_label: str,
    include_proj_numbers: set[int] | None = None,
) -> tuple[io.BytesIO, dict]:
    """Build a formatted Walk Summary .xlsx comparing two models.

    Args:
        include_proj_numbers: If provided, only include projects whose
            proj_number is in this set (i.e. the confirmed portfolio selection).

    Returns (BytesIO with xlsx data, summary dict for UI).
    """
    m1_projects = get_projects(m1_result) or {}
    m2_projects = get_projects(m2_result) or {}

    matched = match_projects(m1_projects, m2_projects)

    # Filter to only portfolio-selected projects when a filter is provided
    if include_proj_numbers is not None:
        matched = [m for m in matched if m["proj_number"] in include_proj_numbers]

    if not matched:
        logger.warning("No projects matched between the two models by Project #.")
        # Flag so we can write an explanatory message into the xlsx below
        _no_matches = True
    else:
        _no_matches = False

    metrics = extract_metrics(matched, m1_projects, m2_projects)
    variances = diff_inputs(matched, m1_projects, m2_projects)

    # Group variances by category
    grouped: dict[str, list[dict]] = {}
    for v in variances:
        grouped.setdefault(v["category"], []).append(v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Build Walk"
    ws.sheet_view.showGridLines = False  # Clean look matching reference files

    case_labels = [m1_label, m2_label]
    n_cases = 2

    # Column layout: B=proj name, C=MWdc, D=units
    # Cases start at col E (5). Each case = 3 cols: NPP, IRR, ∆ Base
    # Case 1: E,F,G (5,6,7)  Case 2: H,I,J (8,9,10)
    def case_cols(case_idx: int) -> tuple[int, int, int]:
        """Return (npp_col, irr_col, delta_col) for a case (0-indexed)."""
        base = 5 + case_idx * 3
        return base, base + 1, base + 2

    last_col = 4 + n_cases * 3  # 10 for 2 cases

    # --- Column widths ---
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    for ci in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ===================================================================
    # TOP SECTION: NPP / IRR comparison table
    # ===================================================================

    # Row 3: Case numbers with merged cells and double border
    for ci, label in enumerate(case_labels):
        npp_c, irr_c, delta_c = case_cols(ci)
        npp_l = get_column_letter(npp_c)
        # Merge across the 3-col case group (or 2 for case 1 which has no delta)
        if ci == 0:
            # Base case: only NPP + IRR (no delta from self)
            merge_end = get_column_letter(irr_c)
        else:
            merge_end = get_column_letter(delta_c)
        ws.merge_cells(f"{npp_l}3:{merge_end}3")
        cell = ws.cell(row=3, column=npp_c, value=ci + 1)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = DOUBLE_BOTTOM
        # Apply border to all merged cells
        for mc in range(npp_c, (delta_c if ci > 0 else irr_c) + 1):
            ws.cell(row=3, column=mc).border = DOUBLE_BOTTOM

    # Row 5: Case labels with merged cells
    for ci, label in enumerate(case_labels):
        npp_c, irr_c, delta_c = case_cols(ci)
        npp_l = get_column_letter(npp_c)
        if ci == 0:
            merge_end = get_column_letter(irr_c)
        else:
            merge_end = get_column_letter(delta_c)
        ws.merge_cells(f"{npp_l}5:{merge_end}5")
        cell = ws.cell(row=5, column=npp_c, value=label)
        cell.font = BOLD_FONT
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BOTTOM
        for mc in range(npp_c, (delta_c if ci > 0 else irr_c) + 1):
            ws.cell(row=5, column=mc).border = THIN_BOTTOM

    # Row 6: Column headers with grey fill
    headers_fixed = [
        (2, "Project Name", None),
        (3, "MWdc", None),
    ]
    for col, text, font_override in headers_fixed:
        cell = ws.cell(row=6, column=col, value=text)
        cell.font = BOLD_FONT if col == 2 else NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = THIN_BOTTOM
        cell.alignment = CENTER if col != 2 else LEFT

    # Grey fill on spacer col D
    ws.cell(row=6, column=4).fill = GREY_FILL
    ws.cell(row=6, column=4).border = THIN_BOTTOM

    for ci in range(n_cases):
        npp_c, irr_c, delta_c = case_cols(ci)
        # NPP header — left border separator
        cell = ws.cell(row=6, column=npp_c, value="NPP ($/W)")
        cell.font = NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = HDR_NPP
        cell.alignment = CENTER
        # IRR header
        cell = ws.cell(row=6, column=irr_c, value="IRR (%)")
        cell.font = NORMAL_FONT
        cell.fill = GREY_FILL
        cell.border = THIN_BOTTOM
        cell.alignment = CENTER
        if ci > 0:
            # Delta header — left+right border
            cell = ws.cell(row=6, column=delta_c, value="\u2206 Base")
            cell.font = BLUE_FONT
            cell.fill = GREY_FILL
            cell.border = HDR_DELTA
            cell.alignment = CENTER

    # Rows 7+: Per-project data with column separators matching reference
    data_start = 7
    for pi, pm in enumerate(metrics):
        r = data_start + pi
        is_last = pi == len(metrics) - 1

        # Project name — navy fill, white bold
        cell = ws.cell(row=r, column=2, value=pm["name"])
        cell.font = WHITE_BOLD
        cell.fill = NAVY_FILL
        if is_last:
            cell.border = THIN_BOTTOM

        # MWdc
        cell = ws.cell(row=r, column=3, value=pm["mwdc"])
        cell.number_format = FMT_MW
        cell.alignment = CENTER
        if is_last:
            cell.border = THIN_BOTTOM

        # Per-case NPP, IRR, delta
        case_vals = [
            (pm["m1_npp"], pm["m1_irr"]),
            (pm["m2_npp"], pm["m2_irr"]),
        ]
        base_npp_col_letter = get_column_letter(case_cols(0)[0])

        for ci, (npp_val, irr_val) in enumerate(case_vals):
            npp_c, irr_c, delta_c = case_cols(ci)

            # NPP — left border separator
            cell = ws.cell(row=r, column=npp_c, value=npp_val)
            cell.number_format = FMT_NPP
            cell.alignment = CENTER
            cell.border = THIN_BOTTOM_LEFT if is_last else THIN_LEFT

            # IRR
            cell = ws.cell(row=r, column=irr_c, value=irr_val)
            cell.number_format = FMT_IRR
            cell.alignment = CENTER
            if is_last:
                cell.border = THIN_BOTTOM

            # Delta (only for non-base cases) — left+right border
            if ci > 0:
                npp_letter = get_column_letter(npp_c)
                formula = f"={npp_letter}{r}-{base_npp_col_letter}{r}"
                cell = ws.cell(row=r, column=delta_c, value=formula)
                cell.number_format = FMT_DELTA
                cell.alignment = CENTER
                cell.border = THIN_BOTTOM_LEFT_RIGHT if is_last else THIN_LEFT_RIGHT

    # Summary row: MW-weighted averages
    summary_r = data_start + len(metrics)
    last_data_r = data_start + len(metrics) - 1
    if metrics:
        # Total MWdc
        cell = ws.cell(
            row=summary_r, column=3,
            value=f"=SUM(C{data_start}:C{last_data_r})",
        )
        cell.number_format = FMT_MW
        cell.alignment = CENTER

        for ci in range(n_cases):
            npp_c, irr_c, delta_c = case_cols(ci)
            npp_l = get_column_letter(npp_c)
            irr_l = get_column_letter(irr_c)

            # MW-weighted NPP
            cell = ws.cell(
                row=summary_r, column=npp_c,
                value=(
                    f"=SUMPRODUCT({npp_l}{data_start}:{npp_l}{last_data_r},"
                    f"$C${data_start}:$C${last_data_r})"
                    f"/SUM($C${data_start}:$C${last_data_r})"
                ),
            )
            cell.number_format = FMT_NPP
            cell.alignment = CENTER

            # MW-weighted IRR
            cell = ws.cell(
                row=summary_r, column=irr_c,
                value=(
                    f"=SUMPRODUCT({irr_l}{data_start}:{irr_l}{last_data_r},"
                    f"$C${data_start}:$C${last_data_r})"
                    f"/SUM($C${data_start}:$C${last_data_r})"
                ),
            )
            cell.number_format = FMT_IRR
            cell.alignment = CENTER

    # ===================================================================
    # BOTTOM SECTION: Variance drivers
    # ===================================================================

    var_start = summary_r + 3

    # "Project Inputs" header
    cell = ws.cell(row=var_start, column=2, value="Project Inputs")
    cell.font = NORMAL_FONT
    cell.border = DOUBLE_BOTTOM
    for vc in range(3, last_col + 1):
        ws.cell(row=var_start, column=vc).border = DOUBLE_BOTTOM

    # Column headers for variance section
    cur_row = var_start + 1
    ws.cell(row=cur_row, column=3, value="Unit").font = Font(size=10, color="7d8694")
    ws.cell(row=cur_row, column=3).alignment = CENTER
    cur_row += 1

    # Compute MW weights for portfolio averaging
    total_mw = sum(pm["mwdc"] for pm in metrics) or 1.0

    for cat_name in _CATEGORY_ORDER:
        cat_vars = grouped.get(cat_name)
        if not cat_vars:
            continue

        # Category header
        cell = ws.cell(row=cur_row, column=2, value=cat_name)
        cell.font = BOLD_FONT
        cell.border = DOUBLE_BOTTOM
        cur_row += 1

        for v in sorted(cat_vars, key=lambda x: x["label"].lower()):
            nfmt = _num_format(v["row"])

            # Label in col B
            cell = ws.cell(row=cur_row, column=2, value=v["label"])
            cell.font = NORMAL_FONT
            cell.alignment = LEFT

            # Unit in col C (blue font)
            if v["unit"]:
                cell = ws.cell(row=cur_row, column=3, value=v["unit"])
                cell.font = BLUE_FONT
                cell.alignment = CENTER

            # MW-weighted portfolio average across matched projects
            m1_sum = 0.0
            m2_sum = 0.0
            m1_count = 0
            m2_count = 0
            is_text_val = False
            first_m1 = None
            first_m2 = None
            for pnum, (m1v, m2v) in v["values"].items():
                mw = next((pm["mwdc"] for pm in metrics if pm["proj_number"] == pnum), 1.0)
                f1 = safe_float(m1v)
                f2 = safe_float(m2v)
                if f1 is not None:
                    m1_sum += f1 * mw
                    m1_count += 1
                elif m1v is not None:
                    is_text_val = True
                    if first_m1 is None:
                        first_m1 = m1v
                if f2 is not None:
                    m2_sum += f2 * mw
                    m2_count += 1
                elif m2v is not None:
                    is_text_val = True
                    if first_m2 is None:
                        first_m2 = m2v

            if is_text_val:
                # Text values: show first project's value
                m1_display = first_m1
                m2_display = first_m2
            else:
                # Numeric: MW-weighted average
                m1_display = m1_sum / total_mw if m1_count else None
                m2_display = m2_sum / total_mw if m2_count else None

            # Case 1 value in col E — boxed cell
            c_e = ws.cell(row=cur_row, column=5, value=m1_display)
            if not is_text_val and m1_display is not None:
                c_e.number_format = nfmt
            c_e.alignment = CENTER_CONT
            c_e.border = THIN_BOX

            # Delta in col G (=H{r}-E{r}) — boxed cell
            if not is_text_val:
                delta_cell = ws.cell(row=cur_row, column=7, value=f"=H{cur_row}-E{cur_row}")
                delta_cell.number_format = FMT_DELTA
                delta_cell.alignment = CENTER
                delta_cell.border = THIN_BOX

            # Case 2 value in col H — boxed cell
            c_h = ws.cell(row=cur_row, column=8, value=m2_display)
            if not is_text_val and m2_display is not None:
                c_h.number_format = nfmt
            c_h.alignment = CENTER_CONT
            c_h.border = THIN_BOX

            # Yellow highlight on changed values
            if m1_display != m2_display:
                f1 = safe_float(m1_display)
                f2 = safe_float(m2_display)
                if f1 is not None and f2 is not None and abs(f1 - f2) > 1e-6:
                    c_h.fill = YELLOW_FILL
                elif is_text_val and str(m1_display or "") != str(m2_display or ""):
                    c_h.fill = YELLOW_FILL

            cur_row += 1

    # If no projects matched, write an explanatory message
    if _no_matches:
        cell = ws.cell(
            row=3, column=2,
            value="No projects matched between models. Check that Project # "
                  "(row 2) is populated in both models.",
        )
        cell.font = Font(color="FF0000", bold=True, size=11)

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Build summary dict for the UI panel
    cats_found = [c for c in _CATEGORY_ORDER if c in grouped]
    summary = {
        "n_matched": len(matched),
        "n_diffs": len(variances),
        "categories": cats_found,
        "m1_label": m1_label,
        "m2_label": m2_label,
    }

    return buf, summary
