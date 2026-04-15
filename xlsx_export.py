"""
38DN Pricing Model Review — Excel Export
Generates branded XLSX with Variance and Full comparison sheets,
per-project Bible comparison, rate curve analysis, and Data Room validation.
Style mirrors 38DN walk/pricing summary formatting (Aptos Narrow, navy headers, etc.)
"""

import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants (from 38DN walk files)
# ---------------------------------------------------------------------------

# Fonts
FONT_LABEL = Font(name="Aptos Narrow", size=11, color="050D25")
FONT_LABEL_BOLD = Font(name="Aptos Narrow", size=11, bold=True, color="050D25")
FONT_HEADER = Font(name="Aptos Display", size=11, bold=True, color="FFFFFF")
FONT_HEADER_REG = Font(name="Aptos Display", size=11, color="FFFFFF")
FONT_UNIT = Font(name="Aptos Narrow", size=11, italic=True, color="0000FF")
FONT_SECTION = Font(name="Aptos Narrow", size=11, bold=True, color="050D25")
FONT_DATA = Font(name="Aptos Narrow", size=11, color="050D25")
FONT_DATA_BOLD = Font(name="Aptos Narrow", size=11, bold=True, color="050D25")
FONT_NEG = Font(name="Aptos Narrow", size=11, color="B83230")
FONT_NEG_BOLD = Font(name="Aptos Narrow", size=11, bold=True, color="B83230")

# Fills
FILL_NAVY = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
FILL_NAVY2 = PatternFill(start_color="212B48", end_color="212B48", fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_ICE_BLUE = PatternFill(start_color="DDEFF7", end_color="DDEFF7", fill_type="solid")
FILL_LAVENDER = PatternFill(start_color="C7CFE6", end_color="C7CFE6", fill_type="solid")
FILL_NONE = PatternFill(fill_type=None)

# Borders
THIN = Side(style="thin")
DOUBLE = Side(style="double")
HAIR = Side(style="hair")
BORDER_STD = Border(top=THIN, bottom=THIN)
BORDER_SECTION = Border(bottom=DOUBLE)
BORDER_HEADER = Border(top=THIN, bottom=THIN, right=THIN)
BORDER_NONE = Border()

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
NF_PCT = "0.00%"
NF_PCT1 = "0.0%"
NF_DPW3 = '0.000_);\\(0.000\\)'  # $/W with parens for neg
NF_DPW3_RED = '0.000_);[Red]\\(0.000\\)'  # red neg
NF_DPW2 = "0.00"
NF_INT = "#,##0"
NF_DOLLAR = '"$"#,##0'
NF_YEAR = "0"
NF_GENERAL = "General"

# Rows that are percentages
PCT_ROWS = {30, 31, 36, 37, 158, 161, 162, 168, 219, 220, 221, 227, 229, 231,
            237, 241, 282, 283, 286, 288, 293, 297, 299, 597, 602}
TEXT_ROWS = {4, 8, 10, 18, 19, 21, 22, 117, 155, 156,
             165, 166, 591, 596}
DATE_ROWS = {68, 69, 70, 71, 72, 73}
# Rows in $/W (3 decimal)
DPW_ROWS = {32, 33, 38, 118, 119, 120, 121, 122, 123, 124, 126, 129, 157, 167, 216, 218}
# Rows that are integers / large numbers
INT_ROWS = {15, 16, 17, 24, 25, 143, 160, 170, 217, 225, 226, 228, 230, 235,
            240, 256, 258, 284, 285, 292, 296, 298, 302}
# Output rows (highlight)
OUTPUT_HIGHLIGHT = {38, 39, 33}


def safe_float(v):
    if v is None: return None
    try: return float(v)
    except (ValueError, TypeError): return None


NF_DATE = "MM/DD/YYYY"

def get_nf(row):
    """Get the appropriate number format for a row."""
    if row in PCT_ROWS: return NF_PCT
    if row in DPW_ROWS: return NF_DPW3_RED
    if row in INT_ROWS: return NF_INT
    if row in TEXT_ROWS: return NF_GENERAL
    if row in DATE_ROWS: return NF_DATE
    return NF_DPW2


def write_comparison_sheet(ws, comp_rows, label1, label2, title, variance_only=False):
    """Write a formatted comparison table to a worksheet.

    Args:
        ws: openpyxl worksheet
        comp_rows: list of dicts with Row, Field, label1, label2, Delta, Delta %, _delta_raw, _pct_raw
        label1, label2: column header labels
        title: sheet title text
        variance_only: if True, only include rows where values differ
    """
    # Filter if variance only
    if variance_only:
        comp_rows = [r for r in comp_rows if r.get("_delta_raw") is not None and r["_delta_raw"] != 0
                     or r.get(label1) != r.get(label2)]

    # Column widths
    ws.column_dimensions["A"].width = 3.0   # spacer
    ws.column_dimensions["B"].width = 8.0   # Row #
    ws.column_dimensions["C"].width = 35.0  # Field
    ws.column_dimensions["D"].width = 8.0   # Units
    ws.column_dimensions["E"].width = 16.0  # Model 1
    ws.column_dimensions["F"].width = 16.0  # Model 2
    ws.column_dimensions["G"].width = 16.0  # Delta
    ws.column_dimensions["H"].width = 14.0  # Delta %

    # Row 2: Title
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = title
    c.font = FONT_LABEL_BOLD
    c.alignment = ALIGN_LEFT

    # Row 4: Header banner
    headers = ["Row", "Field", "Units", label1, label2, "Delta (units)", "Delta (%)"]
    cols = ["B", "C", "D", "E", "F", "G", "H"]
    for i, (col, hdr) in enumerate(zip(cols, headers)):
        cell = ws[f"{col}4"]
        cell.value = hdr
        cell.font = FONT_HEADER if i < 3 else FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    # Row 5: Units sub-header
    ws["D5"].font = FONT_UNIT
    ws["D5"].alignment = ALIGN_CENTER

    # Data rows starting at row 6
    current_row = 6
    prev_section = None

    for rd in comp_rows:
        row_num = rd.get("Row", "")
        field = rd.get("Field", "")
        is_output = row_num in OUTPUT_HIGHLIGHT
        is_text = row_num in TEXT_ROWS or row_num in DATE_ROWS
        is_pct = row_num in PCT_ROWS

        # Detect section breaks (rows with no data = section headers)
        # Insert section separator for key sections
        section_breaks = {
            4: "Project Details", 68: "Milestones",
            118: "CapEx", 143: "Revenue",
            147: "Rate Component 1", 148: "Rate Component 2",
            216: "Incentives", 225: "OpEx", 282: "Decommissioning",
            291: "Property Tax & Insurance", 587: "Tax & Depreciation",
            32: "Outputs",
        }
        if row_num in section_breaks and section_breaks[row_num] != prev_section:
            prev_section = section_breaks[row_num]
            ws[f"B{current_row}"].border = BORDER_SECTION
            ws[f"C{current_row}"].value = prev_section
            ws[f"C{current_row}"].font = FONT_SECTION
            ws[f"C{current_row}"].border = BORDER_SECTION
            for col in ["D", "E", "F", "G", "H"]:
                ws[f"{col}{current_row}"].border = BORDER_SECTION
            current_row += 1

        # Row number
        ws[f"B{current_row}"].value = row_num
        ws[f"B{current_row}"].font = FONT_DATA
        ws[f"B{current_row}"].alignment = ALIGN_CENTER
        ws[f"B{current_row}"].border = BORDER_STD

        # Field name
        ws[f"C{current_row}"].value = field
        ws[f"C{current_row}"].font = FONT_DATA_BOLD if is_output else FONT_DATA
        ws[f"C{current_row}"].alignment = ALIGN_LEFT
        ws[f"C{current_row}"].border = BORDER_STD
        if is_output:
            ws[f"C{current_row}"].fill = FILL_ICE_BLUE

        # Units
        nf = get_nf(row_num)
        unit_text = ""
        if is_pct: unit_text = "%"
        elif row_num in DPW_ROWS: unit_text = "$/W"
        elif row_num in INT_ROWS: unit_text = "#"
        elif is_text: unit_text = "text"
        ws[f"D{current_row}"].value = unit_text
        ws[f"D{current_row}"].font = FONT_UNIT
        ws[f"D{current_row}"].alignment = ALIGN_CENTER
        ws[f"D{current_row}"].border = BORDER_STD

        # Model 1 value
        v1_str = rd.get(label1, "")
        v1_raw = rd.get("_v1_raw")
        cell_e = ws[f"E{current_row}"]
        if v1_raw is not None and not is_text:
            cell_e.value = v1_raw
            cell_e.number_format = nf
        else:
            cell_e.value = v1_str if v1_str != "\u2014" else ""
        cell_e.font = FONT_DATA_BOLD if is_output else FONT_DATA
        cell_e.alignment = ALIGN_CENTER
        cell_e.border = BORDER_STD
        if is_output: cell_e.fill = FILL_ICE_BLUE

        # Model 2 value
        v2_str = rd.get(label2, "")
        v2_raw = rd.get("_v2_raw")
        cell_f = ws[f"F{current_row}"]
        if v2_raw is not None and not is_text:
            cell_f.value = v2_raw
            cell_f.number_format = nf
        else:
            cell_f.value = v2_str if v2_str != "\u2014" else ""
        cell_f.font = FONT_DATA_BOLD if is_output else FONT_DATA
        cell_f.alignment = ALIGN_CENTER
        cell_f.border = BORDER_STD
        if is_output: cell_f.fill = FILL_ICE_BLUE

        # Delta
        delta = rd.get("_delta_raw")
        cell_g = ws[f"G{current_row}"]
        if delta is not None and not is_text:
            cell_g.value = delta
            cell_g.number_format = nf
            cell_g.font = FONT_NEG_BOLD if delta < 0 else (FONT_DATA_BOLD if is_output else FONT_DATA)
        else:
            cell_g.value = ""
        cell_g.alignment = ALIGN_CENTER
        cell_g.border = BORDER_STD

        # Delta %
        pct = rd.get("_pct_raw")
        cell_h = ws[f"H{current_row}"]
        if pct is not None and not is_text:
            cell_h.value = pct
            cell_h.number_format = '0.00%;[Red]\\(0.00%\\);"-"'
            cell_h.font = FONT_NEG if pct < 0 else FONT_DATA
        else:
            cell_h.value = ""
        cell_h.alignment = ALIGN_CENTER
        cell_h.border = BORDER_STD

        # Alternating row fill
        if current_row % 2 == 0 and not is_output:
            for col in ["B", "C", "D", "E", "F", "G", "H"]:
                if ws[f"{col}{current_row}"].fill == FILL_NONE or ws[f"{col}{current_row}"].fill.fgColor.rgb == "00000000":
                    ws[f"{col}{current_row}"].fill = FILL_LIGHT_GRAY

        current_row += 1

    # Freeze panes
    ws.freeze_panes = "E5"

    return current_row


def generate_comparison_xlsx(comp_data, label1, label2, title="Model Comparison"):
    """Generate a branded Excel workbook with Variance and Full sheets.

    Args:
        comp_data: list of dicts from build_comparison_table with added _v1_raw, _v2_raw keys
        label1, label2: model labels
        title: workbook title

    Returns: BytesIO buffer
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Variance (only differences)
    ws_var = wb.active
    ws_var.title = "Variance"
    write_comparison_sheet(ws_var, comp_data, label1, label2,
                           f"{title} \u2014 Variance Only", variance_only=True)

    # Sheet 2: Full (all rows)
    ws_full = wb.create_sheet("Full Comparison")
    write_comparison_sheet(ws_full, comp_data, label1, label2,
                           f"{title} \u2014 Full Inputs", variance_only=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_export_rows(proj1_data, proj2_data, label1, label2):
    """Build comparison rows with raw values for Excel export."""
    from app import INPUT_ROW_LABELS, OUTPUT_ROWS

    all_rows = sorted(set(INPUT_ROW_LABELS.keys()) | set(OUTPUT_ROWS.keys()))
    result = []

    for r in all_rows:
        label = INPUT_ROW_LABELS.get(r, OUTPUT_ROWS.get(r, f"Row {r}"))
        v1_raw_val = proj1_data.get(r) if proj1_data else None
        v2_raw_val = proj2_data.get(r) if proj2_data else None

        is_text = r in TEXT_ROWS
        is_pct = r in PCT_ROWS

        v1 = safe_float(v1_raw_val) if not is_text else None
        v2 = safe_float(v2_raw_val) if not is_text else None

        def fv(v):
            if v is None: return "\u2014"
            if is_pct: return f"{v:.2%}" if abs(v) < 1 else f"{v:.2f}%"
            if abs(v) > 1000: return f"{v:,.0f}"
            return f"{v:.4f}"

        delta = (v2 - v1) if v1 is not None and v2 is not None else None
        pct = (delta / abs(v1)) if delta is not None and v1 and v1 != 0 else None

        result.append({
            "Row": r, "Field": label,
            label1: fv(v1) if not is_text else str(v1_raw_val or ""),
            label2: fv(v2) if not is_text else str(v2_raw_val or ""),
            "_v1_raw": v1 if not is_text else None,
            "_v2_raw": v2 if not is_text else None,
            "_delta_raw": delta,
            "_pct_raw": pct,
        })

    return result


def _truncate_sheet_name(name, suffix, max_len=31):
    """Build a sheet name from a project name + suffix, truncated to Excel's 31-char limit.
    Also replaces characters invalid in sheet names."""
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    prefix_budget = max_len - len(suffix) - 1  # 1 for the separator
    if prefix_budget < 1:
        prefix_budget = 1
    short = name[:prefix_budget]
    result = f"{short} {suffix}"
    return result[:max_len]


# Key metric rows for the summary sheet
_SUMMARY_METRICS = [
    (38,  "NPP ($/W)"),
    (33,  "FMV Calculated ($/W)"),
    (118, "PV EPC Cost ($/W)"),
    (122, "IX Cost ($/W)"),
    (129, "Total Capex Excl. Financing"),
    (597, "ITC Rate (%)"),
    (602, "Eligible Costs (%)"),
    (11,  "System Size (MWdc)"),
    (157, "Energy Rate (at COD)"),
]


def generate_multi_project_xlsx(projects_data, label1, label2, title="Multi-Project Comparison"):
    """Generate a branded Excel workbook with a Summary sheet and per-project Variance/Full sheets.

    Args:
        projects_data: list of (project_name, data1, data2) tuples
        label1, label2: model labels
        title: workbook title

    Returns: BytesIO buffer
    """
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.column_dimensions["A"].width = 3.0
    ws_sum.column_dimensions["B"].width = 28.0  # Project name

    # Title row
    ws_sum.merge_cells("B2:H2")
    ws_sum["B2"].value = title
    ws_sum["B2"].font = FONT_LABEL_BOLD
    ws_sum["B2"].alignment = ALIGN_LEFT

    # Build header columns: B=Project, then for each metric: label1, label2
    headers = ["Project"]
    col_idx = 2  # start at C
    metric_col_pairs = []  # list of (col_letter_1, col_letter_2) for each metric
    for _, metric_label in _SUMMARY_METRICS:
        c1 = get_column_letter(col_idx)
        c2 = get_column_letter(col_idx + 1)
        headers.append(f"{metric_label} ({label1})")
        headers.append(f"{metric_label} ({label2})")
        metric_col_pairs.append((c1, c2))
        # Set column widths
        ws_sum.column_dimensions[c1].width = 18.0
        ws_sum.column_dimensions[c2].width = 18.0
        col_idx += 2

    # Write header row at row 4
    header_row = 4
    ws_sum[f"B{header_row}"].value = "Project"
    ws_sum[f"B{header_row}"].font = FONT_HEADER
    ws_sum[f"B{header_row}"].fill = FILL_NAVY
    ws_sum[f"B{header_row}"].alignment = ALIGN_CENTER
    ws_sum[f"B{header_row}"].border = BORDER_HEADER

    for i, (_, metric_label) in enumerate(_SUMMARY_METRICS):
        c1, c2 = metric_col_pairs[i]
        for col_letter, lbl in [(c1, label1), (c2, label2)]:
            cell = ws_sum[f"{col_letter}{header_row}"]
            cell.value = f"{metric_label}\n({lbl})"
            cell.font = FONT_HEADER
            cell.fill = FILL_NAVY
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_HEADER

    # Data rows
    data_row = 5
    for proj_name, d1, d2 in projects_data:
        ws_sum[f"B{data_row}"].value = proj_name
        ws_sum[f"B{data_row}"].font = FONT_DATA_BOLD
        ws_sum[f"B{data_row}"].alignment = ALIGN_LEFT
        ws_sum[f"B{data_row}"].border = BORDER_STD

        for i, (row_num, _) in enumerate(_SUMMARY_METRICS):
            c1, c2 = metric_col_pairs[i]
            nf = get_nf(row_num)
            v1 = safe_float(d1.get(row_num)) if d1 else None
            v2 = safe_float(d2.get(row_num)) if d2 else None

            cell1 = ws_sum[f"{c1}{data_row}"]
            cell2 = ws_sum[f"{c2}{data_row}"]
            for cell, val in [(cell1, v1), (cell2, v2)]:
                if val is not None:
                    cell.value = val
                    cell.number_format = nf
                else:
                    cell.value = ""
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_STD

        # Alternating row fill
        if data_row % 2 == 0:
            last_col = get_column_letter(1 + len(_SUMMARY_METRICS) * 2)
            for ci in range(2, 2 + len(_SUMMARY_METRICS) * 2 + 1):
                cl = get_column_letter(ci)
                c = ws_sum[f"{cl}{data_row}"]
                if c.fill == FILL_NONE or c.fill.fgColor.rgb == "00000000":
                    c.fill = FILL_LIGHT_GRAY

        data_row += 1

    ws_sum.freeze_panes = "C5"

    # ---- Per-project sheets ----
    for proj_name, d1, d2 in projects_data:
        export_rows = build_export_rows(d1, d2, label1, label2)
        proj_title = f"{proj_name} \u2014 {label1} vs {label2}"

        # Variance sheet
        var_name = _truncate_sheet_name(proj_name, "Var")
        ws_var = wb.create_sheet(var_name)
        write_comparison_sheet(ws_var, export_rows, label1, label2,
                               f"{proj_title} \u2014 Variance Only", variance_only=True)

        # Full sheet
        full_name = _truncate_sheet_name(proj_name, "Full")
        ws_full = wb.create_sheet(full_name)
        write_comparison_sheet(ws_full, export_rows, label1, label2,
                               f"{proj_title} \u2014 Full Inputs", variance_only=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =====================================================================
# ENHANCED REVIEW EXPORT — Bible Comparison + Rate Curve Analysis
# =====================================================================

# Additional style constants for review sheets
FILL_FLAG_RED = PatternFill(start_color="FDE0DF", end_color="FDE0DF", fill_type="solid")
FILL_FLAG_AMBER = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FILL_MATCH_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_DR_MATCH = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FONT_FLAG = Font(name="Aptos Narrow", size=11, bold=True, color="B83230")
FONT_OK = Font(name="Aptos Narrow", size=11, color="45A750")
FONT_NOTE = Font(name="Aptos Narrow", size=10, italic=True, color="666666")
FONT_TITLE = Font(name="Aptos Display", size=14, bold=True, color="050D25")
FONT_SUBTITLE = Font(name="Aptos Display", size=11, bold=True, color="212B48")


def _bible_value_for_row(row_num, bible_benchmarks, state=None):
    """Look up the Pricing Bible benchmark value for a given model row number.

    Returns dict with min/max/unit/label/category or None.
    """
    for cat, checks in bible_benchmarks.items():
        for label, spec in checks.items():
            if spec.get("derived"):
                continue
            if spec.get("row") == row_num:
                return {
                    "min": spec["min"], "max": spec["max"],
                    "unit": spec["unit"], "label": label, "category": cat,
                }
    return None


def _write_project_review_sheet(ws, proj_name, proj_data, bible_benchmarks,
                                 state, data_room=None):
    """Write a per-project review sheet: Model Value vs Bible Range with flags."""
    from config import (INPUT_ROW_LABELS, OUTPUT_ROWS, DISPLAY_ORDER,
                        DPW_ROWS, SECTION_BREAKS)

    ws.column_dimensions["A"].width = 3.0
    ws.column_dimensions["B"].width = 8.0
    ws.column_dimensions["C"].width = 34.0
    ws.column_dimensions["D"].width = 8.0
    ws.column_dimensions["E"].width = 16.0
    ws.column_dimensions["F"].width = 14.0
    ws.column_dimensions["G"].width = 14.0
    ws.column_dimensions["H"].width = 12.0
    ws.column_dimensions["I"].width = 30.0

    # Title
    ws.merge_cells("B2:I2")
    ws["B2"].value = f"{proj_name} \u2014 Pricing Bible Review"
    ws["B2"].font = FONT_TITLE
    ws["B2"].alignment = ALIGN_LEFT

    ws["B3"].value = f"State: {state}"
    ws["B3"].font = FONT_SUBTITLE

    # Header row
    headers = ["Row", "Field", "Units", "Model Value", "Bible Min", "Bible Max", "Status", "Notes"]
    cols = ["B", "C", "D", "E", "F", "G", "H", "I"]
    for col, hdr in zip(cols, headers):
        cell = ws[f"{col}5"]
        cell.value = hdr
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    current_row = 7
    prev_section = None
    flag_count = 0

    all_rows = [r for r in DISPLAY_ORDER if r in INPUT_ROW_LABELS or r in OUTPUT_ROWS]

    for r in all_rows:
        label = INPUT_ROW_LABELS.get(r, OUTPUT_ROWS.get(r, f"Row {r}"))
        is_text = r in TEXT_ROWS or r in DATE_ROWS
        is_pct = r in PCT_ROWS
        is_output = r in OUTPUT_HIGHLIGHT

        # Section break
        if r in SECTION_BREAKS and SECTION_BREAKS[r] != prev_section:
            prev_section = SECTION_BREAKS[r]
            ws[f"B{current_row}"].border = BORDER_SECTION
            ws[f"C{current_row}"].value = prev_section
            ws[f"C{current_row}"].font = FONT_SECTION
            ws[f"C{current_row}"].border = BORDER_SECTION
            for col in ["D", "E", "F", "G", "H", "I"]:
                ws[f"{col}{current_row}"].border = BORDER_SECTION
            current_row += 1

        # Row number
        ws[f"B{current_row}"].value = r
        ws[f"B{current_row}"].font = FONT_DATA
        ws[f"B{current_row}"].alignment = ALIGN_CENTER
        ws[f"B{current_row}"].border = BORDER_STD

        # Field name
        ws[f"C{current_row}"].value = label
        ws[f"C{current_row}"].font = FONT_DATA_BOLD if is_output else FONT_DATA
        ws[f"C{current_row}"].alignment = ALIGN_LEFT
        ws[f"C{current_row}"].border = BORDER_STD

        # Units
        nf = get_nf(r)
        unit_text = ""
        if is_pct:
            unit_text = "%"
        elif r in DPW_ROWS:
            unit_text = "$/W"
        elif r in INT_ROWS:
            unit_text = "#"
        elif is_text:
            unit_text = "text"
        ws[f"D{current_row}"].value = unit_text
        ws[f"D{current_row}"].font = FONT_UNIT
        ws[f"D{current_row}"].alignment = ALIGN_CENTER
        ws[f"D{current_row}"].border = BORDER_STD

        # Model value
        raw_val = proj_data.get(r)
        v = safe_float(raw_val) if not is_text else None
        cell_e = ws[f"E{current_row}"]
        if v is not None:
            cell_e.value = v
            cell_e.number_format = nf
        else:
            cell_e.value = str(raw_val or "") if is_text else ""
        cell_e.font = FONT_DATA_BOLD if is_output else FONT_DATA
        cell_e.alignment = ALIGN_CENTER
        cell_e.border = BORDER_STD
        if is_output:
            cell_e.fill = FILL_ICE_BLUE

        # Bible comparison
        bible = _bible_value_for_row(r, bible_benchmarks, state)
        cell_f = ws[f"F{current_row}"]
        cell_g = ws[f"G{current_row}"]
        cell_h = ws[f"H{current_row}"]
        cell_i = ws[f"I{current_row}"]

        if bible and v is not None:
            cell_f.value = bible["min"]
            cell_f.number_format = nf
            cell_f.font = FONT_DATA
            cell_f.alignment = ALIGN_CENTER
            cell_f.border = BORDER_STD

            cell_g.value = bible["max"]
            cell_g.number_format = nf
            cell_g.font = FONT_DATA
            cell_g.alignment = ALIGN_CENTER
            cell_g.border = BORDER_STD

            if v < bible["min"]:
                cell_h.value = "LOW"
                cell_h.font = FONT_FLAG
                cell_h.fill = FILL_FLAG_RED
                pct_off = (bible["min"] - v) / bible["min"] * 100 if bible["min"] != 0 else 0
                cell_i.value = f"{pct_off:.1f}% below Bible min"
                cell_i.font = FONT_NOTE
                flag_count += 1
            elif v > bible["max"]:
                cell_h.value = "HIGH"
                cell_h.font = FONT_FLAG
                cell_h.fill = FILL_FLAG_RED
                pct_off = (v - bible["max"]) / bible["max"] * 100 if bible["max"] != 0 else 0
                cell_i.value = f"{pct_off:.1f}% above Bible max"
                cell_i.font = FONT_NOTE
                flag_count += 1
            else:
                cell_h.value = "OK"
                cell_h.font = FONT_OK
                cell_h.fill = FILL_MATCH_GREEN
                cell_i.value = ""

            cell_h.alignment = ALIGN_CENTER
            cell_h.border = BORDER_STD
            cell_i.alignment = ALIGN_LEFT
            cell_i.border = BORDER_STD
        else:
            for c in [cell_f, cell_g, cell_h, cell_i]:
                c.value = ""
                c.border = BORDER_STD

        # Alternating fill
        if current_row % 2 == 0 and not is_output:
            for col in ["B", "C", "D", "E", "F", "G"]:
                c = ws[f"{col}{current_row}"]
                if c.fill == FILL_NONE or (c.fill.fgColor and c.fill.fgColor.rgb == "00000000"):
                    c.fill = FILL_LIGHT_GRAY

        current_row += 1

    # Summary row at top
    ws["F3"].value = f"{flag_count} flagged"
    ws["F3"].font = FONT_FLAG if flag_count else FONT_OK

    ws.freeze_panes = "E6"
    return current_row


def _write_rate_curve_sheet(ws, proj_name, proj_rc_data, rc_dates,
                             gh25_ref, state, proj_data):
    """Write a rate curve analysis sheet for one project.

    Compares the model's per-project custom rate curves against the
    GH25 reference curves, calculates implied discount, and flags deviations.
    """
    ws.column_dimensions["A"].width = 3.0
    ws.column_dimensions["B"].width = 12.0
    ws.column_dimensions["C"].width = 16.0
    ws.column_dimensions["D"].width = 16.0
    ws.column_dimensions["E"].width = 16.0
    ws.column_dimensions["F"].width = 16.0
    ws.column_dimensions["G"].width = 14.0
    ws.column_dimensions["H"].width = 12.0
    ws.column_dimensions["I"].width = 24.0

    ws.merge_cells("B2:I2")
    ws["B2"].value = f"{proj_name} \u2014 Rate Curve Analysis"
    ws["B2"].font = FONT_TITLE
    ws["B2"].alignment = ALIGN_LEFT

    # Determine which GH25 curve to use based on utility
    utility = str(proj_data.get(19, "")).strip()
    if "ameren" in utility.lower():
        gh_curve_name = "Ameren GH25"
    elif "comed" in utility.lower() or "com ed" in utility.lower():
        gh_curve_name = "ComEd GH25"
    elif state == "IL":
        gh_curve_name = "Ameren GH25"
    else:
        gh_curve_name = None  # Non-IL: no GH25 reference available

    # Expected GH25 discount based on market config (not customer discount row 161)
    # IL = 17.5%, MD = 22.5% per Pricing Bible
    state_upper = str(state).strip().upper()
    if state_upper == "IL":
        expected_disc = 0.175
    elif state_upper in ("MD", "MD/DE"):
        expected_disc = 0.225
    else:
        # Fall back to the rate discount from Project Inputs (row 161)
        expected_disc = safe_float(proj_data.get(161)) or 0

    cust_disc = safe_float(proj_data.get(161)) or 0
    if gh_curve_name:
        ws["B3"].value = (f"Utility: {utility} | GH25 Ref: {gh_curve_name} | "
                          f"GH25 Discount: {expected_disc:.1%} | Cust Discount: {cust_disc:.1%}")
    else:
        ws["B3"].value = (f"Utility: {utility} | GH25 Discount: {expected_disc:.1%} | "
                          f"Cust Discount: {cust_disc:.1%} | No GH25 IL ref for {state}")
    ws["B3"].font = FONT_SUBTITLE

    gh_annual = gh25_ref.get("annual", {}).get(gh_curve_name, {}) if (gh25_ref and gh_curve_name) else {}
    gh_cagr = gh25_ref.get("cagrs", {}).get(gh_curve_name) if (gh25_ref and gh_curve_name) else None

    if gh_cagr:
        ws["B4"].value = f"GH25 CAGR: {gh_cagr:.2%}"
        ws["B4"].font = FONT_NOTE

    # Annualize the model's RC1 rate curve
    model_rc1 = proj_rc_data.get(1, {})
    model_annual = {}
    for dt, val in model_rc1.items():
        if hasattr(dt, "year"):
            model_annual.setdefault(dt.year, []).append(val)
    model_annual = {yr: sum(v) / len(v) for yr, v in model_annual.items()}

    # Header
    headers = ["Year", "Model Rate", "GH25 Rate", "Implied Disc.", "Expected Disc.",
               "Disc. Variance", "Status", "Notes"]
    cols = ["B", "C", "D", "E", "F", "G", "H", "I"]
    for col, hdr in zip(cols, headers):
        cell = ws[f"{col}6"]
        cell.value = hdr
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    current_row = 7
    all_years = sorted(set(list(model_annual.keys()) + list(gh_annual.keys())))
    cod_year = safe_float(proj_data.get(15))
    if cod_year:
        all_years = [y for y in all_years if y >= int(cod_year)]
    all_years = all_years[:35]

    total_model_disc = []

    for yr in all_years:
        m_rate = model_annual.get(yr)
        g_rate = gh_annual.get(yr)

        ws[f"B{current_row}"].value = yr
        ws[f"B{current_row}"].font = FONT_DATA
        ws[f"B{current_row}"].alignment = ALIGN_CENTER
        ws[f"B{current_row}"].border = BORDER_STD

        cell_c = ws[f"C{current_row}"]
        if m_rate is not None:
            cell_c.value = m_rate
            cell_c.number_format = "$#,##0.0000"
        else:
            cell_c.value = ""
        cell_c.font = FONT_DATA
        cell_c.alignment = ALIGN_CENTER
        cell_c.border = BORDER_STD

        cell_d = ws[f"D{current_row}"]
        if g_rate is not None:
            cell_d.value = g_rate
            cell_d.number_format = "$#,##0.0000"
        else:
            cell_d.value = ""
        cell_d.font = FONT_DATA
        cell_d.alignment = ALIGN_CENTER
        cell_d.border = BORDER_STD

        cell_e = ws[f"E{current_row}"]
        cell_f = ws[f"F{current_row}"]
        cell_g = ws[f"G{current_row}"]
        cell_h = ws[f"H{current_row}"]
        cell_i = ws[f"I{current_row}"]

        if m_rate is not None and g_rate is not None and g_rate > 0:
            implied_disc = 1.0 - (m_rate / g_rate)
            cell_e.value = implied_disc
            cell_e.number_format = "0.0%"

            cell_f.value = expected_disc
            cell_f.number_format = "0.0%"

            disc_var = implied_disc - expected_disc
            cell_g.value = disc_var
            cell_g.number_format = "+0.0%;-0.0%"

            total_model_disc.append(implied_disc)

            if abs(disc_var) > 0.03:
                cell_h.value = "FLAG"
                cell_h.font = FONT_FLAG
                cell_h.fill = FILL_FLAG_RED
                cell_i.value = f"Discount off by {abs(disc_var):.1%} pts"
                cell_i.font = FONT_NOTE
            elif abs(disc_var) > 0.01:
                cell_h.value = "WARN"
                cell_h.font = Font(name="Aptos Narrow", size=11, bold=True, color="CC8800")
                cell_h.fill = FILL_FLAG_AMBER
                cell_i.value = ""
            else:
                cell_h.value = "OK"
                cell_h.font = FONT_OK
                cell_h.fill = FILL_MATCH_GREEN
                cell_i.value = ""
        else:
            for c in [cell_e, cell_f, cell_g, cell_h, cell_i]:
                c.value = ""

        for c in [cell_e, cell_f, cell_g, cell_h, cell_i]:
            c.alignment = ALIGN_CENTER
            c.border = BORDER_STD

        if current_row % 2 == 0:
            for col in ["B", "C", "D"]:
                c = ws[f"{col}{current_row}"]
                if c.fill == FILL_NONE or (c.fill.fgColor and c.fill.fgColor.rgb == "00000000"):
                    c.fill = FILL_LIGHT_GRAY

        current_row += 1

    # Summary row
    current_row += 1
    if total_model_disc:
        avg_disc = sum(total_model_disc) / len(total_model_disc)
        ws[f"B{current_row}"].value = "Overall Average"
        ws[f"B{current_row}"].font = FONT_DATA_BOLD
        ws[f"B{current_row}"].border = BORDER_SECTION
        ws[f"C{current_row}"].border = BORDER_SECTION
        ws[f"D{current_row}"].border = BORDER_SECTION

        ws[f"E{current_row}"].value = avg_disc
        ws[f"E{current_row}"].number_format = "0.0%"
        ws[f"E{current_row}"].font = FONT_DATA_BOLD
        ws[f"E{current_row}"].alignment = ALIGN_CENTER
        ws[f"E{current_row}"].border = BORDER_SECTION

        ws[f"F{current_row}"].value = expected_disc
        ws[f"F{current_row}"].number_format = "0.0%"
        ws[f"F{current_row}"].font = FONT_DATA_BOLD
        ws[f"F{current_row}"].alignment = ALIGN_CENTER
        ws[f"F{current_row}"].border = BORDER_SECTION

        overall_var = avg_disc - expected_disc
        ws[f"G{current_row}"].value = overall_var
        ws[f"G{current_row}"].number_format = "+0.0%;-0.0%"
        ws[f"G{current_row}"].font = FONT_FLAG if abs(overall_var) > 0.02 else FONT_DATA_BOLD
        ws[f"G{current_row}"].alignment = ALIGN_CENTER
        ws[f"G{current_row}"].border = BORDER_SECTION

        ws[f"H{current_row}"].value = "PASS" if abs(overall_var) <= 0.02 else "REVIEW"
        ws[f"H{current_row}"].font = FONT_OK if abs(overall_var) <= 0.02 else FONT_FLAG
        ws[f"H{current_row}"].fill = FILL_MATCH_GREEN if abs(overall_var) <= 0.02 else FILL_FLAG_RED
        ws[f"H{current_row}"].alignment = ALIGN_CENTER
        ws[f"H{current_row}"].border = BORDER_SECTION

    ws.freeze_panes = "C7"
    return current_row


def _write_review_summary_sheet(ws, projects_data, bible_benchmarks, title):
    """Write a summary sheet with key metrics and flag counts per project."""
    ws.column_dimensions["A"].width = 3.0
    ws.column_dimensions["B"].width = 30.0

    ws.merge_cells("B2:H2")
    ws["B2"].value = title
    ws["B2"].font = FONT_TITLE
    ws["B2"].alignment = ALIGN_LEFT

    summary_metrics = [
        (11, "Size MWdc", "MWdc"),
        (118, "EPC Cost", "$/W"),
        (129, "Total CapEx", "$/W"),
        (157, "Energy Rate", "$/kWh"),
        (158, "Escalator", "%"),
        (161, "Discount", "%"),
        (597, "ITC Rate", "%"),
        (33, "FMV ($/W)", "$/W"),
        (38, "NPP ($/W)", "$/W"),
    ]

    header_row = 4
    ws[f"B{header_row}"].value = "Project"
    ws[f"B{header_row}"].font = FONT_HEADER
    ws[f"B{header_row}"].fill = FILL_NAVY
    ws[f"B{header_row}"].alignment = ALIGN_CENTER
    ws[f"B{header_row}"].border = BORDER_HEADER

    col_idx = 3
    for _, metric_label, unit in summary_metrics:
        cl = get_column_letter(col_idx)
        ws.column_dimensions[cl].width = 14.0
        cell = ws[f"{cl}{header_row}"]
        cell.value = f"{metric_label}\n({unit})"
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER
        col_idx += 1

    flag_col = get_column_letter(col_idx)
    ws.column_dimensions[flag_col].width = 12.0
    cell = ws[f"{flag_col}{header_row}"]
    cell.value = "Flags"
    cell.font = FONT_HEADER
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_HEADER

    data_row = 5
    for proj_name, proj_data, state in projects_data:
        ws[f"B{data_row}"].value = proj_name
        ws[f"B{data_row}"].font = FONT_DATA_BOLD
        ws[f"B{data_row}"].alignment = ALIGN_LEFT
        ws[f"B{data_row}"].border = BORDER_STD

        ci = 3
        flag_count = 0
        for row_num, _, _ in summary_metrics:
            cl = get_column_letter(ci)
            v = safe_float(proj_data.get(row_num))
            nf = get_nf(row_num)
            cell = ws[f"{cl}{data_row}"]
            if v is not None:
                cell.value = v
                cell.number_format = nf

                bible = _bible_value_for_row(row_num, bible_benchmarks, state)
                if bible:
                    if v < bible["min"] or v > bible["max"]:
                        cell.fill = FILL_FLAG_RED
                        cell.font = FONT_FLAG
                        flag_count += 1
                    else:
                        cell.font = FONT_DATA
                else:
                    cell.font = FONT_DATA
            else:
                cell.value = ""
                cell.font = FONT_DATA
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_STD
            ci += 1

        cell_flag = ws[f"{flag_col}{data_row}"]
        cell_flag.value = flag_count
        cell_flag.font = FONT_FLAG if flag_count else FONT_OK
        cell_flag.fill = FILL_FLAG_RED if flag_count > 2 else (FILL_FLAG_AMBER if flag_count else FILL_MATCH_GREEN)
        cell_flag.alignment = ALIGN_CENTER
        cell_flag.border = BORDER_STD

        if data_row % 2 == 0:
            for ci2 in range(2, col_idx):
                cl2 = get_column_letter(ci2)
                c = ws[f"{cl2}{data_row}"]
                if c.fill == FILL_NONE or (c.fill.fgColor and c.fill.fgColor.rgb == "00000000"):
                    c.fill = FILL_LIGHT_GRAY

        data_row += 1

    ws.freeze_panes = "C5"


def generate_review_xlsx(projects, bible_benchmarks, rate_curves=None,
                          gh25_ref=None, data_room=None, model_label="Model"):
    """Generate the enhanced review XLSX workbook.

    When only one model is loaded (no Model 2), compares each project
    against Pricing Bible benchmarks with flagged variations.
    Includes rate curve analysis sheets when GH25 reference is available.

    Returns: BytesIO buffer
    """
    wb = openpyxl.Workbook()

    active_projects = {k: v for k, v in projects.items() if v["toggle"]}

    proj_list = []
    for col_idx, proj in active_projects.items():
        pname = proj["name"]
        pdata = proj["data"]
        state = str(pdata.get(18, "")).strip().upper()
        proj_list.append((pname, pdata, state))

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_review_summary_sheet(ws_sum, proj_list, bible_benchmarks,
                                 f"{model_label} \u2014 Pricing Bible Review Summary")

    # Per-project Bible Review sheets
    for pname, pdata, state in proj_list:
        sheet_name = _truncate_sheet_name(pname, "Review")
        ws_proj = wb.create_sheet(sheet_name)
        _write_project_review_sheet(ws_proj, pname, pdata, bible_benchmarks,
                                     state, data_room)

    # Rate Curve Analysis sheets
    rc_proj_data = rate_curves.get("projects", {}) if rate_curves else {}

    for pname, pdata, state in proj_list:
        proj_rc = rc_proj_data.get(pname, {})
        if not proj_rc:
            for rc_name, rc_data in rc_proj_data.items():
                first_part = pname.split(" | ")[0].strip()
                if rc_name in pname or first_part in rc_name or pname in rc_name:
                    proj_rc = rc_data
                    break

        if proj_rc:
            sheet_name = _truncate_sheet_name(pname, "Rates")
            ws_rc = wb.create_sheet(sheet_name)
            rc_dates = rate_curves.get("dates", {}) if rate_curves else {}
            _write_rate_curve_sheet(ws_rc, pname, proj_rc, rc_dates,
                                    gh25_ref, state, pdata)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
