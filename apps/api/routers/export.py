"""Review export endpoint — generate an Excel summary of the review session."""
from __future__ import annotations

import io
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

NAVY_FILL = PatternFill("solid", fgColor="002060")
TEAL_FILL = PatternFill("solid", fgColor="518484")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
THIN_BOTTOM = Border(bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center")


class FindingExport(BaseModel):
    field: str
    status: str
    bible: str
    model: str
    impact: float | None = None
    action: str | None = None  # accept / flag / skip / null
    note: str | None = None


class ProjectExport(BaseModel):
    name: str
    verdict: str
    nppPerW: float
    irrPct: float
    equityK: float
    approved: bool = False
    approvalTimestamp: str | None = None
    approvalReviewer: str | None = None
    projectNote: str | None = None
    findings: list[FindingExport]


class ExportRequest(BaseModel):
    model_label: str = "Model"
    reviewer: str = "Caroline Z."
    bible_label: str = "Q1 '26"
    projects: list[ProjectExport]


@router.post("")
def export_review(req: ExportRequest):
    """Generate an Excel review summary with all findings and approval state."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Summary"

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 30

    # Title
    r = 1
    cell = ws.cell(row=r, column=2, value=f"38°N Pricing Model Review — {req.model_label}")
    cell.font = Font(bold=True, size=14, color="002060")
    r += 1
    ws.cell(row=r, column=2, value=f"Bible: {req.bible_label}  |  Reviewer: {req.reviewer}  |  {datetime.now().strftime('%b %d, %Y %H:%M')}")
    ws.cell(row=r, column=2).font = Font(size=10, color="7d8694")
    r += 2

    # Summary stats
    n_approved = sum(1 for p in req.projects if p.approved)
    n_total = len(req.projects)
    total_fail = sum(len([f for f in p.findings if f.status == "OFF"]) for p in req.projects)
    total_flag = sum(len([f for f in p.findings if f.status == "OUT"]) for p in req.projects)
    ws.cell(row=r, column=2, value=f"{n_approved} of {n_total} projects approved  |  {total_fail} FAIL  |  {total_flag} FLAG")
    ws.cell(row=r, column=2).font = BOLD_FONT
    r += 2

    for proj in req.projects:
        # Project header
        cell = ws.cell(row=r, column=2, value=proj.name)
        cell.font = WHITE_FONT
        cell.fill = NAVY_FILL
        status_cell = ws.cell(row=r, column=3, value="APPROVED" if proj.approved else proj.verdict)
        status_cell.font = WHITE_FONT
        status_cell.fill = TEAL_FILL if proj.approved else NAVY_FILL
        status_cell.alignment = CENTER

        # Metrics
        ws.cell(row=r, column=4, value=proj.nppPerW).number_format = '0.000_);[Red]\\(0.000\\)'
        ws.cell(row=r, column=4).font = Font(color="FFFFFF", size=11)
        ws.cell(row=r, column=4).fill = NAVY_FILL
        ws.cell(row=r, column=5, value=proj.irrPct / 100 if abs(proj.irrPct) < 1 else proj.irrPct).number_format = '0.00%'
        ws.cell(row=r, column=5).font = Font(color="FFFFFF", size=11)
        ws.cell(row=r, column=5).fill = NAVY_FILL

        for c in range(6, 9):
            ws.cell(row=r, column=c).fill = NAVY_FILL
        r += 1

        if proj.approvalTimestamp:
            ws.cell(row=r, column=2, value=f"Approved by {proj.approvalReviewer or '—'} at {proj.approvalTimestamp}")
            ws.cell(row=r, column=2).font = Font(size=9, color="518484", italic=True)
            r += 1

        if proj.projectNote:
            ws.cell(row=r, column=2, value=f"Note: {proj.projectNote}")
            ws.cell(row=r, column=2).font = Font(size=9, color="7d8694", italic=True)
            r += 1

        # Findings header
        headers = ["Field", "Status", "Bible", "Model", "Impact", "Action", "Note"]
        for ci, h in enumerate(headers):
            cell = ws.cell(row=r, column=2 + ci, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = GREY_FILL
            cell.border = THIN_BOTTOM
            cell.alignment = CENTER if ci > 0 else Alignment(horizontal="left")
        r += 1

        for f in proj.findings:
            ws.cell(row=r, column=2, value=f.field).font = NORMAL_FONT
            ws.cell(row=r, column=3, value=f.status).alignment = CENTER
            ws.cell(row=r, column=4, value=f.bible).alignment = CENTER
            ws.cell(row=r, column=5, value=f.model).alignment = CENTER
            if f.impact is not None:
                ws.cell(row=r, column=6, value=f.impact).number_format = '#,##0'
                ws.cell(row=r, column=6).alignment = CENTER
            ws.cell(row=r, column=7, value=f.action or "").alignment = CENTER
            ws.cell(row=r, column=8, value=f.note or "")
            r += 1

        r += 1  # blank row between projects

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Review_Summary_{req.model_label}_{datetime.now().strftime('%Y%m%d')}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
