"""
38DN Pricing Model Review — Data Loading
Reads pricing model workbooks and extracts project data.
"""
import streamlit as st
import openpyxl
from pathlib import Path
import re

from config import (
    INPUT_ROW_LABELS, OUTPUT_ROWS, TEXT_ROWS,
    RATE_COMP_STARTS, EQUITY_RATE_TOGGLE_START,
    DEBT_RATE_TOGGLE_START, APPRAISAL_RATE_TOGGLE_START,
)
from utils import safe_float


# ---------------------------------------------------------------------------
# Dynamic label-based row mapping
# ---------------------------------------------------------------------------

def _normalize_label(s):
    """Normalize a label for fuzzy matching: lowercase, strip, collapse whitespace, remove punctuation noise."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r'\s+', ' ', s)
    return s


def _labels_match(canonical, actual):
    """Check if two normalized labels are equivalent."""
    if canonical == actual:
        return True
    # Common substitutions
    c = canonical.replace("&", "and").replace("-", " ").replace("_", " ").replace("esc", "escalator")
    a = actual.replace("&", "and").replace("-", " ").replace("_", " ").replace("esc", "escalator")
    c = re.sub(r'\s+', ' ', c).strip()
    a = re.sub(r'\s+', ' ', a).strip()
    if c == a:
        return True
    # Check containment for longer labels (>8 chars)
    if len(canonical) > 8 and (canonical in actual or actual in canonical):
        return True
    return False


def _detect_label_column(ws, max_row=620):
    """Detect which column (B-E) contains the input row labels."""
    known = {_normalize_label(v) for v in INPUT_ROW_LABELS.values() if v}
    best_col, best_count = 2, 0
    for col in range(2, 6):
        count = 0
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=col).value
            if val is not None and _normalize_label(val) in known:
                count += 1
        if count > best_count:
            best_col, best_count = col, count
    return best_col


def _build_row_mapping(ws, label_col, max_row=620):
    """Build canonical_row -> actual_row mapping by scanning labels in the model.

    Returns a dict where keys are canonical row numbers (from INPUT_ROW_LABELS)
    and values are the actual row numbers found in this specific model.
    Rows not found fall back to their canonical (default) row number.
    """
    # Read all labels from the label column
    actual_labels = {}  # normalized_label -> actual_row
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=label_col).value
        if val is not None:
            norm = _normalize_label(val)
            if norm and norm not in actual_labels:  # first occurrence wins
                actual_labels[norm] = r

    # Also merge OUTPUT_ROWS into the scan
    all_canonical = dict(INPUT_ROW_LABELS)
    all_canonical.update(OUTPUT_ROWS)

    mapping = {}
    for canonical_row, label in all_canonical.items():
        norm = _normalize_label(label)
        if norm in actual_labels:
            mapping[canonical_row] = actual_labels[norm]
        else:
            # Try fuzzy matching
            matched = False
            for actual_norm, actual_row in actual_labels.items():
                if _labels_match(norm, actual_norm):
                    mapping[canonical_row] = actual_row
                    matched = True
                    break
            if not matched:
                mapping[canonical_row] = canonical_row  # fallback to default

    return mapping


# ---------------------------------------------------------------------------
# Model loading
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def load_pricing_model(file):
    """Load a pricing model workbook and extract all project data.

    Uses read_only=False for fast random cell access (read_only streams rows
    sequentially which is catastrophically slow for sparse random reads:
    456s with read_only vs ~3s without on an 8.5MB xlsm).

    Row positions are detected dynamically per model by scanning column B
    labels, so models with slightly different row layouts are handled correctly.
    """
    wb = openpyxl.load_workbook(file, data_only=True, read_only=False)
    ws = wb["Project Inputs"]

    # Detect label column and build per-model row mapping
    label_col = _detect_label_column(ws)
    row_map = _build_row_mapping(ws, label_col)

    all_needed_canonical = set(INPUT_ROW_LABELS.keys()) | set(OUTPUT_ROWS.keys())
    rate_rows = set()
    for start in RATE_COMP_STARTS:
        for offset in range(1, 9):
            rate_rows.add(start + offset)
    for i in range(6):
        rate_rows.add(EQUITY_RATE_TOGGLE_START + i)
        rate_rows.add(DEBT_RATE_TOGGLE_START + i)
        rate_rows.add(APPRAISAL_RATE_TOGGLE_START + i)
    rate_rows.add(400)   # debt match toggle
    rate_rows.add(512)   # appraisal match toggle

    # Resolve actual row numbers for key fixed-position rows
    name_row = row_map.get(4, 4)
    toggle_row = row_map.get(7, 7)

    projects = {}
    for col in range(6, 88):
        name_cell = ws.cell(row=name_row, column=col).value
        if name_cell is None or not str(name_cell).strip():
            continue
        toggle_cell = ws.cell(row=toggle_row, column=col).value
        is_on = str(toggle_cell).strip().lower() in ("1", "on", "true") if toggle_cell is not None else False

        data = {}
        for canonical_r in all_needed_canonical:
            actual_r = row_map.get(canonical_r, canonical_r)
            data[canonical_r] = ws.cell(row=actual_r, column=col).value

        rate_comps = {}
        for i, start in enumerate(RATE_COMP_STARTS, 1):
            rate_comps[i] = {
                "name": ws.cell(row=start+1, column=col).value,
                "custom_generic": ws.cell(row=start+2, column=col).value,
                "energy_rate": ws.cell(row=start+3, column=col).value,
                "escalator": ws.cell(row=start+4, column=col).value,
                "start_date": ws.cell(row=start+5, column=col).value,
                "term": ws.cell(row=start+6, column=col).value,
                "discount": ws.cell(row=start+7, column=col).value,
                "ucb_fee": ws.cell(row=start+8, column=col).value,
                "equity_on": safe_float(ws.cell(row=EQUITY_RATE_TOGGLE_START+(i-1), column=col).value),
                "debt_on": safe_float(ws.cell(row=DEBT_RATE_TOGGLE_START+(i-1), column=col).value),
                "appraisal_on": safe_float(ws.cell(row=APPRAISAL_RATE_TOGGLE_START+(i-1), column=col).value),
            }

        # For Custom rate components, the escalator input on Project Inputs
        # is not used — the rate curve defines year-by-year rates directly.
        # Suppress the escalator to avoid showing misleading values.
        for i, start in enumerate(RATE_COMP_STARTS, 1):
            cg = ws.cell(row=start+2, column=col).value  # Custom/Generic
            if cg and str(cg).strip().lower() == "custom":
                esc_canonical = start + 4  # escalator row for this rate comp
                if esc_canonical in data:
                    data[esc_canonical] = "N/A (Custom)"

        data["_debt_match_equity"] = ws.cell(row=400, column=col).value
        data["_appraisal_match_equity"] = ws.cell(row=512, column=col).value

        dscr_label = ws.cell(row=341, column=col).value
        dscr_schedule = {}
        for yr in range(1, 31):
            v = safe_float(ws.cell(row=341 + yr, column=col).value)
            if v is not None:
                dscr_schedule[yr] = v

        data["_front_back_toggle"] = ws.cell(row=320, column=col).value
        data["_debt_sizing_method"] = ws.cell(row=321, column=col).value

        clean_name = " | ".join(line.strip() for line in str(name_cell).strip().splitlines() if line.strip())
        projects[col] = {
            "name": clean_name, "toggle": is_on,
            "col_letter": openpyxl.utils.get_column_letter(col),
            "data": data, "rate_comps": rate_comps,
            "dscr_label": dscr_label, "dscr_schedule": dscr_schedule,
        }

    # Ops Sandbox
    ops_sandbox = {"revenue_adders": [], "opex_overrides": []}
    if "Ops Sandbox" in wb.sheetnames:
        ws_ops = wb["Ops Sandbox"]
        ops_sandbox["live_project"] = ws_ops.cell(row=15, column=4).value
        for r in range(19, 29):
            label = ws_ops.cell(row=r, column=2).value
            if label and "Placeholder" not in str(label):
                ops_sandbox["revenue_adders"].append({
                    "label": str(label).strip(),
                    "annual": safe_float(ws_ops.cell(row=r, column=3).value),
                    "equity": safe_float(ws_ops.cell(row=r, column=4).value) or 0,
                    "debt": safe_float(ws_ops.cell(row=r, column=5).value) or 0,
                    "appraisal": safe_float(ws_ops.cell(row=r, column=6).value) or 0,
                    "npv_total": safe_float(ws_ops.cell(row=r, column=8).value),
                })
        for r in range(36, 46):
            label = ws_ops.cell(row=r, column=2).value
            if label and "Placeholder" not in str(label):
                ops_sandbox["opex_overrides"].append({
                    "label": str(label).strip(),
                    "annual": safe_float(ws_ops.cell(row=r, column=3).value),
                    "equity": safe_float(ws_ops.cell(row=r, column=4).value) or 0,
                    "debt": safe_float(ws_ops.cell(row=r, column=5).value) or 0,
                    "appraisal": safe_float(ws_ops.cell(row=r, column=6).value) or 0,
                    "npv_total": safe_float(ws_ops.cell(row=r, column=8).value),
                })
        for attr, row in [("equity_adder_npv", 31), ("debt_adder_npv", 32), ("appraisal_adder_npv", 33),
                          ("equity_opex_npv", 48), ("debt_opex_npv", 49), ("appraisal_opex_npv", 50)]:
            ops_sandbox[attr] = safe_float(ws_ops.cell(row=row, column=8).value)

    wb.close()
    return {"projects": projects, "ops_sandbox": ops_sandbox}


def get_projects(model_result):
    if isinstance(model_result, dict) and "projects" in model_result:
        return model_result["projects"]
    return model_result


def get_ops_sandbox(model_result):
    if isinstance(model_result, dict) and "ops_sandbox" in model_result:
        return model_result["ops_sandbox"]
    return {}


@st.cache_data(show_spinner=False)
def load_data_room(files):
    data_room = {}
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
        fname = Path(f.name).stem if hasattr(f, 'name') else str(f)
        sheets_info = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=False):
                rows.append([c.value for c in row])
            sheets_info[sn] = rows
        data_room[fname] = {"sheets": wb.sheetnames, "data": sheets_info}
        wb.close()
    return data_room
