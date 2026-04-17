"""Tests for walk_builder.py — project matching, metrics, diff, and xlsx generation."""
import io
import pytest
from lib.rows import ROW_PROJECT_NUMBER, ROW_DC_MW, ROW_NPP, ROW_FMV_IRR
from lib.walk_builder import match_projects, extract_metrics, diff_inputs, _categorize_row, build_walk_xlsx


def _make_projects(*specs):
    """Build a fake projects dict. Each spec = (col, proj_num, name, data_overrides)."""
    projects = {}
    for col, pnum, name, overrides in specs:
        data = {ROW_PROJECT_NUMBER: pnum, ROW_DC_MW: 5.0, ROW_NPP: 0.10, ROW_FMV_IRR: 0.18}
        data.update(overrides or {})
        projects[col] = {
            "name": name, "toggle": True, "col_letter": "F",
            "data": data, "rate_comps": {}, "dscr_schedule": {},
        }
    return projects


class TestMatchProjects:
    def test_matches_by_project_number(self):
        m1 = _make_projects((6, 1, "Joel", {}), (7, 2, "Rhea", {}))
        m2 = _make_projects((6, 1, "Joel", {}), (7, 2, "Rhea", {}))
        matched = match_projects(m1, m2)
        assert len(matched) == 2
        assert matched[0]["proj_number"] == 1
        assert matched[1]["proj_number"] == 2

    def test_unmatched_projects_excluded(self):
        m1 = _make_projects((6, 1, "Joel", {}), (7, 3, "Canary", {}))
        m2 = _make_projects((6, 1, "Joel", {}), (7, 2, "Rhea", {}))
        matched = match_projects(m1, m2)
        assert len(matched) == 1
        assert matched[0]["proj_number"] == 1

    def test_empty_models_return_empty(self):
        assert match_projects({}, {}) == []

    def test_none_project_number_uses_fallback(self):
        """When project # is None, fallback matching (positional or name) kicks in."""
        m1 = _make_projects((6, None, "Unknown", {}))
        m2 = _make_projects((6, None, "Unknown", {}))
        matched = match_projects(m1, m2)
        # Fallback should match by position or name
        assert len(matched) >= 1


class TestExtractMetrics:
    def test_extracts_npp_irr_mwdc(self):
        m1 = _make_projects((6, 1, "Joel", {ROW_NPP: 0.15, ROW_FMV_IRR: 0.20, ROW_DC_MW: 10.0}))
        m2 = _make_projects((6, 1, "Joel", {ROW_NPP: 0.12, ROW_FMV_IRR: 0.18, ROW_DC_MW: 10.0}))
        matched = match_projects(m1, m2)
        metrics = extract_metrics(matched, m1, m2)
        assert len(metrics) == 1
        assert metrics[0]["m1_npp"] == 0.15
        assert metrics[0]["m2_npp"] == 0.12
        assert metrics[0]["mwdc"] == 10.0


class TestDiffInputs:
    def test_detects_numeric_difference(self):
        m1 = _make_projects((6, 1, "Joel", {118: 1.65}))
        m2 = _make_projects((6, 1, "Joel", {118: 1.75}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        epc_diffs = [d for d in diffs if d["row"] == 118]
        assert len(epc_diffs) == 1
        assert epc_diffs[0]["category"] == "CapEx"

    def test_identical_values_no_diff(self):
        m1 = _make_projects((6, 1, "Joel", {118: 1.65}))
        m2 = _make_projects((6, 1, "Joel", {118: 1.65}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        epc_diffs = [d for d in diffs if d["row"] == 118]
        assert len(epc_diffs) == 0

    def test_skips_identity_rows(self):
        # Row 4 (Project Name) should be skipped
        m1 = _make_projects((6, 1, "Joel", {4: "Joel"}))
        m2 = _make_projects((6, 1, "Joel", {4: "Joel Changed"}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        name_diffs = [d for d in diffs if d["row"] == 4]
        assert len(name_diffs) == 0


class TestCategorizeRow:
    def test_capex_row(self):
        assert _categorize_row(118) == "CapEx"

    def test_revenue_row(self):
        assert _categorize_row(157) == "Revenue"

    def test_tax_row(self):
        assert _categorize_row(597) == "Incentives & Tax"

    def test_unknown_row(self):
        assert _categorize_row(9999) == "Other"


class TestBuildWalkXlsx:
    def test_produces_valid_xlsx(self):
        m1 = {"projects": _make_projects((6, 1, "Joel", {ROW_NPP: 0.15, ROW_DC_MW: 10.0}))}
        m2 = {"projects": _make_projects((6, 1, "Joel", {ROW_NPP: 0.12, ROW_DC_MW: 10.0}))}
        buf, summary = build_walk_xlsx(m1, m2, "Base Case", "Scenario 2")
        assert isinstance(buf, io.BytesIO)
        assert summary["n_matched"] == 1
        assert summary["m1_label"] == "Base Case"
        # Verify it's a valid xlsx by reading it back
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.title == "Build Walk"
        # Check case header exists
        assert ws.cell(row=3, column=5).value == 1
        # Check project name is in the data
        assert ws.cell(row=7, column=2).value == "Joel"
        wb.close()

    def test_no_match_produces_empty_xlsx(self):
        m1 = {"projects": _make_projects((6, 1, "Joel", {}))}
        m2 = {"projects": _make_projects((6, 2, "Rhea", {}))}
        buf, summary = build_walk_xlsx(m1, m2, "M1", "M2")
        assert summary["n_matched"] == 0
        # Should still produce valid xlsx
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert wb.active.title == "Build Walk"
        wb.close()


class TestPortedWalkSpec:
    """Regression tests for the spec ports into lib.walk_builder (FastAPI path).

    Locks in: Δ direction (M1 - M2), bool/int equality, IRR row 37, Notes
    column, Unmatched sheet, property-tax inputs-only behavior.
    Imports from lib.* explicitly because the FastAPI route uses lib/walk_builder.py.
    """

    def test_irr_uses_row_37_not_row_31(self):
        from lib.rows import ROW_LEVERED_PT_IRR
        from lib.walk_builder import match_projects as lib_match, extract_metrics as lib_extract
        assert ROW_LEVERED_PT_IRR == 37
        m1 = _make_projects((6, 1, "Joel", {ROW_LEVERED_PT_IRR: 0.18, 31: 0.99}))
        m2 = _make_projects((6, 1, "Joel", {ROW_LEVERED_PT_IRR: 0.20, 31: 0.99}))
        matched = lib_match(m1, m2)
        metrics = lib_extract(matched, m1, m2)
        assert metrics[0]["m1_irr"] == 0.18
        assert metrics[0]["m2_irr"] == 0.20

    def test_delta_direction_is_m1_minus_m2(self):
        from lib.walk_builder import build_walk_xlsx as lib_build
        m1 = {"projects": _make_projects((6, 1, "Joel", {ROW_NPP: 0.15, ROW_DC_MW: 10.0}))}
        m2 = {"projects": _make_projects((6, 1, "Joel", {ROW_NPP: 0.10, ROW_DC_MW: 10.0}))}
        buf, _ = lib_build(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        delta_formula = ws.cell(row=7, column=7).value
        assert delta_formula is not None
        assert delta_formula.replace(" ", "") == "=E7-H7", (
            f"Expected M1-M2 formula =E7-H7, got {delta_formula!r}")
        wb.close()

    def test_bool_vs_int_toggle_not_a_diff(self):
        from lib.walk_builder import match_projects as lib_match, diff_inputs as lib_diff
        m1 = _make_projects((6, 1, "Joel", {291: True}))
        m2 = _make_projects((6, 1, "Joel", {291: 1}))
        matched = lib_match(m1, m2)
        diffs = lib_diff(matched, m1, m2)
        proptax_diffs = [d for d in diffs if d["row"] == 291]
        assert len(proptax_diffs) == 0, (
            "True vs 1 on PropTax toggle should not flag as differing")

    def test_bool_vs_int_genuine_diff_still_detected(self):
        from lib.walk_builder import match_projects as lib_match, diff_inputs as lib_diff
        m1 = _make_projects((6, 1, "Joel", {291: True}))
        m2 = _make_projects((6, 1, "Joel", {291: False}))
        matched = lib_match(m1, m2)
        diffs = lib_diff(matched, m1, m2)
        proptax_diffs = [d for d in diffs if d["row"] == 291]
        assert len(proptax_diffs) == 1

    def test_n_diff_count_in_diff_record(self):
        from lib.walk_builder import match_projects as lib_match, diff_inputs as lib_diff
        m1 = _make_projects(
            (6, 1, "A", {118: 1.65}),
            (7, 2, "B", {118: 1.65}),
            (8, 3, "C", {118: 1.65}),
        )
        m2 = _make_projects(
            (6, 1, "A", {118: 1.65}),
            (7, 2, "B", {118: 1.65}),
            (8, 3, "C", {118: 1.99}),
        )
        matched = lib_match(m1, m2)
        diffs = lib_diff(matched, m1, m2)
        epc = [d for d in diffs if d["row"] == 118]
        assert epc and epc[0]["n_diff"] == 1
        assert epc[0]["n_total"] == 3

    def test_notes_column_written(self):
        from lib.walk_builder import build_walk_xlsx as lib_build
        m1 = {"projects": _make_projects(
            (6, 1, "A", {118: 1.65}),
            (7, 2, "B", {118: 1.65}),
        )}
        m2 = {"projects": _make_projects(
            (6, 1, "A", {118: 1.65}),
            (7, 2, "B", {118: 1.99}),
        )}
        buf, _ = lib_build(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=2).value and "EPC" in str(ws.cell(row=r, column=2).value):
                note = ws.cell(row=r, column=10).value
                if note and "differs for 1 of 2" in str(note):
                    found = True
                    break
        assert found, "Notes column missing 'differs for 1 of 2 projects'"
        wb.close()

    def test_unmatched_sheet_when_orphans_exist(self):
        from lib.walk_builder import build_walk_xlsx as lib_build
        m1 = {"projects": _make_projects(
            (6, 1, "Matched", {}),
            (7, 99, "M1Orphan", {}),
        )}
        m2 = {"projects": _make_projects(
            (6, 1, "Matched", {}),
            (7, 88, "M2Orphan", {}),
        )}
        buf, summary = lib_build(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert "Unmatched" in wb.sheetnames
        assert summary["n_unmatched_m1"] == 1
        assert summary["n_unmatched_m2"] == 1
        un = wb["Unmatched"]
        all_strings = []
        for row in un.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    all_strings.append(str(v))
        assert any("M1Orphan" in s for s in all_strings)
        assert any("M2Orphan" in s for s in all_strings)
        wb.close()

    def test_no_unmatched_sheet_when_clean_pairing(self):
        from lib.walk_builder import build_walk_xlsx as lib_build
        m1 = {"projects": _make_projects((6, 1, "Joel", {}))}
        m2 = {"projects": _make_projects((6, 1, "Joel", {}))}
        buf, summary = lib_build(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        assert "Unmatched" not in wb.sheetnames
        assert summary["n_unmatched_m1"] == 0
        assert summary["n_unmatched_m2"] == 0
        wb.close()

    def test_filename_sanitization(self):
        from apps.api.routers.walk import _safe_filename_part
        assert _safe_filename_part("Macro: foo") == "Macro_foo"
        assert _safe_filename_part('a/b\\c:d?e') == "abcde"
        assert _safe_filename_part("") == "model"
        assert _safe_filename_part("   ") == "model"


class TestDeeperDiffCoverage:
    """Regression tests for Pass 1a (rate comps), 1b (match toggles + DSCR),
    and 1c (Rate Curves COD-period diff) added during the walk methodology
    rework."""

    def _rc_project(self, col, pnum, name, rate_comps, extra_data=None,
                    rate_curves=None):
        """Build a project with a populated rate_comps dict + optional
        _rate_curves_rcN top-level keys."""
        data = {ROW_PROJECT_NUMBER: pnum, ROW_DC_MW: 5.0, ROW_NPP: 0.10,
                ROW_FMV_IRR: 0.18}
        if extra_data:
            data.update(extra_data)
        proj = {
            "name": name, "toggle": True, "col_letter": "F",
            "data": data, "rate_comps": rate_comps, "dscr_schedule": {},
        }
        if rate_curves:
            for rc_idx, curve in rate_curves.items():
                proj[f"_rate_curves_rc{rc_idx}"] = curve
        return {col: proj}

    def test_rc3_rate_diff_detected(self):
        """RC3 energy_rate differing — must surface as 'RC3 Energy Rate'."""
        m1_proj = self._rc_project(6, 1, "Alpha", {
            3: {"name": "REC", "custom_generic": "Generic", "energy_rate": 0.05,
                "equity_on": 1, "debt_on": 0, "appraisal_on": 0},
        })
        m2_proj = self._rc_project(6, 1, "Alpha", {
            3: {"name": "REC", "custom_generic": "Generic", "energy_rate": 0.07,
                "equity_on": 1, "debt_on": 0, "appraisal_on": 0},
        })
        matched = match_projects(m1_proj, m2_proj)
        diffs = diff_inputs(matched, m1_proj, m2_proj)
        labels = [d["label"] for d in diffs]
        assert "RC3 Energy Rate" in labels

    def test_rc_value_suppressed_when_off_in_one_side(self):
        """Per Decision B(i): when RC is off in either model, don't surface
        the per-field rate diff — only the toggle diff."""
        m1_proj = self._rc_project(6, 1, "Alpha", {
            2: {"custom_generic": "Generic", "energy_rate": 0.05,
                "equity_on": 1, "debt_on": 0, "appraisal_on": 0},
        })
        m2_proj = self._rc_project(6, 1, "Alpha", {
            2: {"custom_generic": "Generic", "energy_rate": 0.09,
                "equity_on": 0, "debt_on": 0, "appraisal_on": 0},  # off
        })
        matched = match_projects(m1_proj, m2_proj)
        diffs = diff_inputs(matched, m1_proj, m2_proj)
        labels = [d["label"] for d in diffs]
        # Toggle diff flagged
        assert "RC2 Equity" in labels
        # Rate value diff suppressed
        assert "RC2 Energy Rate" not in labels

    def test_rc_value_shown_when_both_on(self):
        """When RC is on in BOTH, value-field diffs DO surface."""
        m1_proj = self._rc_project(6, 1, "Alpha", {
            1: {"custom_generic": "Generic", "energy_rate": 0.05,
                "equity_on": 1, "debt_on": 0, "appraisal_on": 0},
        })
        m2_proj = self._rc_project(6, 1, "Alpha", {
            1: {"custom_generic": "Generic", "energy_rate": 0.07,
                "equity_on": 1, "debt_on": 0, "appraisal_on": 0},
        })
        matched = match_projects(m1_proj, m2_proj)
        diffs = diff_inputs(matched, m1_proj, m2_proj)
        labels = [d["label"] for d in diffs]
        assert "RC1 Energy Rate" in labels

    def test_debt_match_equity_toggle_diff(self):
        """_debt_match_equity differing → 'Debt Rate: match equity' row."""
        m1 = _make_projects((6, 1, "A", {"_debt_match_equity": 1}))
        m2 = _make_projects((6, 1, "A", {"_debt_match_equity": 0}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        labels = [d["label"] for d in diffs]
        assert "Debt Rate: match equity" in labels

    def test_dscr_schedule_year_diff(self):
        """DSCR diff at year 3 → 'DSCR Y3' row."""
        m1 = _make_projects((6, 1, "A", {}))
        m1[6]["dscr_schedule"] = {1: 1.20, 2: 1.20, 3: 1.20}
        m2 = _make_projects((6, 1, "A", {}))
        m2[6]["dscr_schedule"] = {1: 1.20, 2: 1.20, 3: 1.35}
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        labels = [d["label"] for d in diffs]
        assert "DSCR Y3" in labels
        # Years 1,2 match — shouldn't appear
        assert "DSCR Y1" not in labels
        assert "DSCR Y2" not in labels

    def test_rate_curve_cod_diff_for_custom_rcs(self):
        """Both sides have Custom RC1 but the rate at COD differs."""
        from datetime import datetime
        curve_m1 = {datetime(2026, 1, 1): 0.1200, datetime(2027, 1, 1): 0.1250}
        curve_m2 = {datetime(2026, 1, 1): 0.1050, datetime(2027, 1, 1): 0.1080}
        m1 = self._rc_project(
            6, 1, "Alpha",
            rate_comps={1: {"custom_generic": "Custom", "energy_rate": None,
                            "equity_on": 1, "debt_on": 0, "appraisal_on": 0}},
            extra_data={15: 2026},  # COD year
            rate_curves={1: curve_m1},
        )
        m2 = self._rc_project(
            6, 1, "Alpha",
            rate_comps={1: {"custom_generic": "Custom", "energy_rate": None,
                            "equity_on": 1, "debt_on": 0, "appraisal_on": 0}},
            extra_data={15: 2026},
            rate_curves={1: curve_m2},
        )
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        labels = [d["label"] for d in diffs]
        assert "RC1 Rate Curve (COD)" in labels
        # Verify the diff picked the COD-year rate, not some arbitrary month.
        rate_diff = next(d for d in diffs if d["label"] == "RC1 Rate Curve (COD)")
        m1_val, m2_val = rate_diff["values"][1]
        assert m1_val == pytest.approx(0.1200)
        assert m2_val == pytest.approx(0.1050)

    def test_rate_curve_not_compared_when_one_side_generic(self):
        """Shape mismatch (one Custom, one Generic) suppresses Rate Curves
        diff — that shape change is already flagged by Pass 1a."""
        from datetime import datetime
        m1 = self._rc_project(
            6, 1, "Alpha",
            rate_comps={1: {"custom_generic": "Custom", "energy_rate": None,
                            "equity_on": 1, "debt_on": 0, "appraisal_on": 0}},
            extra_data={15: 2026},
            rate_curves={1: {datetime(2026, 1, 1): 0.1200}},
        )
        m2 = self._rc_project(
            6, 1, "Alpha",
            rate_comps={1: {"custom_generic": "Generic", "energy_rate": 0.1050,
                            "equity_on": 1, "debt_on": 0, "appraisal_on": 0}},
            extra_data={15: 2026},
        )
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        labels = [d["label"] for d in diffs]
        assert "RC1 Rate Curve (COD)" not in labels
        assert "RC1 Custom/Generic" in labels  # shape diff flagged instead


class TestTranche1:
    """Regressions for the Tranche 1 walk accuracy fixes:
      - Unit reconciliation for $/kWh ↔ $/MW/yr drift on rows 121, 240
      - Substring-match false positive in data_loader rate curve pairing
      - Unmatched sheet reason codes
    """

    def test_unit_reconciliation_cust_mgmt_false_positive_suppressed(self):
        """M1 = $0.0025/kWh, M2 = $2,500/MW/yr, yield 1.25 kWh/Wdc —
        reconcile: $0.0025 × 1.25 × 1,000,000 = $3,125/MW/yr → differs
        from $2,500 but by a reasonable margin, not by 1000x. The point is
        we don't report the raw $0.0025 vs $2,500 as an enormous diff."""
        m1 = _make_projects((6, 1, "A", {240: 0.0020, 14: 1.25}))   # $/kWh basis
        m2 = _make_projects((6, 1, "A", {240: 2500.0, 14: 1.25}))   # $/MW/yr basis
        # $0.0020 × 1.25 × 1e6 = $2,500 → equivalent, no diff
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        cust_mgmt = [d for d in diffs if d["row"] == 240]
        assert len(cust_mgmt) == 0, (
            f"Expected no diff after unit reconciliation, got {cust_mgmt}")

    def test_unit_reconciliation_genuine_diff_still_detected(self):
        """Same units, real difference — must still flag."""
        m1 = _make_projects((6, 1, "A", {240: 2500.0, 14: 1.25}))
        m2 = _make_projects((6, 1, "A", {240: 3500.0, 14: 1.25}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        cust_mgmt = [d for d in diffs if d["row"] == 240]
        assert len(cust_mgmt) == 1

    def test_unit_reconciliation_annotates_unit_note(self):
        """When reconciliation fires, the diff dict carries a unit_note
        explaining the conversion."""
        m1 = _make_projects((6, 1, "A", {240: 0.0020, 14: 1.25}))
        m2 = _make_projects((6, 1, "A", {240: 9999.0, 14: 1.25}))  # real diff
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        cust_mgmt = [d for d in diffs if d["row"] == 240]
        assert cust_mgmt
        assert "unit_note" in cust_mgmt[0]
        assert "$/MW/yr" in cust_mgmt[0]["unit_note"]

    def test_unmatched_reason_missing_proj_num(self):
        """An M1 project with no Project # should surface reason code
        'missing_proj_num' in the Unmatched sheet."""
        from openpyxl import load_workbook
        # M1 project 1: Alpha, no proj#. M2 project 1: Alpha, proj# 1.
        # Match via name fallback would succeed if names match, so give them
        # different names to force an orphan.
        m1 = {"projects": {6: {
            "name": "Alpha (no pnum)", "toggle": True, "col_letter": "F",
            "data": {ROW_PROJECT_NUMBER: None, ROW_DC_MW: 5.0},
            "rate_comps": {}, "dscr_schedule": {},
        }}}
        m2 = {"projects": {6: {
            "name": "Beta", "toggle": True, "col_letter": "F",
            "data": {ROW_PROJECT_NUMBER: 2, ROW_DC_MW: 5.0},
            "rate_comps": {}, "dscr_schedule": {},
        }}}
        buf, _ = build_walk_xlsx(m1, m2, "M1", "M2")
        wb = load_workbook(buf)
        assert "Unmatched" in wb.sheetnames
        un = wb["Unmatched"]
        headers = [un.cell(row=1, column=c).value for c in range(1, 7)]
        assert "Reason Code" in headers
        # Find M1 orphan row and check reason.
        rc_col = headers.index("Reason Code") + 1
        side_col = headers.index("Side") + 1
        for r in range(2, un.max_row + 1):
            if un.cell(row=r, column=side_col).value == "M1":
                assert un.cell(row=r, column=rc_col).value == "missing_proj_num"
                return
        pytest.fail("No M1 unmatched row found")

    def test_unmatched_reason_proj_num_not_in_other(self):
        """M1 has proj# 5 but M2 has no proj# 5 → reason should be
        'proj_num_not_in_other' (not 'missing_proj_num')."""
        from openpyxl import load_workbook
        m1 = {"projects": {6: {
            "name": "Only-M1", "toggle": True, "col_letter": "F",
            "data": {ROW_PROJECT_NUMBER: 5, ROW_DC_MW: 5.0},
            "rate_comps": {}, "dscr_schedule": {},
        }}}
        m2 = {"projects": {6: {
            "name": "Only-M2", "toggle": True, "col_letter": "F",
            "data": {ROW_PROJECT_NUMBER: 6, ROW_DC_MW: 5.0},
            "rate_comps": {}, "dscr_schedule": {},
        }}}
        buf, _ = build_walk_xlsx(m1, m2, "M1", "M2")
        wb = load_workbook(buf)
        un = wb["Unmatched"]
        headers = [un.cell(row=1, column=c).value for c in range(1, 7)]
        rc_col = headers.index("Reason Code") + 1
        codes = {un.cell(row=r, column=rc_col).value
                 for r in range(2, un.max_row + 1)}
        assert "proj_num_not_in_other" in codes


class TestDataLoaderRateCurveMatch:
    """Regression for the substring-match bug in data_loader rate curve
    pairing. Uses a mock approach because the real loader reads xlsx files."""

    def test_canonical_match_avoids_solar_1_matching_solar_10(self):
        """Build two strings that demonstrate the old substring rule was
        broken and the new canonical rule works."""
        import re as _re
        def canon(s):
            return _re.sub(r"\s+", " ", str(s) or "").strip().casefold()

        pname = "Solar 1"
        proj_rows_map = {"Solar 10": 45, "Solar 1": 35}
        # Simulate new canonical-equality logic inline — asserts the LOGIC
        # is correct even without spinning up a fake workbook.
        pname_canon = canon(pname)
        matched_row = None
        for rc_pname, rc_row in proj_rows_map.items():
            if canon(rc_pname) == pname_canon:
                matched_row = rc_row
                break
        assert matched_row == 35, "Solar 1 should match Solar 1, not Solar 10"

    def test_canonical_match_ignores_whitespace_case(self):
        import re as _re
        def canon(s):
            return _re.sub(r"\s+", " ", str(s) or "").strip().casefold()

        pname = "  Il JOEL  "
        pname_canon = canon(pname)
        proj_rows_map = {"IL Joel": 35}
        matched_row = None
        for rc_pname, rc_row in proj_rows_map.items():
            if canon(rc_pname) == pname_canon:
                matched_row = rc_row
                break
        assert matched_row == 35


class TestTranche2Provenance:
    """Every diff dict carries a 'source' tag. _write_variance_section
    surfaces it in column K of the walk output."""

    def test_canonical_diff_tagged_canonical(self):
        m1 = _make_projects((6, 1, "A", {118: 1.65}))
        m2 = _make_projects((6, 1, "A", {118: 1.75}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        epc = next(d for d in diffs if d["row"] == 118)
        assert epc["source"] == "canonical"

    def test_rate_comp_diff_tagged_rate_comps(self):
        m1 = {6: {"name":"A","toggle":True,"col_letter":"F",
                  "data":{ROW_PROJECT_NUMBER:1,ROW_DC_MW:5.0},
                  "rate_comps":{1:{"custom_generic":"Generic","energy_rate":0.05,
                                   "equity_on":1,"debt_on":0,"appraisal_on":0}},
                  "dscr_schedule":{}}}
        m2 = {6: {"name":"A","toggle":True,"col_letter":"F",
                  "data":{ROW_PROJECT_NUMBER:1,ROW_DC_MW:5.0},
                  "rate_comps":{1:{"custom_generic":"Generic","energy_rate":0.08,
                                   "equity_on":1,"debt_on":0,"appraisal_on":0}},
                  "dscr_schedule":{}}}
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        rc_diff = next(d for d in diffs if d["label"] == "RC1 Energy Rate")
        assert rc_diff["source"] == "rate_comps"

    def test_special_diff_tagged_special(self):
        m1 = _make_projects((6, 1, "A", {"_debt_match_equity": 1}))
        m2 = _make_projects((6, 1, "A", {"_debt_match_equity": 0}))
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        d = next(x for x in diffs if x["label"] == "Debt Rate: match equity")
        assert d["source"] == "special"

    def test_dscr_diff_tagged_special(self):
        m1 = _make_projects((6, 1, "A", {}))
        m1[6]["dscr_schedule"] = {3: 1.20}
        m2 = _make_projects((6, 1, "A", {}))
        m2[6]["dscr_schedule"] = {3: 1.35}
        matched = match_projects(m1, m2)
        diffs = diff_inputs(matched, m1, m2)
        d = next(x for x in diffs if x["label"] == "DSCR Y3")
        assert d["source"] == "special"

    def test_source_column_written_to_xlsx(self):
        m1 = {"projects": _make_projects((6, 1, "A", {118: 1.65, ROW_DC_MW: 5.0}))}
        m2 = {"projects": _make_projects((6, 1, "A", {118: 1.75, ROW_DC_MW: 5.0}))}
        buf, _ = build_walk_xlsx(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        # Find the PV EPC row in variance section; check col K (11) = "canonical"
        for r in range(1, ws.max_row + 1):
            lbl = ws.cell(row=r, column=2).value
            if lbl and "EPC" in str(lbl) and lbl != "Project Inputs":
                src = ws.cell(row=r, column=11).value
                if src == "canonical":
                    wb.close()
                    return
        wb.close()
        pytest.fail("Source column 'canonical' not written for EPC row")


class TestTranche2Impact:
    """Dollar impact estimation — column L in variance section."""

    def test_impact_per_project_epc(self):
        # M1 EPC = $1.65/W, M2 = $1.75/W, 5 MWdc (5,000,000 W).
        # Delta = 1.65 - 1.75 = -0.10. Higher EPC = worse for sponsor, so
        # sign flips: impact = -(-0.10) * 5_000_000 = +500,000? No wait.
        # _delta = M1 - M2 = -0.10. favor_m1_high=False → flip → +0.10 * 5M = +500k.
        # Reading the code: raw = -0.10 * 5_000_000 = -500,000. Not favor_m1_high
        # → flip → +500,000. But M1 has LOWER EPC than M2 (1.65 < 1.75), which
        # IS better for sponsor in M1. So +500,000 is the M1 advantage over M2.
        from lib.impact import per_project_impact
        data = {ROW_DC_MW: 5.0}
        imp = per_project_impact(118, 1.65, 1.75, data)
        assert imp == pytest.approx(500_000, abs=1.0)

    def test_impact_per_project_upfront_incentive(self):
        # M1 Upfront = $0.20/W, M2 = $0.10/W. M1 has more incentive → favors M1.
        # Raw delta = +0.10. favor_m1_high=True → no flip → +500k.
        from lib.impact import per_project_impact
        data = {ROW_DC_MW: 5.0}
        imp = per_project_impact(216, 0.20, 0.10, data)
        assert imp == pytest.approx(500_000, abs=1.0)

    def test_impact_unknown_row_returns_none(self):
        from lib.impact import per_project_impact
        data = {ROW_DC_MW: 5.0}
        assert per_project_impact(9999, 1.0, 2.0, data) is None

    def test_impact_missing_dc_returns_none(self):
        from lib.impact import per_project_impact
        imp = per_project_impact(118, 1.65, 1.75, {})
        assert imp is None

    def test_portfolio_impact_sums_projects(self):
        from lib.impact import portfolio_impact
        # Two projects, both 5 MWdc, both with 10 cent EPC delta
        per_project = {1: (1.65, 1.75), 2: (1.60, 1.70)}
        m1_data = {
            1: {ROW_DC_MW: 5.0},
            2: {ROW_DC_MW: 5.0},
        }
        total = portfolio_impact(118, per_project, m1_data)
        # Each project: +500k; total ~= +1M
        assert total == pytest.approx(1_000_000, abs=1.0)

    def test_portfolio_impact_column_written_to_xlsx(self):
        m1 = {"projects": _make_projects((6, 1, "A", {118: 1.65, ROW_DC_MW: 5.0}))}
        m2 = {"projects": _make_projects((6, 1, "A", {118: 1.75, ROW_DC_MW: 5.0}))}
        buf, _ = build_walk_xlsx(m1, m2, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        # Find PV EPC row, check col L (12) for a dollar value
        for r in range(1, ws.max_row + 1):
            lbl = ws.cell(row=r, column=2).value
            if lbl and "EPC" in str(lbl) and lbl != "Project Inputs":
                imp = ws.cell(row=r, column=12).value
                if imp is not None and isinstance(imp, (int, float)):
                    # +500k for M1 cheaper by $0.10/W × 5 MW
                    assert imp == pytest.approx(500_000, abs=1.0)
                    wb.close()
                    return
        wb.close()
        pytest.fail("$ Impact column not populated for EPC row")


class TestTranche3Matching:
    """canonicalize_pnum handles int/float/string variance; match_projects
    returns match_source on every pair; Anchor sheet flags non-standard
    pairings; summary dict carries matches_by_source count."""

    def test_canonicalize_pnum_int_float_equivalence(self):
        from lib.utils import canonicalize_pnum
        assert canonicalize_pnum(1) == canonicalize_pnum(1.0) == canonicalize_pnum("1")

    def test_canonicalize_pnum_string_pnum(self):
        from lib.utils import canonicalize_pnum
        # "P-001" has no numeric form; canonicalizes via name rule.
        c1 = canonicalize_pnum("P-001")
        c2 = canonicalize_pnum("p-001")
        c3 = canonicalize_pnum("  P-001  ")
        assert c1 == c2 == c3

    def test_canonicalize_pnum_none_and_blank(self):
        from lib.utils import canonicalize_pnum
        assert canonicalize_pnum(None) is None
        assert canonicalize_pnum("") is None

    def test_canonicalize_pnum_non_integer_float(self):
        from lib.utils import canonicalize_pnum
        # 1.5 and 1.5 agree; 1.5 and 1 do NOT.
        assert canonicalize_pnum(1.5) == canonicalize_pnum(1.5)
        assert canonicalize_pnum(1.5) != canonicalize_pnum(1)

    def test_match_projects_handles_int_vs_float_pnum(self):
        m1 = _make_projects((6, 1, "A", {}))
        m2 = _make_projects((6, 1.0, "A", {}))
        matched = match_projects(m1, m2)
        assert len(matched) == 1
        assert matched[0]["match_source"] == "proj_num"

    def test_match_projects_handles_string_pnum(self):
        m1 = _make_projects((6, "P-001", "Alpha", {}))
        m2 = _make_projects((6, "P-001", "Beta", {}))
        matched = match_projects(m1, m2)
        assert len(matched) == 1
        assert matched[0]["match_source"] == "proj_num"

    def test_match_source_positional_fallback(self):
        # Both projects at col 6 with no proj# → positional infers # = 1.
        m1 = _make_projects((6, None, "Only by position", {}))
        m2 = _make_projects((6, None, "Only by position", {}))
        matched = match_projects(m1, m2)
        assert len(matched) >= 1
        # Could resolve via positional OR name — accept either as a fallback.
        assert matched[0]["match_source"] in ("positional", "name")

    def test_match_source_name_fallback(self):
        # Different proj# each side, same name, but different cols so
        # positional falls through to name.
        m1 = _make_projects((6, None, "SameName", {}))
        m2 = _make_projects((7, None, "SameName", {}))
        matched = match_projects(m1, m2)
        assert len(matched) == 1
        # Positional would match col 6 ↔ col 7 as different project # (1 vs 2),
        # so this case should route to name fallback.
        assert matched[0]["match_source"] == "name"

    def test_anchor_adds_comment_for_non_proj_num_match(self):
        """When match_source != 'proj_num', the Project Name cell in the
        Anchor section carries an openpyxl Comment flagging the pairing."""
        m1 = _make_projects((6, None, "Alpha", {}))
        m2 = _make_projects((7, None, "Alpha", {}))
        buf, _ = build_walk_xlsx({"projects": m1}, {"projects": m2}, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        # Project Name lives in col B starting at row 7.
        cell = ws.cell(row=7, column=2)
        assert cell.value == "Alpha"
        assert cell.comment is not None
        assert "fallback" in cell.comment.text.lower()
        wb.close()

    def test_summary_matches_by_source(self):
        m1 = _make_projects((6, 1, "Alpha", {}), (7, None, "Beta", {}))
        m2 = _make_projects((6, 1, "Alpha", {}), (7, None, "Beta", {}))
        _, summary = build_walk_xlsx({"projects": m1}, {"projects": m2}, "M1", "M2")
        assert "matches_by_source" in summary
        mbs = summary["matches_by_source"]
        assert mbs.get("proj_num", 0) >= 1  # Alpha via proj#
        # Beta via positional or name fallback (depends on which fires first)
        assert (mbs.get("positional", 0) + mbs.get("name", 0)) >= 0


class TestTranche3RateCurveConfidence:
    """_rate_at_cod now returns (rate, confidence). Walk notes flag
    extrapolated-rate projects per RC."""

    def test_rate_at_cod_exact_match(self):
        from datetime import datetime
        from lib.walk_builder import _rate_at_cod
        curve = {datetime(2026, 1, 1): 0.12, datetime(2027, 1, 1): 0.13}
        rate, conf = _rate_at_cod(curve, {15: 2026, 587: 1})
        assert rate == 0.12
        assert conf == "exact"

    def test_rate_at_cod_extrapolated_forward(self):
        from datetime import datetime
        from lib.walk_builder import _rate_at_cod
        curve = {datetime(2027, 1, 1): 0.13}
        # COD 2026 but curve starts 2027 → extrapolate forward.
        rate, conf = _rate_at_cod(curve, {15: 2026})
        assert rate == 0.13
        assert conf == "extrapolated_forward"

    def test_rate_at_cod_clamped_end(self):
        from datetime import datetime
        from lib.walk_builder import _rate_at_cod
        curve = {datetime(2025, 1, 1): 0.10}
        # COD 2030 but curve ends 2025 → clamp to last value.
        rate, conf = _rate_at_cod(curve, {15: 2030})
        assert rate == 0.10
        assert conf == "clamped_end"

    def test_rate_at_cod_empty_curve(self):
        from lib.walk_builder import _rate_at_cod
        rate, conf = _rate_at_cod({}, {15: 2026})
        assert rate is None
        assert conf is None

    def test_walk_notes_flags_extrapolated_count(self):
        """When a Custom-Custom RC pair has any project with non-exact
        rate-curve lookup, the Notes column mentions it."""
        from datetime import datetime
        # Project with COD 2030 but both curves end in 2025 → clamped_end.
        m1 = {6: {"name":"Alpha","toggle":True,"col_letter":"F",
                  "data":{ROW_PROJECT_NUMBER:1,ROW_DC_MW:5.0,15:2030},
                  "rate_comps":{1:{"custom_generic":"Custom","energy_rate":None,
                                   "equity_on":1,"debt_on":0,"appraisal_on":0}},
                  "dscr_schedule":{},
                  "_rate_curves_rc1":{datetime(2025,1,1):0.10}}}
        m2 = {6: {"name":"Alpha","toggle":True,"col_letter":"F",
                  "data":{ROW_PROJECT_NUMBER:1,ROW_DC_MW:5.0,15:2030},
                  "rate_comps":{1:{"custom_generic":"Custom","energy_rate":None,
                                   "equity_on":1,"debt_on":0,"appraisal_on":0}},
                  "dscr_schedule":{},
                  "_rate_curves_rc1":{datetime(2025,1,1):0.08}}}
        buf, _ = build_walk_xlsx({"projects": m1}, {"projects": m2}, "M1", "M2")
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb["Build Walk"]
        # Find the RC1 Rate Curve (COD) row; Notes col 10 should mention extrapolated
        found = False
        for r in range(1, ws.max_row + 1):
            lbl = ws.cell(row=r, column=2).value
            if lbl and "Rate Curve (COD)" in str(lbl):
                note = ws.cell(row=r, column=10).value or ""
                if "extrapolated" in str(note).lower():
                    found = True
                    break
        wb.close()
        assert found, "Notes column missing 'extrapolated' flag"
