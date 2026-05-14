"""Bible Excel loader tests.

Build synthetic mini-workbooks rather than checking in real Pricing Bible
fixtures. The loader's job is structural — it must:
  - reject files without a CS tab
  - reject .xlsx without any overlayable rows
  - overlay numeric values from CS!H<row> onto the bundled cs_average
  - skip sentinel strings ("See Market Specific Assumptions") and blanks
  - preserve int-vs-float typing so golden snapshots stay stable
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from lib.bible_loader import (
    AVERAGE_COL_IDX,
    CS_TAB_ROW_MAP,
    BibleParseError,
    load_bible_from_excel,
)

# ---------------------------------------------------------------------------
# Synthetic workbook builders
# ---------------------------------------------------------------------------


def _make_cs_workbook(
    row_values: dict[int, object],
    *,
    sheet_name: str = "CS",
) -> io.BytesIO:
    """Build a mini xlsx with a single sheet of given name and CS!H cells.

    row_values maps Bible-row → cell value (number, string, or None).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row, val in row_values.items():
        ws.cell(row=row, column=AVERAGE_COL_IDX, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _all_bible_rows_populated(value: float = 1234.0) -> dict[int, object]:
    """Populate every mapped Bible row with the same value."""
    return {br: value for br in CS_TAB_ROW_MAP}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestRejection:
    def test_no_cs_tab_raises(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"  # not "CS"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with pytest.raises(BibleParseError, match="no 'CS' worksheet"):
            load_bible_from_excel(buf, filename="bad.xlsx")

    def test_unopenable_file_raises(self):
        buf = io.BytesIO(b"not a real xlsx file")
        with pytest.raises(BibleParseError, match="Could not open"):
            load_bible_from_excel(buf, filename="garbage.xlsx")

    def test_no_overlayable_rows_raises(self):
        # CS tab exists but every Bible row is a sentinel string
        buf = _make_cs_workbook({br: "See Market Specific Assumptions" for br in CS_TAB_ROW_MAP})
        with pytest.raises(BibleParseError, match="no numeric values"):
            load_bible_from_excel(buf, filename="sentinel-only.xlsx")


class TestOverlay:
    def test_overlay_populates_metadata(self):
        buf = _make_cs_workbook(_all_bible_rows_populated())
        bible = load_bible_from_excel(buf, label="Test Vintage", filename="t.xlsx")

        assert bible.label == "Test Vintage"
        assert bible.source == "t.xlsx"
        assert bible.vintage_id.startswith("upload-")
        # ISO datetime, ends in +00:00 (UTC)
        assert bible.uploaded_at.endswith("+00:00")

    def test_overlay_replaces_cs_average_values(self):
        # Use a distinctive value not in the bundled vintage
        buf = _make_cs_workbook(_all_bible_rows_populated(value=9999.0))
        bible = load_bible_from_excel(buf, filename="t.xlsx")

        for model_row in CS_TAB_ROW_MAP.values():
            assert bible.cs_average[model_row]["value"] == 9999

    def test_label_falls_back_to_filename(self):
        buf = _make_cs_workbook(_all_bible_rows_populated())
        bible = load_bible_from_excel(buf, label="", filename="fallback.xlsx")
        assert bible.label == "fallback.xlsx"

    def test_skip_blank_cells(self):
        # Populate only one row, leave the rest blank — should still succeed
        # since 1 row was overlaid
        bundled_row_118 = None
        from lib.audit.bible import Bible

        bundled = Bible.bundled_q1_2026()
        bundled_row_118 = bundled.cs_average[118]["value"]

        buf = _make_cs_workbook({11: 1.42})  # only EPC ($/W) row
        bible = load_bible_from_excel(buf, filename="sparse.xlsx")

        # Row 118 (EPC, mapped from Bible row 11) gets overlaid
        assert bible.cs_average[118]["value"] == pytest.approx(1.42)
        # Other rows keep their bundled values
        assert bible.cs_average[123]["value"] == bundled.cs_average[123]["value"]
        # Sanity: the overlay actually changed something
        assert bible.cs_average[118]["value"] != bundled_row_118


class TestTyping:
    """Int/float typing must survive — golden snapshots depend on it."""

    def test_integer_valued_floats_become_ints(self):
        # openpyxl typically returns floats for numeric cells. Loader must
        # downcast integer-valued floats so JSON snapshots match bundled.
        buf = _make_cs_workbook({35: 4750.0})  # O&M Preventative → row 225
        bible = load_bible_from_excel(buf, filename="t.xlsx")

        v = bible.cs_average[225]["value"]
        assert v == 4750
        assert isinstance(v, int)

    def test_non_integer_floats_stay_float(self):
        buf = _make_cs_workbook({11: 1.65})  # EPC $/W → row 118
        bible = load_bible_from_excel(buf, filename="t.xlsx")

        v = bible.cs_average[118]["value"]
        assert v == pytest.approx(1.65)
        assert isinstance(v, float)


class TestSentinels:
    def test_sentinel_string_skipped(self):
        # All-sentinel raised earlier; here, one sentinel + one number =
        # one overlay, no raise
        from lib.audit.bible import Bible

        bundled = Bible.bundled_q1_2026()
        bundled_row_225 = bundled.cs_average[225]["value"]

        buf = _make_cs_workbook(
            {
                11: 1.70,  # numeric — overlays
                35: "See Market Specific Assumptions",  # sentinel — skipped
            }
        )
        bible = load_bible_from_excel(buf, filename="mixed.xlsx")

        assert bible.cs_average[118]["value"] == pytest.approx(1.70)
        # Row 225 (O&M) keeps bundled value because Bible row 35 was sentinel
        assert bible.cs_average[225]["value"] == bundled_row_225
