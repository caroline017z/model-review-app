"""FastAPI tests for /api/bible/* endpoints.

Each test starts with a fresh BibleStore (bundled vintage active).
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from apps.api import bible_store as bible_store_module
from apps.api.bible_store import BibleStore
from apps.api.main import app
from lib.bible_loader import AVERAGE_COL_IDX, CS_TAB_ROW_MAP


@pytest.fixture
def client(monkeypatch):
    """TestClient with a fresh BibleStore, patched in both the module and
    the router import."""
    fresh = BibleStore()
    monkeypatch.setattr(bible_store_module, "bible_store", fresh)
    # The router imports `bible_store` by name from the module, so patch
    # the router's reference too.
    from apps.api.routers import bible as bible_router

    monkeypatch.setattr(bible_router, "bible_store", fresh)
    return TestClient(app)


def _make_valid_xlsx() -> bytes:
    """Build a minimum-viable Pricing Bible xlsx (CS sheet + one numeric row)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CS"
    for row in CS_TAB_ROW_MAP:
        ws.cell(row=row, column=AVERAGE_COL_IDX, value=1234.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestListAndActive:
    def test_list_returns_bundled(self, client):
        resp = client.get("/api/bible/vintages")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) == 1
        assert data[0]["vintage_id"] == "bundled-q1-2026"
        assert data[0]["is_active"] is True

    def test_active_returns_bundled(self, client):
        resp = client.get("/api/bible/active")
        assert resp.status_code == 200
        data = resp.json()
        assert data["vintage_id"] == "bundled-q1-2026"
        assert data["is_active"] is True
        # Detail shape: row counts present
        assert "cs_average_row_count" in data
        assert "market_entries_count" in data

    def test_get_vintage_by_id(self, client):
        resp = client.get("/api/bible/vintages/bundled-q1-2026")
        assert resp.status_code == 200
        assert resp.json()["vintage_id"] == "bundled-q1-2026"

    def test_get_missing_vintage_404(self, client):
        resp = client.get("/api/bible/vintages/does-not-exist")
        assert resp.status_code == 404


class TestUpload:
    def test_upload_valid_xlsx(self, client):
        content = _make_valid_xlsx()
        resp = client.post(
            "/api/bible/vintages",
            files={"file": ("test.xlsx", content, "application/octet-stream")},
            params={"label": "My Test Vintage", "set_active": True},
        )
        assert resp.status_code == 200, resp.text
        data = resp.json()
        assert data["label"] == "My Test Vintage"
        assert data["source"] == "test.xlsx"
        assert data["is_active"] is True
        assert data["overlaid_row_count"] > 0

    def test_upload_makes_it_active(self, client):
        content = _make_valid_xlsx()
        upload = client.post(
            "/api/bible/vintages",
            files={"file": ("makeactive.xlsx", content, "application/octet-stream")},
        )
        vid = upload.json()["vintage_id"]

        # Verify via /active
        active = client.get("/api/bible/active").json()
        assert active["vintage_id"] == vid

    def test_upload_without_setting_active(self, client):
        content = _make_valid_xlsx()
        resp = client.post(
            "/api/bible/vintages",
            files={"file": ("hold.xlsx", content, "application/octet-stream")},
            params={"set_active": False},
        )
        assert resp.status_code == 200
        assert resp.json()["is_active"] is False

        # Bundled is still active
        assert client.get("/api/bible/active").json()["vintage_id"] == "bundled-q1-2026"

    def test_upload_rejects_non_xlsx(self, client):
        resp = client.post(
            "/api/bible/vintages",
            files={"file": ("doc.pdf", b"%PDF-1.4", "application/pdf")},
        )
        assert resp.status_code == 400
        assert "Unsupported file type" in resp.json()["detail"]

    def test_upload_rejects_garbage_xlsx(self, client):
        resp = client.post(
            "/api/bible/vintages",
            files={"file": ("fake.xlsx", b"not really xlsx", "application/octet-stream")},
        )
        assert resp.status_code == 400

    def test_upload_rejects_missing_cs_tab(self, client):
        # Build an xlsx with only a "Summary" tab
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        buf = io.BytesIO()
        wb.save(buf)

        resp = client.post(
            "/api/bible/vintages",
            files={"file": ("nocs.xlsx", buf.getvalue(), "application/octet-stream")},
        )
        assert resp.status_code == 400
        assert "CS" in resp.json()["detail"]


class TestSetActive:
    def test_switch_active_to_uploaded(self, client):
        content = _make_valid_xlsx()
        up = client.post(
            "/api/bible/vintages",
            files={"file": ("a.xlsx", content, "application/octet-stream")},
            params={"set_active": False},
        )
        vid = up.json()["vintage_id"]

        resp = client.post("/api/bible/active", json={"vintage_id": vid})
        assert resp.status_code == 200
        assert resp.json()["vintage_id"] == vid
        assert resp.json()["is_active"] is True

    def test_switch_to_missing_returns_404(self, client):
        resp = client.post("/api/bible/active", json={"vintage_id": "does-not-exist"})
        assert resp.status_code == 404


class TestDelete:
    def test_delete_non_active(self, client):
        content = _make_valid_xlsx()
        up = client.post(
            "/api/bible/vintages",
            files={"file": ("a.xlsx", content, "application/octet-stream")},
            params={"set_active": False},
        )
        vid = up.json()["vintage_id"]

        resp = client.delete(f"/api/bible/vintages/{vid}")
        assert resp.status_code == 200
        assert resp.json()["deleted"] is True

        # Gone from list
        ids = [v["vintage_id"] for v in client.get("/api/bible/vintages").json()]
        assert vid not in ids

    def test_cannot_delete_active(self, client):
        content = _make_valid_xlsx()
        up = client.post(
            "/api/bible/vintages",
            files={"file": ("a.xlsx", content, "application/octet-stream")},
            params={"set_active": True},
        )
        vid = up.json()["vintage_id"]

        resp = client.delete(f"/api/bible/vintages/{vid}")
        assert resp.status_code == 409
        assert "active" in resp.json()["detail"].lower()

    def test_delete_missing_404(self, client):
        resp = client.delete("/api/bible/vintages/does-not-exist")
        assert resp.status_code == 404
