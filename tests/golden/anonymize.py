"""Anonymize a 38DN pricing-model .xlsm into a golden-fixture safe copy.

Strategy
--------
1. Open each raw .xlsm under tests/golden/fixtures/raw/.
2. Read project names (PI row 4) and developer names (PI row 10) from
   columns H..N.
3. Build a mapping: original → generic ("Project A1", "Developer A", …).
4. Walk every cell in every worksheet. If a cell value EQUALS any
   mapped key (full-string match), replace with the mapped value.
5. Save to tests/golden/fixtures/<original_name>__anon.xlsm.

State / utility / financial values are NOT touched — those are industry-
typical, not deal-confidential, and meaningful for the audit pipeline.

Run from repo root:
    uv run python tests/golden/anonymize.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

RAW_DIR = Path("tests/golden/fixtures/raw")
OUT_DIR = Path("tests/golden/fixtures")


def collect_names(ws) -> tuple[list[str], list[str]]:
    """Read project names (row 4) and developer names (row 10) from cols H..N."""
    project_names: list[str] = []
    developer_names: list[str] = []
    for col_letter in "HIJKLMN":
        n = ws[f"{col_letter}4"].value
        d = ws[f"{col_letter}10"].value
        if n and isinstance(n, str) and n.strip():
            project_names.append(n)
        if d and isinstance(d, str) and d.strip():
            developer_names.append(d)
    return project_names, developer_names


def build_mapping(project_names: list[str], developer_names: list[str]) -> dict[str, str]:
    """Stable mapping: each original name → a generic placeholder.
    Preserves uniqueness — every original gets a unique replacement.
    """
    mapping: dict[str, str] = {}
    # Project names: "Project A1", "A2", …
    for i, name in enumerate(dict.fromkeys(project_names), start=1):
        mapping[name] = f"Project A{i}"
    # Developer names: "Developer A", "B", …
    for i, dev in enumerate(dict.fromkeys(developer_names)):
        mapping[dev] = f"Developer {chr(ord('A') + i)}"
    return mapping


def replace_all(wb, mapping: dict[str, str]) -> tuple[int, int]:
    """Walk every cell in every sheet and replace exact-match strings.
    Returns (cells_touched, sheets_scanned).
    """
    cells_touched = 0
    sheets_scanned = 0
    for ws in wb.worksheets:
        sheets_scanned += 1
        # iter_rows() over the worksheet's used range
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                # Full-string match (case-sensitive) against any mapped key.
                if v in mapping:
                    cell.value = mapping[v]
                    cells_touched += 1
    return cells_touched, sheets_scanned


def anonymize_one(src: Path) -> Path:
    print(f"\n=== {src.name} ===")
    # Load with data_only=False so we KEEP formulas intact (anonymize string
    # cells only). Cached values get refreshed by Excel on next open.
    wb = openpyxl.load_workbook(src, data_only=False, keep_vba=True)
    ws = wb["Project Inputs"]
    projects, developers = collect_names(ws)
    print(f"  Found {len(projects)} project names, {len(developers)} developer names")
    mapping = build_mapping(projects, developers)
    for orig, anon in mapping.items():
        orig_disp = orig.replace("\n", "\\n")
        print(f"    {orig_disp!r:40s} -> {anon!r}")
    touched, sheets = replace_all(wb, mapping)
    print(f"  Replaced {touched} cell value(s) across {sheets} sheet(s)")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = src.stem
    # Strip the deal-identifying parts of the filename too
    # e.g. "38DN-IL_US Solar_PricingModel_PV Only_2026.05.13_TEST" -> "38DN_fixture_us_solar_il_pv"
    out_name = f"{stem}__anon.xlsm"
    out_path = OUT_DIR / out_name
    wb.save(out_path)
    print(f"  Saved {out_path}  ({out_path.stat().st_size / 1024 / 1024:.1f} MB)")
    return out_path


def main() -> int:
    if not RAW_DIR.exists():
        print(f"ERROR: {RAW_DIR} does not exist", file=sys.stderr)
        return 1
    sources = sorted(RAW_DIR.glob("*.xlsm"))
    if not sources:
        print(f"ERROR: no .xlsm files in {RAW_DIR}", file=sys.stderr)
        return 1
    for src in sources:
        anonymize_one(src)
    print(f"\nDone. {len(sources)} fixture(s) written to {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
